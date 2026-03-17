"""
Seed Steph's Microsoft 365 account with Vantage demo data.

Seeds:
  - Outlook  (~70 emails across Inbox and Sent Items)
  - OneNote  (20 renewal call pages in "Member Notes" notebook)
  - OneDrive (Excel benchmarking files, workshop transcripts,
              case studies, membership data in Vantage/ folder)

Optionally wipes existing data before seeding with --wipe flag.

Usage:
    uv run seed.py                        # Auto-refresh token, then seed
    uv run seed.py --wipe                 # Wipe everything, then seed
    uv run seed.py --token-file path.txt  # Use a specific access token file
    MS365_ACCESS_TOKEN=xxx uv run seed.py # Use an explicit access token
"""

import argparse
import io
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from .m365_auth import resolve_access_token
except ImportError:
    from m365_auth import resolve_access_token

BASE = "https://graph.microsoft.com/v1.0"
STEPH = "steph@unifyailtd123.onmicrosoft.com"
ROB = "yasser@unifyailtd123.onmicrosoft.com"

TOKEN = None
H_JSON = {}
H_HTML = {}
H_AUTH = {}


def _init_headers(token):
    global TOKEN, H_JSON, H_HTML, H_AUTH
    TOKEN = token
    H_JSON = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    H_HTML = {"Authorization": f"Bearer {token}", "Content-Type": "text/html"}
    H_AUTH = {"Authorization": f"Bearer {token}"}


# =========================================================================
# CENTRAL ORG DATA — single source of truth for all 28 Carbon Club members
# =========================================================================

SESSION_DATES = [
    ("April 2024", "02/04/2024"),
    ("May 2024", "01/05/2024"),
    ("June 2024", "05/06/2024"),
    ("July 2024", "03/07/2024"),
    ("August 2024", "07/08/2024"),
    ("September 2024", "04/09/2024"),
    ("October 2024", "02/10/2024"),
    ("November 2024", "06/11/2024"),
    ("December 2024", "04/12/2024"),
    ("January 2025", "08/01/2025"),
    ("February 2025", "05/02/2025"),
    ("March 2025", "05/03/2025"),
]

ORGS = [
    {
        "name": "Greendale Homes",
        "city": "Nottingham",
        "homes": 12500,
        "contact": "Mark Jennings",
        "role": "Head of Sustainability",
        "email": "m.jennings@greendalehousing.org.uk",
        "since": 2019,
        "clubs": {"Carbon Club": 4200, "Performance Club": 5000},
        "renewal": "01/07/2025",
        "fin": {
            "cpu25": 4200,
            "cpu24": 3900,
            "reinvest": 8.5,
            "margin_sh": 28.1,
            "margin_all": 24.2,
            "gearing": 42.0,
            "ebitda": 210,
            "roce": 4.8,
        },
        "tsm": {
            "overall": 75,
            "repairs": 76,
            "time": 68,
            "safety": 80,
            "informed": 72,
            "fair": 79,
            "complaint": 38,
            "asb": 58,
            "neighbourhood": 63,
            "home": 68,
            "communal": 61,
        },
        "ops": {
            "void_days": 19,
            "rent_loss": 1.4,
            "epc_c": 78,
            "retrofit": 1800,
            "pv": 1200,
            "complaints": 42,
            "ev": 35,
            "damp": 12,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Northfield Housing Association",
        "city": "Birmingham",
        "homes": 8200,
        "contact": "Fiona Clarke",
        "role": "Director of Assets",
        "email": "f.clarke@northfieldhousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 4200, "Operations Club": 3900},
        "renewal": "01/08/2025",
        "fin": {
            "cpu25": 4800,
            "cpu24": 4500,
            "reinvest": 7.8,
            "margin_sh": 25.5,
            "margin_all": 22.0,
            "gearing": 45.0,
            "ebitda": 195,
            "roce": 4.2,
        },
        "tsm": {
            "overall": 74,
            "repairs": 75,
            "time": 66,
            "safety": 79,
            "informed": 70,
            "fair": 78,
            "complaint": 40,
            "asb": 57,
            "neighbourhood": 62,
            "home": 67,
            "communal": 59,
        },
        "ops": {
            "void_days": 16,
            "rent_loss": 1.1,
            "epc_c": 76,
            "retrofit": 1100,
            "pv": 850,
            "complaints": 38,
            "ev": 28,
            "damp": 14,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Beacon Dwellings",
        "city": "Coventry",
        "homes": 5800,
        "contact": "David Osei",
        "role": "Sustainability Manager",
        "email": "d.osei@beaconhousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 4200, "Finance Directors Club": 3600},
        "renewal": "01/09/2025",
        "fin": {
            "cpu25": 4600,
            "cpu24": 4300,
            "reinvest": 7.2,
            "margin_sh": 23.8,
            "margin_all": 20.5,
            "gearing": 48.0,
            "ebitda": 180,
            "roce": 3.8,
        },
        "tsm": {
            "overall": 71,
            "repairs": 72,
            "time": 64,
            "safety": 77,
            "informed": 68,
            "fair": 76,
            "complaint": 35,
            "asb": 55,
            "neighbourhood": 60,
            "home": 65,
            "communal": 57,
        },
        "ops": {
            "void_days": 21,
            "rent_loss": 1.6,
            "epc_c": 70,
            "retrofit": 680,
            "pv": 420,
            "complaints": 48,
            "ev": 15,
            "damp": 18,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Broadland Housing",
        "city": "Norwich",
        "homes": 4200,
        "contact": "Lena Morris",
        "role": "Director of Assets",
        "email": "l.morris@broadlandhousing.org.uk",
        "since": 2021,
        "clubs": {"Carbon Club": 4200, "Executive Club": 4300},
        "renewal": "01/04/2025",
        "fin": {
            "cpu25": 5800,
            "cpu24": 5500,
            "reinvest": 9.2,
            "margin_sh": 24.2,
            "margin_all": 20.8,
            "gearing": 52.0,
            "ebitda": 165,
            "roce": 3.5,
        },
        "tsm": {
            "overall": 72,
            "repairs": 74,
            "time": 65,
            "safety": 78,
            "informed": 69,
            "fair": 77,
            "complaint": 31,
            "asb": 54,
            "neighbourhood": 61,
            "home": 66,
            "communal": 58,
        },
        "ops": {
            "void_days": 24,
            "rent_loss": 1.9,
            "epc_c": 63,
            "retrofit": 280,
            "pv": 0,
            "complaints": 55,
            "ev": 5,
            "damp": 22,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 1, 1],
    },
    {
        "name": "Westmoor Housing Group",
        "city": "Sheffield",
        "homes": 9400,
        "contact": "Karen Blackwell",
        "role": "Head of Sustainability",
        "email": "k.blackwell@westmoorhousing.org.uk",
        "since": 2019,
        "clubs": {"Carbon Club": 4200},
        "renewal": "01/06/2025",
        "fin": {
            "cpu25": 5200,
            "cpu24": 4900,
            "reinvest": 6.8,
            "margin_sh": 22.5,
            "margin_all": 19.2,
            "gearing": 50.0,
            "ebitda": 155,
            "roce": 3.2,
        },
        "tsm": {
            "overall": 70,
            "repairs": 71,
            "time": 63,
            "safety": 76,
            "informed": 67,
            "fair": 76,
            "complaint": 33,
            "asb": 53,
            "neighbourhood": 59,
            "home": 64,
            "communal": 56,
        },
        "ops": {
            "void_days": 28,
            "rent_loss": 2.4,
            "epc_c": 71,
            "retrofit": 450,
            "pv": 0,
            "complaints": 62,
            "ev": 8,
            "damp": 20,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Riverside Community Housing",
        "city": "Manchester",
        "homes": 15300,
        "contact": "James Thornton",
        "role": "Head of Property Services",
        "email": "j.thornton@riversidehousing.org.uk",
        "since": 2018,
        "clubs": {
            "Carbon Club": 4200,
            "Operations Club": 3900,
            "Customer Experience Club": 4100,
        },
        "renewal": "01/05/2025",
        "fin": {
            "cpu25": 4500,
            "cpu24": 4200,
            "reinvest": 8.1,
            "margin_sh": 26.2,
            "margin_all": 22.8,
            "gearing": 44.0,
            "ebitda": 200,
            "roce": 4.5,
        },
        "tsm": {
            "overall": 73,
            "repairs": 74,
            "time": 67,
            "safety": 79,
            "informed": 71,
            "fair": 78,
            "complaint": 38,
            "asb": 57,
            "neighbourhood": 62,
            "home": 67,
            "communal": 60,
        },
        "ops": {
            "void_days": 22,
            "rent_loss": 1.7,
            "epc_c": 72,
            "retrofit": 950,
            "pv": 600,
            "complaints": 45,
            "ev": 30,
            "damp": 15,
        },
        "att": [1, 1, 1, 1, 1, 1, 0, 1, 1, 1, 1, 1],
    },
    {
        "name": "Oaktree Living",
        "city": "Leeds",
        "homes": 6100,
        "contact": "Priya Sharma",
        "role": "Sustainability Lead",
        "email": "p.sharma@oaktreehousing.org.uk",
        "since": 2021,
        "clubs": {"Carbon Club": 3800},
        "renewal": "01/06/2025",
        "fin": {
            "cpu25": 5600,
            "cpu24": 5300,
            "reinvest": 6.5,
            "margin_sh": 21.8,
            "margin_all": 18.5,
            "gearing": 53.0,
            "ebitda": 150,
            "roce": 3.0,
        },
        "tsm": {
            "overall": 69,
            "repairs": 70,
            "time": 62,
            "safety": 75,
            "informed": 66,
            "fair": 75,
            "complaint": 32,
            "asb": 52,
            "neighbourhood": 58,
            "home": 63,
            "communal": 55,
        },
        "ops": {
            "void_days": 25,
            "rent_loss": 2.1,
            "epc_c": 68,
            "retrofit": 400,
            "pv": 0,
            "complaints": 58,
            "ev": 5,
            "damp": 25,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Millbrook Homes",
        "city": "Bristol",
        "homes": 7600,
        "contact": "Tom Henderson",
        "role": "Director of Development",
        "email": "t.henderson@millbrookhousing.org.uk",
        "since": 2022,
        "clubs": {"Carbon Club": 4200, "Executive Club": 4300},
        "renewal": "01/08/2025",
        "fin": {
            "cpu25": 4400,
            "cpu24": 4100,
            "reinvest": 7.5,
            "margin_sh": 24.8,
            "margin_all": 21.5,
            "gearing": 43.0,
            "ebitda": 205,
            "roce": 4.5,
        },
        "tsm": {
            "overall": 74,
            "repairs": 75,
            "time": 67,
            "safety": 80,
            "informed": 71,
            "fair": 79,
            "complaint": 39,
            "asb": 57,
            "neighbourhood": 63,
            "home": 68,
            "communal": 60,
        },
        "ops": {
            "void_days": 17,
            "rent_loss": 1.2,
            "epc_c": 80,
            "retrofit": 620,
            "pv": 550,
            "complaints": 40,
            "ev": 25,
            "damp": 10,
        },
        "att": [1, 1, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Pennine Valleys Housing",
        "city": "Halifax",
        "homes": 4800,
        "contact": "Claire Whitfield",
        "role": "Asset Strategy Manager",
        "email": "c.whitfield@penninehousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 4200},
        "renewal": "01/07/2025",
        "fin": {
            "cpu25": 5200,
            "cpu24": 4900,
            "reinvest": 8.8,
            "margin_sh": 23.5,
            "margin_all": 20.2,
            "gearing": 47.0,
            "ebitda": 175,
            "roce": 3.6,
        },
        "tsm": {
            "overall": 71,
            "repairs": 72,
            "time": 64,
            "safety": 77,
            "informed": 68,
            "fair": 77,
            "complaint": 36,
            "asb": 55,
            "neighbourhood": 60,
            "home": 65,
            "communal": 57,
        },
        "ops": {
            "void_days": 24,
            "rent_loss": 1.8,
            "epc_c": 69,
            "retrofit": 450,
            "pv": 0,
            "complaints": 50,
            "ev": 8,
            "damp": 19,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Severn Vale Homes",
        "city": "Gloucester",
        "homes": 3900,
        "contact": "Andrew Marsh",
        "role": "Operations Director",
        "email": "a.marsh@severnvalehousing.org.uk",
        "since": 2019,
        "clubs": {"Carbon Club": 4200, "Customer Experience Club": 3400},
        "renewal": "01/05/2025",
        "fin": {
            "cpu25": 3800,
            "cpu24": 3500,
            "reinvest": 9.5,
            "margin_sh": 27.4,
            "margin_all": 24.5,
            "gearing": 38.0,
            "ebitda": 230,
            "roce": 5.2,
        },
        "tsm": {
            "overall": 78,
            "repairs": 79,
            "time": 72,
            "safety": 82,
            "informed": 75,
            "fair": 81,
            "complaint": 44,
            "asb": 61,
            "neighbourhood": 66,
            "home": 72,
            "communal": 64,
        },
        "ops": {
            "void_days": 15,
            "rent_loss": 0.9,
            "epc_c": 82,
            "retrofit": 700,
            "pv": 400,
            "complaints": 32,
            "ev": 18,
            "damp": 8,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Riverview Estates",
        "city": "Leicester",
        "homes": 11200,
        "contact": "Helen Foster",
        "role": "Head of Housing",
        "email": "h.foster@riverviewhousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 4800},
        "renewal": "01/04/2025",
        "fin": {
            "cpu25": 5500,
            "cpu24": 5200,
            "reinvest": 6.2,
            "margin_sh": 21.5,
            "margin_all": 18.0,
            "gearing": 55.0,
            "ebitda": 145,
            "roce": 2.8,
        },
        "tsm": {
            "overall": 66,
            "repairs": 67,
            "time": 59,
            "safety": 74,
            "informed": 63,
            "fair": 73,
            "complaint": 28,
            "asb": 48,
            "neighbourhood": 55,
            "home": 60,
            "communal": 52,
        },
        "ops": {
            "void_days": 26,
            "rent_loss": 2.2,
            "epc_c": 65,
            "retrofit": 180,
            "pv": 0,
            "complaints": 68,
            "ev": 3,
            "damp": 42,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0],
    },
    {
        "name": "Ashworth Housing Trust",
        "city": "Bolton",
        "homes": 5500,
        "contact": "Nadeem Hussain",
        "role": "Property Director",
        "email": "n.hussain@ashworthhousing.org.uk",
        "since": 2021,
        "clubs": {"Carbon Club": 4200, "Finance Directors Club": 3000},
        "renewal": "01/09/2025",
        "fin": {
            "cpu25": 5100,
            "cpu24": 4800,
            "reinvest": 7.4,
            "margin_sh": 24.0,
            "margin_all": 20.5,
            "gearing": 49.0,
            "ebitda": 170,
            "roce": 3.5,
        },
        "tsm": {
            "overall": 72,
            "repairs": 73,
            "time": 65,
            "safety": 78,
            "informed": 69,
            "fair": 77,
            "complaint": 37,
            "asb": 56,
            "neighbourhood": 61,
            "home": 66,
            "communal": 58,
        },
        "ops": {
            "void_days": 20,
            "rent_loss": 1.5,
            "epc_c": 74,
            "retrofit": 200,
            "pv": 0,
            "complaints": 46,
            "ev": 10,
            "damp": 16,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Meridian Housing Group",
        "city": "Southampton",
        "homes": 14700,
        "contact": "Sarah Linehan",
        "role": "Chief Operating Officer",
        "email": "s.linehan@meridianhousing.org.uk",
        "since": 2018,
        "clubs": {"Carbon Club": 4200, "Executive Club": 5600},
        "renewal": "01/06/2025",
        "fin": {
            "cpu25": 4300,
            "cpu24": 4000,
            "reinvest": 8.2,
            "margin_sh": 25.8,
            "margin_all": 22.5,
            "gearing": 41.0,
            "ebitda": 215,
            "roce": 4.6,
        },
        "tsm": {
            "overall": 75,
            "repairs": 76,
            "time": 69,
            "safety": 81,
            "informed": 73,
            "fair": 80,
            "complaint": 40,
            "asb": 59,
            "neighbourhood": 64,
            "home": 69,
            "communal": 62,
        },
        "ops": {
            "void_days": 18,
            "rent_loss": 1.3,
            "epc_c": 79,
            "retrofit": 820,
            "pv": 750,
            "complaints": 36,
            "ev": 40,
            "damp": 11,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 1, 1],
    },
    {
        "name": "Crestwood Homes",
        "city": "Wolverhampton",
        "homes": 6300,
        "contact": "Paul Gorman",
        "role": "Head of Asset Management",
        "email": "p.gorman@crestwoodhousing.org.uk",
        "since": 2022,
        "clubs": {"Carbon Club": 4200},
        "renewal": "01/10/2025",
        "fin": {
            "cpu25": 5200,
            "cpu24": 4900,
            "reinvest": 7.0,
            "margin_sh": 23.2,
            "margin_all": 19.8,
            "gearing": 49.0,
            "ebitda": 160,
            "roce": 3.3,
        },
        "tsm": {
            "overall": 70,
            "repairs": 71,
            "time": 63,
            "safety": 76,
            "informed": 67,
            "fair": 76,
            "complaint": 34,
            "asb": 53,
            "neighbourhood": 59,
            "home": 64,
            "communal": 56,
        },
        "ops": {
            "void_days": 22,
            "rent_loss": 1.7,
            "epc_c": 72,
            "retrofit": 380,
            "pv": 0,
            "complaints": 52,
            "ev": 6,
            "damp": 21,
        },
        "att": [1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0],
    },
    {
        "name": "Thornbury Housing Association",
        "city": "Bradford",
        "homes": 8900,
        "contact": "Rachel Iqbal",
        "role": "Sustainability Director",
        "email": "r.iqbal@thornburyhousing.org.uk",
        "since": 2019,
        "clubs": {"Carbon Club": 4200, "Operations Club": 3900},
        "renewal": "01/07/2025",
        "fin": {
            "cpu25": 5500,
            "cpu24": 5200,
            "reinvest": 6.8,
            "margin_sh": 22.0,
            "margin_all": 18.8,
            "gearing": 52.0,
            "ebitda": 148,
            "roce": 3.0,
        },
        "tsm": {
            "overall": 69,
            "repairs": 70,
            "time": 62,
            "safety": 75,
            "informed": 66,
            "fair": 75,
            "complaint": 33,
            "asb": 52,
            "neighbourhood": 58,
            "home": 63,
            "communal": 55,
        },
        "ops": {
            "void_days": 23,
            "rent_loss": 1.8,
            "epc_c": 70,
            "retrofit": 350,
            "pv": 0,
            "complaints": 56,
            "ev": 7,
            "damp": 24,
        },
        "att": [1, 1, 1, 1, 1, 1, 0, 1, 0, 0, 1, 0],
    },
    {
        "name": "Lakeside Living",
        "city": "Liverpool",
        "homes": 10100,
        "contact": "Chris Doyle",
        "role": "Head of Repairs & Maintenance",
        "email": "c.doyle@lakesidehousing.org.uk",
        "since": 2019,
        "clubs": {"Carbon Club": 4200, "Operations Club": 3200},
        "renewal": "01/06/2025",
        "fin": {
            "cpu25": 4800,
            "cpu24": 4500,
            "reinvest": 7.8,
            "margin_sh": 25.0,
            "margin_all": 21.8,
            "gearing": 43.0,
            "ebitda": 198,
            "roce": 4.3,
        },
        "tsm": {
            "overall": 74,
            "repairs": 75,
            "time": 68,
            "safety": 80,
            "informed": 71,
            "fair": 79,
            "complaint": 38,
            "asb": 57,
            "neighbourhood": 63,
            "home": 68,
            "communal": 60,
        },
        "ops": {
            "void_days": 19,
            "rent_loss": 1.4,
            "epc_c": 76,
            "retrofit": 550,
            "pv": 550,
            "complaints": 40,
            "ev": 22,
            "damp": 13,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Harrowfield Homes",
        "city": "Derby",
        "homes": 4400,
        "contact": "Joanna Briggs",
        "role": "Asset Manager",
        "email": "j.briggs@harrowfieldhousing.org.uk",
        "since": 2022,
        "clubs": {"Carbon Club": 4200},
        "renewal": "01/10/2025",
        "fin": {
            "cpu25": 5000,
            "cpu24": 4700,
            "reinvest": 7.2,
            "margin_sh": 24.2,
            "margin_all": 20.8,
            "gearing": 46.0,
            "ebitda": 178,
            "roce": 3.7,
        },
        "tsm": {
            "overall": 71,
            "repairs": 72,
            "time": 64,
            "safety": 77,
            "informed": 68,
            "fair": 77,
            "complaint": 35,
            "asb": 55,
            "neighbourhood": 60,
            "home": 65,
            "communal": 57,
        },
        "ops": {
            "void_days": 21,
            "rent_loss": 1.6,
            "epc_c": 73,
            "retrofit": 320,
            "pv": 0,
            "complaints": 50,
            "ev": 4,
            "damp": 17,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 1],
    },
    {
        "name": "Summit Housing Partnership",
        "city": "Huddersfield",
        "homes": 7200,
        "contact": "Darren Walsh",
        "role": "Director of Property",
        "email": "d.walsh@summithousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 4200, "Operations Club": 4000},
        "renewal": "01/08/2025",
        "fin": {
            "cpu25": 4700,
            "cpu24": 4400,
            "reinvest": 8.0,
            "margin_sh": 24.5,
            "margin_all": 21.2,
            "gearing": 45.0,
            "ebitda": 188,
            "roce": 4.0,
        },
        "tsm": {
            "overall": 73,
            "repairs": 74,
            "time": 66,
            "safety": 79,
            "informed": 70,
            "fair": 78,
            "complaint": 37,
            "asb": 56,
            "neighbourhood": 62,
            "home": 67,
            "communal": 59,
        },
        "ops": {
            "void_days": 20,
            "rent_loss": 1.5,
            "epc_c": 75,
            "retrofit": 600,
            "pv": 380,
            "complaints": 44,
            "ev": 15,
            "damp": 14,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Wychwood Housing Association",
        "city": "Oxford",
        "homes": 3200,
        "contact": "Mei-Lin Chen",
        "role": "Sustainability Lead",
        "email": "m.chen@wychwoodhousing.org.uk",
        "since": 2021,
        "clubs": {"Carbon Club": 3800},
        "renewal": "01/09/2025",
        "fin": {
            "cpu25": 3600,
            "cpu24": 3400,
            "reinvest": 9.0,
            "margin_sh": 26.5,
            "margin_all": 23.5,
            "gearing": 36.0,
            "ebitda": 240,
            "roce": 5.0,
        },
        "tsm": {
            "overall": 77,
            "repairs": 78,
            "time": 71,
            "safety": 82,
            "informed": 74,
            "fair": 81,
            "complaint": 41,
            "asb": 60,
            "neighbourhood": 65,
            "home": 71,
            "communal": 63,
        },
        "ops": {
            "void_days": 17,
            "rent_loss": 1.1,
            "epc_c": 84,
            "retrofit": 420,
            "pv": 300,
            "complaints": 30,
            "ev": 20,
            "damp": 7,
        },
        "att": [1, 1, 1, 1, 0, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Dales & Moorland Housing",
        "city": "Skipton",
        "homes": 2800,
        "contact": "Ian Calvert",
        "role": "Chief Executive",
        "email": "i.calvert@daleshousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 3400},
        "renewal": "01/07/2025",
        "fin": {
            "cpu25": 3400,
            "cpu24": 3200,
            "reinvest": 9.8,
            "margin_sh": 28.0,
            "margin_all": 25.0,
            "gearing": 34.0,
            "ebitda": 250,
            "roce": 5.5,
        },
        "tsm": {
            "overall": 79,
            "repairs": 80,
            "time": 73,
            "safety": 83,
            "informed": 76,
            "fair": 82,
            "complaint": 45,
            "asb": 62,
            "neighbourhood": 67,
            "home": 73,
            "communal": 65,
        },
        "ops": {
            "void_days": 16,
            "rent_loss": 1.0,
            "epc_c": 85,
            "retrofit": 380,
            "pv": 250,
            "complaints": 28,
            "ev": 12,
            "damp": 6,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 0, 1, 1, 1, 1],
    },
    {
        "name": "Ironbridge Homes",
        "city": "Telford",
        "homes": 5100,
        "contact": "Samira Begum",
        "role": "Head of Compliance",
        "email": "s.begum@ironbridgehousing.org.uk",
        "since": 2022,
        "clubs": {"Carbon Club": 4100},
        "renewal": "01/09/2025",
        "fin": {
            "cpu25": 5200,
            "cpu24": 4900,
            "reinvest": 7.1,
            "margin_sh": 23.0,
            "margin_all": 19.5,
            "gearing": 50.0,
            "ebitda": 158,
            "roce": 3.2,
        },
        "tsm": {
            "overall": 70,
            "repairs": 71,
            "time": 63,
            "safety": 76,
            "informed": 67,
            "fair": 76,
            "complaint": 34,
            "asb": 53,
            "neighbourhood": 59,
            "home": 64,
            "communal": 56,
        },
        "ops": {
            "void_days": 22,
            "rent_loss": 1.7,
            "epc_c": 71,
            "retrofit": 290,
            "pv": 0,
            "complaints": 54,
            "ev": 5,
            "damp": 20,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Thameside Housing Trust",
        "city": "London (SE)",
        "homes": 18500,
        "contact": "Oliver Grant",
        "role": "Group Director of Assets",
        "email": "o.grant@thamesidehousing.org.uk",
        "since": 2018,
        "clubs": {
            "Carbon Club": 4200,
            "Executive Club": 4300,
            "Finance Directors Club": 5000,
            "Operations Club": 4500,
            "Customer Experience Club": 4500,
        },
        "renewal": "01/04/2025",
        "fin": {
            "cpu25": 5800,
            "cpu24": 5500,
            "reinvest": 8.5,
            "margin_sh": 22.8,
            "margin_all": 19.5,
            "gearing": 55.0,
            "ebitda": 142,
            "roce": 2.9,
        },
        "tsm": {
            "overall": 71,
            "repairs": 72,
            "time": 64,
            "safety": 77,
            "informed": 68,
            "fair": 77,
            "complaint": 35,
            "asb": 54,
            "neighbourhood": 60,
            "home": 65,
            "communal": 57,
        },
        "ops": {
            "void_days": 20,
            "rent_loss": 1.5,
            "epc_c": 74,
            "retrofit": 1200,
            "pv": 1200,
            "complaints": 50,
            "ev": 45,
            "damp": 18,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Avondale Community Homes",
        "city": "Bath",
        "homes": 3600,
        "contact": "Emma Stubbs",
        "role": "Operations Manager",
        "email": "e.stubbs@avondalehousing.org.uk",
        "since": 2022,
        "clubs": {"Carbon Club": 3600},
        "renewal": "01/10/2025",
        "fin": {
            "cpu25": 3900,
            "cpu24": 3700,
            "reinvest": 8.8,
            "margin_sh": 26.0,
            "margin_all": 23.0,
            "gearing": 37.0,
            "ebitda": 235,
            "roce": 5.0,
        },
        "tsm": {
            "overall": 76,
            "repairs": 77,
            "time": 70,
            "safety": 81,
            "informed": 73,
            "fair": 80,
            "complaint": 40,
            "asb": 59,
            "neighbourhood": 64,
            "home": 70,
            "communal": 62,
        },
        "ops": {
            "void_days": 18,
            "rent_loss": 1.2,
            "epc_c": 80,
            "retrofit": 340,
            "pv": 200,
            "complaints": 34,
            "ev": 10,
            "damp": 9,
        },
        "att": [1, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Stonebridge Housing Group",
        "city": "Stoke-on-Trent",
        "homes": 9800,
        "contact": "Wayne Kirkpatrick",
        "role": "Director of Neighbourhoods",
        "email": "w.kirkpatrick@stonebridgehousing.org.uk",
        "since": 2019,
        "clubs": {"Carbon Club": 4200, "Customer Experience Club": 3800},
        "renewal": "01/07/2025",
        "fin": {
            "cpu25": 5000,
            "cpu24": 4700,
            "reinvest": 7.6,
            "margin_sh": 23.8,
            "margin_all": 20.5,
            "gearing": 47.0,
            "ebitda": 182,
            "roce": 3.8,
        },
        "tsm": {
            "overall": 72,
            "repairs": 73,
            "time": 65,
            "safety": 78,
            "informed": 69,
            "fair": 78,
            "complaint": 36,
            "asb": 56,
            "neighbourhood": 61,
            "home": 66,
            "communal": 58,
        },
        "ops": {
            "void_days": 21,
            "rent_loss": 1.6,
            "epc_c": 73,
            "retrofit": 680,
            "pv": 900,
            "complaints": 46,
            "ev": 20,
            "damp": 16,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Foxley Homes",
        "city": "Swindon",
        "homes": 4500,
        "contact": "Danielle Webb",
        "role": "Head of Sustainability",
        "email": "d.webb@foxleyhousing.org.uk",
        "since": 2022,
        "clubs": {"Carbon Club": 4200},
        "renewal": "01/10/2025",
        "fin": {
            "cpu25": 4300,
            "cpu24": 4100,
            "reinvest": 7.8,
            "margin_sh": 24.5,
            "margin_all": 21.2,
            "gearing": 44.0,
            "ebitda": 192,
            "roce": 4.1,
        },
        "tsm": {
            "overall": 73,
            "repairs": 74,
            "time": 66,
            "safety": 79,
            "informed": 70,
            "fair": 78,
            "complaint": 37,
            "asb": 56,
            "neighbourhood": 62,
            "home": 67,
            "communal": 59,
        },
        "ops": {
            "void_days": 19,
            "rent_loss": 1.4,
            "epc_c": 76,
            "retrofit": 350,
            "pv": 180,
            "complaints": 42,
            "ev": 8,
            "damp": 14,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Eastgate Housing Association",
        "city": "Ipswich",
        "homes": 3400,
        "contact": "Martin Osborne",
        "role": "Asset Management Lead",
        "email": "m.osborne@eastgatehousing.org.uk",
        "since": 2023,
        "clubs": {"Carbon Club": 3800},
        "renewal": "01/11/2025",
        "fin": {
            "cpu25": 4000,
            "cpu24": 3800,
            "reinvest": 8.2,
            "margin_sh": 25.2,
            "margin_all": 22.0,
            "gearing": 40.0,
            "ebitda": 218,
            "roce": 4.5,
        },
        "tsm": {
            "overall": 74,
            "repairs": 75,
            "time": 67,
            "safety": 80,
            "informed": 71,
            "fair": 79,
            "complaint": 38,
            "asb": 57,
            "neighbourhood": 63,
            "home": 68,
            "communal": 60,
        },
        "ops": {
            "void_days": 18,
            "rent_loss": 1.3,
            "epc_c": 78,
            "retrofit": 420,
            "pv": 310,
            "complaints": 38,
            "ev": 12,
            "damp": 10,
        },
        "att": [1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Chiltern Edge Housing",
        "city": "High Wycombe",
        "homes": 6700,
        "contact": "Abigail Reeves",
        "role": "Property Services Director",
        "email": "a.reeves@chilternhousing.org.uk",
        "since": 2020,
        "clubs": {"Carbon Club": 4200, "Finance Directors Club": 3400},
        "renewal": "01/08/2025",
        "fin": {
            "cpu25": 4700,
            "cpu24": 4400,
            "reinvest": 7.5,
            "margin_sh": 24.0,
            "margin_all": 20.8,
            "gearing": 46.0,
            "ebitda": 185,
            "roce": 3.9,
        },
        "tsm": {
            "overall": 72,
            "repairs": 73,
            "time": 65,
            "safety": 78,
            "informed": 69,
            "fair": 78,
            "complaint": 36,
            "asb": 55,
            "neighbourhood": 61,
            "home": 66,
            "communal": 58,
        },
        "ops": {
            "void_days": 20,
            "rent_loss": 1.5,
            "epc_c": 74,
            "retrofit": 520,
            "pv": 420,
            "complaints": 44,
            "ev": 18,
            "damp": 15,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    },
    {
        "name": "Maplewood Living",
        "city": "Cambridge",
        "homes": 5900,
        "contact": "George Kaplan",
        "role": "Head of Development",
        "email": "g.kaplan@maplewoodhousing.org.uk",
        "since": 2021,
        "clubs": {"Carbon Club": 4200, "Executive Club": 4300},
        "renewal": "01/09/2025",
        "fin": {
            "cpu25": 4300,
            "cpu24": 4000,
            "reinvest": 8.0,
            "margin_sh": 25.5,
            "margin_all": 22.2,
            "gearing": 42.0,
            "ebitda": 208,
            "roce": 4.4,
        },
        "tsm": {
            "overall": 74,
            "repairs": 75,
            "time": 68,
            "safety": 80,
            "informed": 72,
            "fair": 79,
            "complaint": 39,
            "asb": 58,
            "neighbourhood": 63,
            "home": 68,
            "communal": 61,
        },
        "ops": {
            "void_days": 19,
            "rent_loss": 1.3,
            "epc_c": 77,
            "retrofit": 580,
            "pv": 560,
            "complaints": 38,
            "ev": 22,
            "damp": 11,
        },
        "att": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 1],
    },
]


# =========================================================================
# WIPE
# =========================================================================


def wipe_emails():
    print("Wiping emails...")
    for folder in ["Inbox", "SentItems"]:
        while True:
            r = requests.get(
                f"{BASE}/me/mailFolders/{folder}/messages",
                headers=H_AUTH,
                params={"$top": 50, "$select": "id"},
            )
            msgs = r.json().get("value", [])
            if not msgs:
                break
            for m in msgs:
                requests.delete(f"{BASE}/me/messages/{m['id']}", headers=H_AUTH)
            print(f"  Deleted {len(msgs)} from {folder}")
            time.sleep(0.5)
    print("  Done.")


def wipe_onenote():
    print("Wiping OneNote pages from 'Member Notes'...")
    r = requests.get(f"{BASE}/me/onenote/notebooks", headers=H_AUTH)
    for nb in r.json().get("value", []):
        if nb["displayName"] == "Member Notes":
            r2 = requests.get(
                f"{BASE}/me/onenote/notebooks/{nb['id']}/sections",
                headers=H_AUTH,
            )
            for sec in r2.json().get("value", []):
                r3 = requests.get(
                    f"{BASE}/me/onenote/sections/{sec['id']}/pages",
                    headers=H_AUTH,
                    params={"$top": 100},
                )
                for page in r3.json().get("value", []):
                    requests.delete(
                        f"{BASE}/me/onenote/pages/{page['id']}",
                        headers=H_AUTH,
                    )
                    time.sleep(0.3)
                print(f"  Deleted pages from section: {sec['displayName']}")
    print("  Done.")


def wipe_onedrive():
    print("Wiping OneDrive Vantage/ folder...")
    r = requests.get(
        f"{BASE}/me/drive/root:/Vantage",
        headers=H_AUTH,
    )
    if r.status_code == 200:
        item_id = r.json()["id"]
        requests.delete(f"{BASE}/me/drive/items/{item_id}", headers=H_AUTH)
        print("  Deleted Vantage/ folder.")
        time.sleep(2)
    elif r.status_code == 404:
        print("  Vantage/ folder not found — nothing to delete.")
    else:
        print(f"  Warning: could not check Vantage/ folder ({r.status_code})")
    print("  Done.")


# =========================================================================
# EMAIL DATA
# =========================================================================


def _e(subject, date, frm, to, body, direction):
    return {
        "subject": subject,
        "date": date,
        "from": {"name": frm[0], "email": frm[1]},
        "to": {"name": to[0], "email": to[1]},
        "body": body,
        "direction": direction,
    }


S = ("Steph Hosny", STEPH)
R = ("Rob Bryan", ROB)
LM = ("Lena Morris", "l.morris@broadlandhousing.org.uk")
MJ = ("Mark Jennings", "m.jennings@greendalehousing.org.uk")
FC = ("Fiona Clarke", "f.clarke@northfieldhousing.org.uk")
KB = ("Karen Blackwell", "k.blackwell@westmoorhousing.org.uk")
DO = ("David Osei", "d.osei@beaconhousing.org.uk")
JT = ("James Thornton", "j.thornton@riversidehousing.org.uk")
PS = ("Priya Sharma", "p.sharma@oaktreehousing.org.uk")
TH = ("Tom Henderson", "t.henderson@millbrookhousing.org.uk")
AM = ("Andrew Marsh", "a.marsh@severnvalehousing.org.uk")
HF = ("Helen Foster", "h.foster@riverviewhousing.org.uk")
CW = ("Claire Whitfield", "c.whitfield@penninehousing.org.uk")
OG = ("Oliver Grant", "o.grant@thamesidehousing.org.uk")
CD = ("Chris Doyle", "c.doyle@lakesidehousing.org.uk")
SL = ("Sarah Linehan", "s.linehan@meridianhousing.org.uk")
NH = ("Nadeem Hussain", "n.hussain@ashworthhousing.org.uk")
DW = ("Darren Walsh", "d.walsh@summithousing.org.uk")
SB = ("Samira Begum", "s.begum@ironbridgehousing.org.uk")
JB = ("Joanna Briggs", "j.briggs@harrowfieldhousing.org.uk")
DnW = ("Danielle Webb", "d.webb@foxleyhousing.org.uk")
PG = ("Paul Gorman", "p.gorman@crestwoodhousing.org.uk")
RI = ("Rachel Iqbal", "r.iqbal@thornburyhousing.org.uk")
MC = ("Mei-Lin Chen", "m.chen@wychwoodhousing.org.uk")
IC = ("Ian Calvert", "i.calvert@daleshousing.org.uk")
ES = ("Emma Stubbs", "e.stubbs@avondalehousing.org.uk")
WK = ("Wayne Kirkpatrick", "w.kirkpatrick@stonebridgehousing.org.uk")
MO = ("Martin Osborne", "m.osborne@eastgatehousing.org.uk")
AR = ("Abigail Reeves", "a.reeves@chilternhousing.org.uk")
GK = ("George Kaplan", "g.kaplan@maplewoodhousing.org.uk")

EMAILS = [
    # === BROADLAND HOUSING — full thread ===
    _e(
        "Solar PV case studies - Greendale & Northfield",
        "2024-10-15T14:30:00Z",
        S,
        LM,
        "Hi Lena,\n\nFollowing our one-to-one last month, please find attached the case studies from Greendale Homes and Northfield HA on their solar PV programmes.\n\nKey highlights:\n- Greendale: 1,200 installations at \u00a34,200/unit via Solarfix framework, 89% tenant satisfaction\n- Northfield: 850 installations with battery storage at \u00a36,200/unit, tenants saving additional \u00a3150/yr vs PV-only\n\nBoth are happy to host peer visits if you'd like to see the programmes in action. Let me know and I can make introductions.\n\nBest,\nSteph",
        "sent",
    ),
    _e(
        "RE: Solar PV case studies - Greendale & Northfield",
        "2024-10-16T09:15:00Z",
        LM,
        S,
        "Hi Steph,\n\nThese are really helpful, thank you. The Greendale numbers in particular are very encouraging \u2014 their costs are lower than I expected.\n\nI'd definitely be interested in a peer visit to Greendale. Could you set that up for me? We're looking at a programme of about 500 properties initially and their experience with the Solarfix framework would be really valuable.\n\nI've shared the case studies with our Director of Finance and he's keen to understand the SHDF funding angle better. Could you send over the latest SHDF guidance when you get a chance?\n\nThanks,\nLena",
        "inbox",
    ),
    _e(
        "Void benchmarking data pack",
        "2024-11-22T11:00:00Z",
        S,
        LM,
        "Hi Lena,\n\nAs promised, please find attached the void turnaround benchmarking data compiled from Carbon Club members.\n\nThe pack includes:\n- Target vs actual turnaround times for 15 organisations\n- Breakdown by lettable standard (Basic, Standard, Enhanced)\n- Rent loss calculations per void\n- Best practice examples from top performers\n\nSevern Vale Homes are the standout at 15 days average \u2014 they attribute this to their in-house DLO. Might be worth a conversation with Andrew Marsh there if you're looking to improve your turnaround times.\n\nBest,\nSteph",
        "sent",
    ),
    _e(
        "SHDF Wave 2 funding guidance document",
        "2025-01-08T10:20:00Z",
        S,
        LM,
        "Hi Lena,\n\nAs requested, here is the DESNZ guidance on the SHDF Wave 2 application process and eligibility criteria.\n\nKey points for Broadland:\n- Application window for Wave 2.2 opens 1 April 2025, closes 30 June 2025\n- PV on its own IS eligible as long as properties are below EPC C\n- Minimum 33% match funding from the housing association\n- You'll need property-level EPC data and contractor quotes in the application\n\nClaire Whitfield at Pennine Valleys went through the Wave 2.1 process and said the application took about 6 weeks \u2014 I'd recommend starting to pull your data together now.\n\nBest,\nSteph",
        "sent",
    ),
    _e(
        "RE: SHDF Wave 2 funding guidance document",
        "2025-01-09T11:45:00Z",
        LM,
        S,
        "Thanks Steph, this is really useful. I've shared it with our finance team and we're starting to pull the property-level data together.\n\nOne question \u2014 do we need actual contractor quotes at application stage, or just indicative pricing? We haven't gone to market yet.\n\nAlso, could you set up that peer visit to Greendale? I'd like to go before the application window opens if possible.\n\nLena",
        "inbox",
    ),
    _e(
        "RE: SHDF Wave 2 funding guidance document",
        "2025-01-10T09:00:00Z",
        S,
        LM,
        "Hi Lena,\n\nIndicative pricing should be fine at application stage \u2014 you just need to demonstrate you've got a realistic cost estimate. Claire at Pennine Valleys can confirm but I believe they used framework prices rather than actual quotes.\n\nI'll set up the Greendale visit \u2014 will email Mark today.\n\nSteph",
        "sent",
    ),
    # === GREENDALE HOMES ===
    _e(
        "Carbon Club spotlight - February session",
        "2025-01-20T16:45:00Z",
        MJ,
        S,
        "Hi Steph,\n\nJust confirming I'm happy to do the PV spotlight at the February Carbon Club session. I'll prepare a short presentation covering:\n\n- Programme progress (1,200 installs completed)\n- Costs and procurement approach\n- Tenant satisfaction results\n- Plans for the remaining 600 properties\n\nShould be about 15 minutes. Let me know if you want me to cover anything else specifically.\n\nMark",
        "inbox",
    ),
    _e(
        "RE: Carbon Club spotlight - February session",
        "2025-01-21T09:30:00Z",
        S,
        MJ,
        "Perfect, thanks Mark. That all sounds great.\n\nOne thing that might be useful to mention is the grid connection process with Western Power Distribution \u2014 a few members have been asking about potential bottlenecks with the DNO. Your experience with batch applications would be really helpful for them to hear.\n\nAlso, Lena Morris from Broadland is keen to visit your programme. Would you be open to hosting a peer visit? She's looking at starting a PV programme of about 500 properties.\n\nSteph",
        "sent",
    ),
    _e(
        "RE: Carbon Club spotlight - February session",
        "2025-01-21T14:20:00Z",
        MJ,
        S,
        "Good shout on the DNO bit \u2014 I'll definitely include that. Happy to host Lena. Get her to drop me a line and we'll sort a date. February or early March would work well.\n\nMark",
        "inbox",
    ),
    _e(
        "Greendale PV procurement framework details",
        "2025-03-06T14:00:00Z",
        MJ,
        S,
        "Hi Steph,\n\nAs promised in yesterday's session, here are the details of the Solarfix procurement framework:\n\n- Framework operator: Solarfix Ltd (accessed via Fusion21)\n- Contract value: \u00a35.4m over 3 years\n- Procurement timeline: 12 weeks from going to market to appointment\n- Key benefits: pre-vetted contractors, competitive pricing, simplified process\n- Contact at Solarfix: James Wright, j.wright@solarfix.co.uk\n\nHappy for you to share this with anyone who's interested. And yes, Lena from Broadland is very welcome to visit \u2014 just get her to drop me a line and we'll sort a date.\n\nMark",
        "inbox",
    ),
    _e(
        "Greendale 12-month PV performance report",
        "2025-02-18T10:30:00Z",
        MJ,
        S,
        "Hi Steph,\n\nAs mentioned at the session, here's our 12-month PV performance report that went to the board in January. Key findings:\n\n- Average generation: 3,800 kWh per property per year (vs predicted 4,000)\n- Average tenant saving: \u00a3180/year on electricity\n- 8% tenant refusal rate (mostly elderly residents concerned about disruption)\n- 96% system reliability (only 48 faults across 1,200 installations)\n\nFeel free to share the format with other members \u2014 I think having a standardised performance report template could be useful for the group.\n\nMark",
        "inbox",
    ),
    # === NORTHFIELD HA ===
    _e(
        "Battery storage cost-benefit analysis",
        "2025-03-07T11:30:00Z",
        FC,
        S,
        "Hi Steph,\n\nAs discussed at Wednesday's session, here's our battery storage cost-benefit analysis.\n\nSummary of key findings:\n- Additional cost per property for battery: \u00a31,800\n- Additional tenant saving per year: \u00a3150 vs PV-only\n- Battery payback period: 10-12 years\n- System used: GivEnergy 5.2kWh, 12-year warranty\n- Failure rate: <1% (6 out of 850)\n\nThe full spreadsheet model is attached. Feel free to share with the group.\n\nFiona",
        "inbox",
    ),
    _e(
        "Heat pump pilot update",
        "2024-11-28T15:00:00Z",
        FC,
        S,
        "Hi Steph,\n\nQuick update on our heat pump pilot \u2014 we've now completed all 50 installations.\n\nResults are mixed honestly:\n- Energy cost savings are good for most tenants\n- But 15% have complained about noise (mainly semi-detached where the unit is near a neighbour's window)\n- Installation costs came in at \u00a39,500 vs our \u00a38,000 estimate\n- Overall satisfaction at 76%\n\nWe're holding off on expanding until we've resolved the noise issue. Might need to look at different unit placements or acoustic enclosures.\n\nFiona",
        "inbox",
    ),
    # === WESTMOOR HOUSING GROUP ===
    _e(
        "RE: Void standards - urgent",
        "2024-09-15T08:45:00Z",
        KB,
        S,
        "Hi Steph,\n\nThanks for the call last week. The board meeting was tough \u2014 they're really pushing us on void turnaround times. We're at 28 days average which is well above what other HAs seem to be achieving.\n\nCould you share any benchmarking data you have on void standards? Specifically I'd like to know:\n- What targets other Carbon Club members are working to\n- What lettable standards they use (Basic/Standard/Enhanced)\n- Any practical tips that have made a real difference\n\nI also posted on the portal about this so hopefully we'll get some responses there too.\n\nKaren",
        "inbox",
    ),
    _e(
        "RE: Void standards - urgent",
        "2024-09-16T10:20:00Z",
        S,
        KB,
        "Hi Karen,\n\nSorry to hear the board meeting was difficult. I'll pull together a benchmarking data pack for you \u2014 should have it ready by end of next week.\n\nIn the meantime, a few quick pointers from what I'm hearing across the club:\n- Severn Vale Homes are the best performer at 15 days avg \u2014 they run an in-house DLO which makes a big difference\n- Northfield are at 16 days with pre-void inspections 4 weeks before tenancy end \u2014 that seems to be a key lever\n- Millbrook integrated their void and lettings teams into one \u2014 eliminated a 3-4 day handover gap\n\nI'll include all this detail in the data pack.\n\nSteph",
        "sent",
    ),
    _e(
        "RE: Void standards - urgent",
        "2024-09-17T11:30:00Z",
        KB,
        S,
        "That's really helpful Steph, thank you. The Severn Vale example is interesting \u2014 we've been thinking about going in-house for a while but the setup costs have put us off. Would be good to understand Andrew's experience.\n\nLooking forward to the data pack.\n\nKaren",
        "inbox",
    ),
    _e(
        "Insulation programme tenant engagement plan",
        "2025-01-22T14:15:00Z",
        KB,
        S,
        "Hi Steph,\n\nWe're finalising our tenant engagement plan for the insulation programme starting in April. Karen mentioned at the last one-to-one that she'd value any examples from other members.\n\nWe're planning face-to-face visits for elderly tenants (which seems to be best practice from what we've heard) and letters for everyone else. James at Riverside offered to share his templates \u2014 has he sent those through yet?\n\nKaren",
        "inbox",
    ),
    _e(
        "RE: Insulation programme tenant engagement plan",
        "2025-01-23T09:00:00Z",
        S,
        KB,
        "Hi Karen,\n\nYes, James sent the templates through last week \u2014 I'll forward them to you now. They're really well structured, especially the part about being specific with savings figures.\n\nYour plan sounds solid. The face-to-face approach for elderly tenants is definitely the way to go \u2014 several members have seen refusal rates drop from 15% to under 5% with home visits.\n\nSteph",
        "sent",
    ),
    # === SEVERN VALE HOMES ===
    _e(
        "Whole-house retrofit spotlight offer",
        "2025-03-06T16:15:00Z",
        AM,
        S,
        "Hi Steph,\n\nFollowing yesterday's session, I'd be happy to do a spotlight on our whole-house retrofit programme at a future session. We've done 200 properties with insulation, windows, and heat pumps \u2014 EPC E to B in most cases.\n\nI could also cover the in-house DLO approach since a few people were asking about it, particularly Priya from Oaktree who's thinking about going in-house for their void works.\n\nLet me know what month works best.\n\nAndrew",
        "inbox",
    ),
    _e(
        "DLO setup costs breakdown",
        "2024-10-22T10:00:00Z",
        AM,
        S,
        "Hi Steph,\n\nFollowing the interest from a few members about our in-house DLO, I've put together a breakdown of the setup costs:\n\n- 8 operatives recruited (painters x3, plumber x1, electrician x1, general x2, supervisor x1)\n- Vehicles: \u00a3145k (8 vans, bought outright)\n- Tools and equipment: \u00a328k\n- Recruitment costs: \u00a312k\n- Total setup: approximately \u00a3185k\n\nWe broke even within 14 months through reduced contractor costs and dramatically lower rent loss. Our void turnaround went from 24 days to 15 days.\n\nHappy to share more detail with anyone who's interested.\n\nAndrew",
        "inbox",
    ),
    # === RIVERVIEW ESTATES ===
    _e(
        "Checking in - Carbon Club attendance",
        "2025-02-10T09:00:00Z",
        S,
        HF,
        "Hi Helen,\n\nI just wanted to check in \u2014 I noticed Riverview Estates hasn't attended the last couple of Carbon Club sessions. I hope everything is okay.\n\nWe've had some really useful discussions recently on solar PV programmes, EPC data quality, and SHDF funding \u2014 I can send you the notes if that would be helpful.\n\nYour renewal is coming up in April so it would be good to catch up before then. Are you free for a one-to-one in the next couple of weeks?\n\nBest,\nSteph",
        "sent",
    ),
    _e(
        "RE: Checking in - Carbon Club attendance",
        "2025-02-12T14:30:00Z",
        HF,
        S,
        "Hi Steph,\n\nThanks for reaching out. To be honest, things have been very stretched here. We've had a significant increase in damp and mould cases \u2014 we're at 42 per 1,000 homes now \u2014 and that's been taking up most of the team's time.\n\nI know we should be attending the sessions but it's been hard to prioritise with everything going on. The notes would be really helpful though, yes please.\n\nI can do a one-to-one on the 25th if that works?\n\nHelen",
        "inbox",
    ),
    _e(
        "RE: Checking in - Carbon Club attendance",
        "2025-02-13T08:30:00Z",
        S,
        HF,
        "Hi Helen,\n\nCompletely understand \u2014 damp and mould is a huge pressure for everyone right now. The 25th works for me, let's do 10am.\n\nIn the meantime, Chris Doyle at Lakeside Living has done some really innovative work with environmental monitoring sensors for damp prevention. Might be worth a conversation? I can connect you if useful.\n\nI'll send the session notes over today.\n\nSteph",
        "sent",
    ),
    # === OAKTREE LIVING ===
    _e(
        "In-house void team - business case help",
        "2025-02-20T11:00:00Z",
        PS,
        S,
        "Hi Steph,\n\nFollowing the discussion at the last Carbon Club about void turnaround times, I'm now seriously looking at setting up an in-house void team. Andrew at Severn Vale mentioned he broke even within 14 months.\n\nCould you connect me with Andrew? I'd like to understand the setup costs, team structure, and how they manage the recruitment. Our current contractor availability is terrible \u2014 25 days average turnaround and it's costing us over \u00a3400k a year in lost rent.\n\nPriya",
        "inbox",
    ),
    _e(
        "RE: In-house void team - business case help",
        "2025-02-21T09:15:00Z",
        S,
        PS,
        "Hi Priya,\n\nAbsolutely \u2014 I've CC'd Andrew Marsh on this email. Andrew, would you mind sharing the DLO setup costs breakdown you sent me last month? I think it would be really helpful for Priya.\n\nPriya, Andrew's team went from 24 days average to 15 days after setting up the DLO. The setup cost was about \u00a3185k but they broke even within 14 months through reduced contractor costs and lower rent loss.\n\nSteph",
        "sent",
    ),
    # === BEACON DWELLINGS ===
    _e(
        "EPC data quality - resurvey programme",
        "2024-12-05T15:30:00Z",
        DO,
        S,
        "Hi Steph,\n\nWe've been digging into our EPC data following the discussions at Carbon Club and it's worse than I thought \u2014 about 25% of our solid wall stock has inaccurate EPCs.\n\nMark at Greendale mentioned they're working with Elmhurst Energy on a resurvey. Could you connect us? We're budgeting for a full resurvey in 2025/26 and it would be good to compare approaches.\n\nAlso, our PV programme is at 340 installs now. Hoping to get to 500 by end of the financial year.\n\nDavid",
        "inbox",
    ),
    _e(
        "RE: EPC data quality - resurvey programme",
        "2024-12-06T09:45:00Z",
        S,
        DO,
        "Hi David,\n\nThanks for the update. 25% inaccuracy is significant \u2014 you're not alone though, we're hearing similar numbers from several members.\n\nI've copied Mark Jennings (Greendale) on this email \u2014 Mark, would you mind sharing your Elmhurst Energy contact with David?\n\nGood progress on the PV programme. 500 by year end sounds achievable.\n\nSteph",
        "sent",
    ),
    _e(
        "Beacon PV programme Q3 update",
        "2025-02-05T16:00:00Z",
        DO,
        S,
        "Hi Steph,\n\nQuick update \u2014 we hit 420 PV installs by end of January. On track for 500 by end of March. Costs holding steady at around \u00a34,400/unit through the Fusion21 framework.\n\nWe've also started the EPC resurvey with a firm called EPC Solutions (not Elmhurst in the end \u2014 they were cheaper). First batch of 300 properties done, confirming about 22% inaccuracy rate.\n\nDavid",
        "inbox",
    ),
    # === PENNINE VALLEYS HOUSING ===
    _e(
        "SHDF Wave 2.1 success - thank you",
        "2024-11-05T10:00:00Z",
        CW,
        S,
        "Hi Steph,\n\nJust wanted to let you know we were successful with our SHDF Wave 2.1 bid \u2014 \u00a31.2 million for 450 measures across 280 properties!\n\nThe guidance notes you shared and the examples from other Carbon Club members were really helpful in putting the application together. Took about 6 weeks but definitely worth the effort.\n\nHappy to share our experience with anyone else in the club who's thinking of applying for Wave 2.2.\n\nClaire",
        "inbox",
    ),
    _e(
        "SHDF Wave 2.1 - delivery progress",
        "2025-02-15T11:00:00Z",
        CW,
        S,
        "Hi Steph,\n\nThought you'd like an update on our SHDF programme. We've completed 180 of the 450 planned measures so far \u2014 mainly loft insulation and cavity wall. The remaining 270 are scheduled for Q1-Q2 2025/26.\n\nThe trickiest part has been the tenant engagement for external wall insulation \u2014 scaffolding access is always a challenge. We're using face-to-face visits which has helped but it's time-consuming.\n\nClaire",
        "inbox",
    ),
    # === THAMESIDE HOUSING TRUST ===
    _e(
        "EPC regulatory reporting concerns",
        "2024-11-18T13:00:00Z",
        OG,
        S,
        "Hi Steph,\n\nI've been thinking more about the EPC data quality issue we discussed at the last session. I'm worried about the regulatory implications.\n\nIf we're reporting 75% of stock at EPC C or above but the reality is 65%, that's a governance issue. The consumer standards require accurate data and I don't think we can defend inaccurate EPC reporting.\n\nAre any other members looking at this from a compliance angle? Would be useful to know how others are approaching the risk.\n\nOliver",
        "inbox",
    ),
    _e(
        "Heat pump programme - board presentation",
        "2025-01-28T09:30:00Z",
        OG,
        S,
        "Hi Steph,\n\nWe've got a board presentation next month on our heat pump programme. Our satisfaction is still at 78% which the board aren't happy about.\n\nDo you have any examples of how other HAs have improved their heat pump satisfaction scores? I know the sector is struggling with this generally but any evidence of what works would be helpful.\n\nOliver",
        "inbox",
    ),
    _e(
        "RE: Heat pump programme - board presentation",
        "2025-01-29T10:00:00Z",
        S,
        OG,
        "Hi Oliver,\n\nThe 78% is actually in line with what we're seeing across the club \u2014 heat pumps are a harder sell than PV because of the behavioural change required.\n\nKey things that seem to help:\n- Education sessions before installation (Severn Vale do a 1-hour home visit)\n- Setting expectations about how heat pumps work differently to gas boilers\n- Follow-up visits 2 weeks post-install to address any issues early\n\nI'll pull together a summary of heat pump satisfaction data across all members \u2014 it might help to show the board that 78% is actually mid-table, not bottom.\n\nSteph",
        "sent",
    ),
    # === LAKESIDE LIVING ===
    _e(
        "Environmental monitoring sensors - results",
        "2025-01-15T14:30:00Z",
        CD,
        S,
        "Hi Steph,\n\nYou asked me to share an update on our environmental monitoring sensors programme. We've now got 500 sensors deployed across our highest-risk properties.\n\nResults so far:\n- 23 properties flagged for high humidity before any mould developed\n- Cost: about \u00a380 per sensor including installation\n- Reduced reactive damp & mould callouts by 30% in monitored properties\n- Tenants like the proactive approach \u2014 satisfaction up 12 points\n\nHappy to present on this at a future session if useful.\n\nChris",
        "inbox",
    ),
    _e(
        "Sensor supplier details",
        "2025-02-25T10:00:00Z",
        CD,
        S,
        "Hi Steph,\n\nA few members asked about our sensor supplier after I mentioned it on the portal. We use Switchee \u2014 they offer a housing association package at \u00a380/unit installed with a 5-year data contract.\n\nContact: Sarah Mitchell, s.mitchell@switchee.co, 020 7946 0958.\n\nThey also do a free pilot of 50 units so you can test before committing to a larger rollout.\n\nChris",
        "inbox",
    ),
    # === MERIDIAN HOUSING GROUP ===
    _e(
        "Asset database approach to EPC tracking",
        "2024-11-20T10:15:00Z",
        SL,
        S,
        "Hi Steph,\n\nFollowing the EPC data quality discussion, a few members asked about our approach to tracking installed measures.\n\nWe've built a simple Excel-based asset database that tracks what measures have actually been installed in each property (wall insulation, loft insulation, glazing, heating system) and calculates what the EPC should be. It's not official but it gives us a much more accurate picture than relying on the EPC register.\n\nI've attached the template if anyone wants to adapt it. Happy to walk through it on a call.\n\nSarah",
        "inbox",
    ),
    # === RIVERSIDE COMMUNITY HOUSING ===
    _e(
        "Tenant comms templates for retrofit works",
        "2025-02-03T09:30:00Z",
        JT,
        S,
        "Hi Steph,\n\nAs discussed, I've attached our template tenant notification letters for different types of retrofit works \u2014 PV installations, insulation, window replacements, and heat pumps.\n\nThe key thing we learned is to be specific about savings \u2014 'you could save \u00a3150-200/year' rather than just 'you'll save money'. That made a big difference to our take-up rates.\n\nHappy for you to share with the group.\n\nJames",
        "inbox",
    ),
    _e(
        "Stock condition survey results",
        "2025-01-08T15:45:00Z",
        JT,
        S,
        "Hi Steph,\n\nOur stock condition survey is now complete. Key findings:\n- 20% discrepancy between EPC records and actual condition\n- 340 properties need urgent window replacement\n- 85 properties identified as potential damp & mould risks\n- Average SAP rating across stock: 62 (target: 69)\n\nWe're using this to prioritise our retrofit investment plan for 2025-2028. Happy to share the methodology if other members are planning similar surveys.\n\nJames",
        "inbox",
    ),
    # === ASHWORTH HOUSING TRUST ===
    _e(
        "SHDF match funding question",
        "2024-12-08T14:00:00Z",
        NH,
        S,
        "Hi Steph,\n\nFollowing the SHDF discussion on the portal, I'm still trying to clarify the match funding rules. Does the 33% match have to come from our own reserves, or can we use Affordable Homes Programme funding?\n\nOur finance director is asking and I can't find a clear answer in the guidance.\n\nNadeem",
        "inbox",
    ),
    _e(
        "RE: SHDF match funding question",
        "2024-12-09T09:30:00Z",
        S,
        NH,
        "Hi Nadeem,\n\nI checked with DESNZ and the answer is: the match funding must come from the housing association's own resources or private borrowing. You cannot use other government grant funding (including AHP) as match.\n\nHowever, you CAN use the Boiler Upgrade Scheme alongside SHDF for heat pump installations specifically \u2014 they count as different pots.\n\nSteph",
        "sent",
    ),
    # === SUMMIT HOUSING PARTNERSHIP ===
    _e(
        "Tenant liaison officer - job description",
        "2025-02-18T11:30:00Z",
        DW,
        S,
        "Hi Steph,\n\nFollowing the tenant comms discussion, we've been looking at hiring a dedicated tenant liaison officer for our retrofit programme. It's made such a difference for Riverside and Lakeside.\n\nWould you be able to share any example job descriptions from other members who've recruited for this role?\n\nDarren",
        "inbox",
    ),
    _e(
        "RE: Tenant liaison officer - job description",
        "2025-02-19T09:00:00Z",
        S,
        DW,
        "Hi Darren,\n\nGood move! I'll ask James at Riverside and Chris at Lakeside if they're happy to share their JDs. Both have had great success with dedicated liaison officers.\n\nKey things to include based on what I've heard:\n- Single point of contact for tenants throughout the process\n- Both pre-visit engagement and post-install follow-up\n- Ability to handle complaints and escalate where needed\n- Some technical knowledge (doesn't need to be an expert but should understand the basics of the measures being installed)\n\nSteph",
        "sent",
    ),
    # === IRONBRIDGE HOMES ===
    _e(
        "Regulatory compliance query - Awaab's Law",
        "2025-01-20T16:00:00Z",
        SB,
        S,
        "Hi Steph,\n\nWith Awaab's Law now in force, we're reviewing our damp and mould response times. We currently target 5 working days for initial inspection but I'm not sure that's fast enough.\n\nWhat are other members targeting? I saw on the portal that Northfield are doing 48 hours which seems ambitious but they've had great results.\n\nSamira",
        "inbox",
    ),
    _e(
        "RE: Regulatory compliance query - Awaab's Law",
        "2025-01-21T09:45:00Z",
        S,
        SB,
        "Hi Samira,\n\nGood question and very timely. The range across Carbon Club members is:\n- Northfield: 48 hours (best in class)\n- Lakeside: 72 hours (supported by sensor monitoring)\n- Most others: 3-5 working days\n\nI'd recommend tightening to at least 72 hours for initial inspection. The key is having a triage system \u2014 Northfield categorise into emergency (same day), urgent (48 hours), and routine (5 days).\n\nI'll share the full damp & mould portal discussion with you.\n\nSteph",
        "sent",
    ),
    # === HARROWFIELD HOMES ===
    _e(
        "Damp & mould training programme",
        "2025-02-08T13:00:00Z",
        JB,
        S,
        "Hi Steph,\n\nWe've just rolled out damp and mould assessment training for all our housing officers. The idea is early identification during routine visits before tenants even report it.\n\nIt's a half-day session covering:\n- Visual signs of damp vs condensation\n- Moisture meter use\n- When to escalate to a specialist surveyor\n- How to advise tenants on ventilation\n\nWould other members be interested in the training materials? We developed it with a surveying firm and they've said we can share it.\n\nJoanna",
        "inbox",
    ),
    # === FOXLEY HOMES ===
    _e(
        "SHDF guidance request",
        "2025-03-04T10:00:00Z",
        DnW,
        S,
        "Hi Steph,\n\nI know you shared the SHDF Wave 2 guidance with some members individually but I don't think everyone got it. Could you send it round to the whole group? We're starting to think about an application.\n\nAlso, are there any upcoming sessions focused specifically on SHDF bid preparation? I think there's quite a lot of interest.\n\nDanielle",
        "inbox",
    ),
    # === INTERNAL — STEPH <> ROB ===
    _e(
        "Carbon Club - March session agenda",
        "2025-02-28T16:00:00Z",
        S,
        R,
        "Hi Rob,\n\nDraft agenda for next week's Carbon Club session (5 March):\n\n1. Welcome and housekeeping\n2. Spotlight: Greendale PV update (Mark Jennings)\n3. Spotlight: Northfield battery storage (Fiona Clarke)\n4. SHDF Wave 2 funding update\n5. EPC data quality roundtable\n6. Tenant communications for retrofit\n7. Wrap-up and actions\n\nI want to flag that Riverview Estates haven't attended the last 3 sessions. Helen says they're stretched with damp & mould cases. Their renewal is in April \u2014 I'm worried we might lose them.\n\nSteph",
        "sent",
    ),
    _e(
        "RE: Carbon Club - March session agenda",
        "2025-03-01T08:45:00Z",
        R,
        S,
        "Agenda looks good Steph.\n\nRe Riverview \u2014 yes, let's discuss. Maybe we need to offer them some targeted support on the damp & mould side to demonstrate value before the renewal conversation. Chris at Lakeside has done great work with those sensors \u2014 could we connect them?\n\nRob",
        "inbox",
    ),
    _e(
        "Broadland Housing renewal - April 2025",
        "2025-03-03T11:00:00Z",
        S,
        R,
        "Hi Rob,\n\nHeads up that Broadland Housing's renewal is coming up on 1 April. They're in Carbon Club (\u00a34,200/yr) and Exec Club (\u00a34,300/yr) \u2014 total \u00a38,500.\n\nThey've been good members \u2014 attended 10 of 11 sessions this year. Lena Morris has been very engaged, particularly around starting a PV programme. I've sent her case studies, connected her with Mark at Greendale for a peer visit, and shared SHDF funding guidance.\n\nOne coaching point for the renewal call: their complaint handling satisfaction is at 31%, below the club average of 35%. And their EPC C+ stock is only 63% vs club average of 72% \u2014 which is actually a good argument for staying in the Carbon Club as they've got a lot to learn from peers.\n\nI'll prepare a full renewal brief before the call.\n\nSteph",
        "sent",
    ),
    _e(
        "Q3 club performance summary",
        "2025-01-12T10:00:00Z",
        S,
        R,
        "Hi Rob,\n\nHere's the Q3 performance summary across all clubs:\n\n- Carbon Club: 28 members, 85% average session attendance, 2 renewals at risk (Riverview, Thornbury)\n- Executive Club: 50 members, 78% attendance, no immediate concerns\n- Finance Directors: 42 members, 72% attendance, 1 renewal at risk\n- Operations: 45 members, 80% attendance, strong engagement\n- Customer Experience: 42 members, 75% attendance, growing well\n- Performance Improvement: 45 members, 82% attendance, flagship club\n\nTotal membership revenue on track at \u00a31.1m for the year. Main risk is Riverview \u2014 they're in Carbon Club only at \u00a34,800/yr but the disengagement pattern is concerning.\n\nSteph",
        "sent",
    ),
    _e(
        "New member enquiry - Pinnacle Living",
        "2025-02-24T14:00:00Z",
        R,
        S,
        "Steph,\n\nHad a call with Sarah Thompson at Pinnacle Living (Doncaster, ~6,800 homes). They're interested in joining the Carbon Club \u2014 they're just starting their decarbonisation strategy and want to learn from peers.\n\nCan you follow up and send them the membership info? I think they'd be a good fit.\n\nRob",
        "inbox",
    ),
    _e(
        "RE: New member enquiry - Pinnacle Living",
        "2025-02-25T09:00:00Z",
        S,
        R,
        "Will do \u2014 I'll reach out to Sarah today. Another Carbon Club member would be great, especially as they're early in their journey. Lots to learn from the established members.\n\nSteph",
        "sent",
    ),
    # === CRESTWOOD HOMES ===
    _e(
        "Contractor frameworks for insulation programme",
        "2025-01-14T10:00:00Z",
        PG,
        S,
        "Hi Steph,\n\nWe're planning a major cavity wall insulation programme for 2025/26 \u2014 about 400 properties. I've been hearing good things about how Greendale managed their PV procurement through Solarfix/Fusion21.\n\nDo you know of any similar frameworks specifically for insulation works? We've struggled with contractor quality in the past and want to use a framework that pre-vets suppliers.\n\nPaul",
        "inbox",
    ),
    _e(
        "RE: Contractor frameworks for insulation programme",
        "2025-01-15T09:30:00Z",
        S,
        PG,
        "Hi Paul,\n\nGood timing \u2014 a few members have used frameworks successfully for insulation:\n- Fusion21 has an insulation-specific lot that Beacon and Pennine Valleys have used\n- Procurement for Housing (PfH) has a retrofit framework covering cavity wall, EWI, and loft insulation\n- Northfield went direct to market but used Savills as employer's agent for quality assurance\n\nI'd recommend talking to Claire at Pennine Valleys \u2014 they recently completed a similar-sized programme.\n\nSteph",
        "sent",
    ),
    # === THORNBURY HA ===
    _e(
        "EPC targets for 2026 and regulatory pressure",
        "2024-12-10T14:30:00Z",
        RI,
        S,
        "Hi Steph,\n\nOur board has set an ambitious target of 80% of stock at EPC C or above by March 2026. We're currently at 70% which means we need to improve about 900 properties in 15 months.\n\nAre other members facing similar pressure? I'm worried about the pace we need to sustain, especially given contractor availability challenges.\n\nRachel",
        "inbox",
    ),
    _e(
        "RE: EPC targets for 2026 and regulatory pressure",
        "2024-12-11T09:15:00Z",
        S,
        RI,
        "Hi Rachel,\n\nYes, several members are facing similar board-level EPC targets. The range I'm seeing across the club is:\n- Top performers: 82-85% at C+ (Severn Vale, Wychwood, Dales & Moorland)\n- Mid-range: 72-78% (most members)\n- Those with more work to do: 63-70% (Broadland, Riverview)\n\n900 properties in 15 months is ambitious but achievable with the right contractor capacity. SHDF Wave 2.2 could help fund some of this \u2014 application window opens April.\n\nSteph",
        "sent",
    ),
    # === WYCHWOOD HA ===
    _e(
        "EV charging infrastructure - what are others doing?",
        "2025-02-10T11:00:00Z",
        MC,
        S,
        "Hi Steph,\n\nWe've started getting requests from tenants about EV charging points. We've installed 20 so far in our newer developments but we need a strategy for older stock where off-street parking is limited.\n\nAre any Carbon Club members doing EV charging at scale? I'd be interested in costs per unit and any experience with communal charging hubs.\n\nMei-Lin",
        "inbox",
    ),
    _e(
        "RE: EV charging infrastructure",
        "2025-02-11T09:30:00Z",
        S,
        MC,
        "Hi Mei-Lin,\n\nEV charging is a growing topic. Here's what I know from across the club:\n- Thameside have the most points (45) \u2014 mainly in new builds and car parks\n- Greendale have 35 across their estate car parks\n- Riverside installed 30 \u2014 they partnered with Pod Point who funded installation in exchange for a revenue share\n\nThe communal hub approach is worth exploring for older stock. I'll add this as a discussion topic for a future session.\n\nSteph",
        "sent",
    ),
    # === DALES & MOORLAND ===
    _e(
        "Whole-house retrofit approach - interested",
        "2025-01-28T15:30:00Z",
        IC,
        S,
        "Hi Steph,\n\nI've been very impressed by what Andrew at Severn Vale has achieved with whole-house retrofit \u2014 taking properties from EPC E to B. We've got about 400 off-gas properties in rural locations where a combined approach (insulation + heat pump + PV) could make real sense.\n\nCould you set up a conversation between me and Andrew? I'd also like to understand the funding model they used.\n\nIan",
        "inbox",
    ),
    _e(
        "RE: Whole-house retrofit approach",
        "2025-01-29T10:00:00Z",
        S,
        IC,
        "Hi Ian,\n\nAbsolutely \u2014 Andrew has offered to present on the whole-house approach at a future session but I'll also connect you directly. He used a combination of SHDF funding and own capital, and the in-house DLO kept costs down significantly.\n\nYour rural off-gas stock is exactly the type where this approach makes the strongest financial case. With the Boiler Upgrade Scheme providing \u00a37,500 per heat pump and potential SHDF match funding, the economics could be very favourable.\n\nSteph",
        "sent",
    ),
    # === STONEBRIDGE HOUSING GROUP ===
    _e(
        "Window replacement programme - scale and costs",
        "2025-02-05T13:00:00Z",
        WK,
        S,
        "Hi Steph,\n\nWe're planning our largest ever window replacement programme \u2014 1,200 properties over two years. It's our biggest single EPC improvement measure.\n\nDo you have any benchmarking data on window replacement costs? We're getting quotes ranging from \u00a33,500 to \u00a35,500 per property and I want to make sure we're not being overcharged.\n\nWayne",
        "inbox",
    ),
    _e(
        "RE: Window replacement programme",
        "2025-02-06T09:45:00Z",
        S,
        WK,
        "Hi Wayne,\n\nThat's a significant programme! From the data I have across Carbon Club members:\n- Average window replacement cost: \u00a34,200 per property (typical 3-bed)\n- Range: \u00a33,500-\u00a35,000 depending on specification and volume\n- Key factors: double vs triple glazing, number of windows, access requirements\n\nRiverside recently did 340 properties and got costs down to \u00a33,800/unit through a framework. Happy to connect you with James Thornton.\n\nSteph",
        "sent",
    ),
    # === AVONDALE COMMUNITY HOMES ===
    _e(
        "Tenant satisfaction - improvement strategies",
        "2025-02-28T11:00:00Z",
        ES,
        S,
        "Hi Steph,\n\nOur overall satisfaction improved from 73% to 76% this year which we're pleased about, but our complaint handling score is still only at 40%. We know the club average is higher.\n\nCould you share any examples of what top-performing members are doing differently on complaints? We've invested in a new CRM system but I think the issue is more about process and culture than technology.\n\nEmma",
        "inbox",
    ),
    # === EASTGATE HA ===
    _e(
        "New member - Carbon Club introduction",
        "2025-03-01T10:00:00Z",
        MO,
        S,
        "Hi Steph,\n\nAs a relatively new member (joined late 2023), I just wanted to say how valuable the Carbon Club has been. The February session on PV was excellent and we're now seriously considering starting our own programme.\n\nIs there a summary of key contacts across the group? I'd like to reach out to a few members directly to compare notes on decarbonisation strategies.\n\nMartin",
        "inbox",
    ),
    # === CHILTERN EDGE HOUSING ===
    _e(
        "SHDF Wave 2.2 - timeline concerns",
        "2025-03-03T14:00:00Z",
        AR,
        S,
        "Hi Steph,\n\nWe're planning a SHDF Wave 2.2 application but I'm concerned about the timeline. Claire mentioned it took six weeks at Pennine Valleys and the window opens 1 April.\n\nWe haven't started pulling our property-level data together yet. Is it realistic to get an application in by 30 June, or should we wait for the next round?\n\nAbigail",
        "inbox",
    ),
    _e(
        "RE: SHDF Wave 2.2 - timeline concerns",
        "2025-03-04T09:00:00Z",
        S,
        AR,
        "Hi Abigail,\n\nYou've got three months (April to June) which should be enough if you start now. Key things to prepare:\n1. Property-level EPC data (you should have this already)\n2. Indicative contractor pricing (framework prices are fine at application stage)\n3. Tenant engagement plan (can be high-level)\n4. Match funding confirmation from your finance team\n\nI'd strongly recommend connecting with Claire at Pennine Valleys and David at Beacon \u2014 both can share their actual applications as templates.\n\nSteph",
        "sent",
    ),
    # === MAPLEWOOD LIVING ===
    _e(
        "Benchmarking data access request",
        "2025-02-20T10:30:00Z",
        GK,
        S,
        "Hi Steph,\n\nI'm preparing a board paper on our decarbonisation programme and need some comparative data. Specifically:\n- How do our PV installations (560) compare to other members?\n- Where do we sit on void turnaround (19 days)?\n- How does our EPC C+ stock (77%) compare to the club average?\n\nCould you pull together a quick benchmarking snapshot?\n\nGeorge",
        "inbox",
    ),
    _e(
        "RE: Benchmarking data access request",
        "2025-02-21T09:30:00Z",
        S,
        GK,
        "Hi George,\n\nHappy to help. Quick snapshot:\n- PV installations: your 560 is above the club average. Only a handful have done more (Greendale 1,200, Thameside 1,200, Stonebridge 900, Northfield 850)\n- Void turnaround: 19 days puts you slightly below the club average of about 21 days \u2014 good position\n- EPC C+ at 77%: above the club average of around 73%\n\nOverall you're performing well relative to peers. I'll prepare a more detailed pack for your board paper.\n\nSteph",
        "sent",
    ),
    # === ADDITIONAL RIVERVIEW ===
    _e(
        "Missed session again - apologies",
        "2025-03-06T10:00:00Z",
        HF,
        S,
        "Hi Steph,\n\nI'm so sorry we missed the session again yesterday. I know it's the fourth time and I feel terrible about it. The damp and mould situation has been overwhelming \u2014 15 new cases this month alone on top of the existing backlog.\n\nI do value the club and want to keep attending. Could you send me the notes from yesterday? And I'd really like to take Chris up on the offer about the environmental sensors \u2014 if they can help us get ahead of cases, that could be transformative.\n\nHelen",
        "inbox",
    ),
    _e(
        "RE: Missed session again - apologies",
        "2025-03-07T08:30:00Z",
        S,
        HF,
        "Hi Helen,\n\nPlease don't worry \u2014 I completely understand the pressure you're under. Session notes attached.\n\nI've copied Chris Doyle (Lakeside Living) \u2014 Chris, would you mind sharing the Switchee sensor details with Helen? Lakeside reduced reactive D&M callouts by 30% with environmental monitoring and I think it could really help Riverview.\n\nHelen, your renewal is coming up in April. Let's have a proper conversation about how we can support you better \u2014 maybe we can tailor the engagement to fit your current situation.\n\nSteph",
        "sent",
    ),
    # === INTERNAL ===
    _e(
        "Q4 revenue forecast and renewal pipeline",
        "2025-03-10T11:00:00Z",
        S,
        R,
        "Hi Rob,\n\nQuick Q4 update:\n\n- Total membership revenue on track at \u00a31.1m\n- April renewals: Broadland (\u00a38,500 \u2014 confident), Riverview (\u00a34,800 \u2014 at risk), Thameside (\u00a322,500 \u2014 confident)\n- Pinnacle Living enquiry progressing \u2014 Sarah Thompson keen, likely Carbon Club initially at \u00a34,200\n\nKey risk is Riverview. Helen's engagement has dropped significantly but she's reaching out now. I've connected her with Chris at Lakeside for the sensor programme.\n\nSteph",
        "sent",
    ),
    _e(
        "RE: Q4 revenue forecast and renewal pipeline",
        "2025-03-11T09:00:00Z",
        R,
        S,
        "Thanks Steph, good update. Let's make sure we have a clear value proposition for Helen before the renewal conversation. The sensor connection is a good move.\n\nAlso, can you schedule Andrew Marsh's whole-house retrofit spotlight for the April or May session? A few members have been asking about it and it would be a strong retention argument.\n\nRob",
        "inbox",
    ),
]

# =========================================================================
# ONENOTE DATA — expanded to 15 renewal notes
# =========================================================================

RENEWAL_NOTES = [
    {
        "org": "Broadland Housing",
        "date": "5 September 2024",
        "attendees": "Rob Bryan, Lena Morris (Director of Assets, Broadland)",
        "bullets": [
            "Broadland very interested in starting a PV programme \u2014 asked for contacts at orgs who have done it at scale",
            "Renewed for Carbon Club + Exec Club (2-year deal, \u00a38,500/yr)",
            "Lena mentioned they're also looking at air source heat pumps for off-gas properties",
            "Currently 63% of stock at EPC C or above \u2014 wants to improve significantly",
            "Lena very engaged \u2014 attended 10 of 11 sessions this year",
        ],
        "action": "Steph to send examples of PV case studies from other members",
    },
    {
        "org": "Westmoor Housing Group",
        "date": "12 September 2024",
        "attendees": "Steph Hosny, Karen Blackwell (Head of Sustainability, Westmoor)",
        "bullets": [
            "Renewed Carbon Club only this year (\u00a34,200/yr)",
            "Karen flagged concerns about void standards \u2014 turnaround times increased to 28 days avg, board asking questions",
            "Interested in benchmarking void costs against peers",
            "Also asked about tenant engagement approaches for upcoming insulation programme",
            "Mentioned budget for face-to-face home visits as part of insulation rollout",
        ],
        "action": "Share void standards thread from Zoho portal",
    },
    {
        "org": "Greendale Homes",
        "date": "20 September 2024",
        "attendees": "Steph Hosny, Mark Jennings (Head of Sustainability, Greendale)",
        "bullets": [
            "PV programme going very well \u2014 on track for 1,200 installs by year end",
            "Mark happy to do a spotlight at Carbon Club and host peer visits",
            "Renewed Carbon Club + Performance Club (\u00a39,200/yr)",
            "Raised EPC data quality as emerging concern \u2014 working with Elmhurst Energy to resurvey 2,000 properties",
            "Shared 12-month PV performance report \u2014 3,800 kWh avg per property",
        ],
        "action": "Schedule Greendale PV spotlight for Feb 2025 session",
    },
    {
        "org": "Northfield Housing Association",
        "date": "3 October 2024",
        "attendees": "Steph Hosny, Fiona Clarke (Director of Assets, Northfield)",
        "bullets": [
            "PV+battery programme at 500 installs, targeting 850 by end of Q4",
            "Fiona keen to share the cost-benefit analysis showing battery value",
            "Also piloting 50 heat pumps \u2014 mixed results so far on tenant satisfaction (76%)",
            "Noise complaints from 15% of tenants with heat pumps",
            "Renewed Carbon Club + Operations Club (\u00a38,100/yr)",
        ],
        "action": "Schedule Northfield battery storage spotlight for Feb session",
    },
    {
        "org": "Thameside Housing Trust",
        "date": "15 October 2024",
        "attendees": "Rob Bryan, Oliver Grant (Group Director of Assets, Thameside)",
        "bullets": [
            "120 heat pump installs completed across two estates",
            "Satisfaction at 78% \u2014 lower than target. Noise and instant heat concerns",
            "Oliver worried about EPC data quality and regulatory reporting implications",
            "Renewed all five clubs (\u00a322,500/yr) \u2014 largest single client",
            "Board presentation on heat pump programme scheduled for February",
        ],
        "action": "Connect Oliver with Sarah (Meridian) re asset database approach to EPC tracking",
    },
    {
        "org": "Oaktree Living",
        "date": "28 October 2024",
        "attendees": "Jane, Priya Sharma (Sustainability Lead, Oaktree)",
        "bullets": [
            "No PV programme yet but watching Greendale and Northfield closely",
            "Main focus currently on cavity wall insulation \u2014 400 properties targeted",
            "Contractor availability a major challenge \u2014 impacting both voids and retrofit",
            "Seriously considering in-house void team \u2014 current 25 day avg costing \u00a3400k/yr in lost rent",
            "Renewed Carbon Club only (\u00a33,800/yr)",
        ],
        "action": "Share contractor framework information from other members; connect with Andrew Marsh (Severn Vale) re DLO setup",
    },
    {
        "org": "Severn Vale Homes",
        "date": "8 November 2024",
        "attendees": "Steph Hosny, Andrew Marsh (Operations Director, Severn Vale)",
        "bullets": [
            "Whole-house retrofit programme completed \u2014 200 properties E to B",
            "Andrew very positive about in-house DLO approach \u2014 wants to present to the group",
            "DLO setup cost was \u00a3185k, broke even in 14 months",
            "Looking at expanding programme to additional 300 properties in 2025/26",
            "Renewed Carbon Club + Customer Experience Club (\u00a37,600/yr)",
        ],
        "action": "Schedule Severn Vale retrofit spotlight for future session",
    },
    {
        "org": "Riverview Estates",
        "date": "14 November 2024",
        "attendees": "Steph Hosny, Helen Foster (Head of Housing, Riverview)",
        "bullets": [
            "Helen seemed disengaged \u2014 said team is stretched and struggling to attend sessions",
            "Riverview haven't attended last two Carbon Club sessions",
            "Main focus currently on damp and mould compliance \u2014 42 cases per 1,000 homes",
            "Renewal coming up April 2025 \u2014 risk of non-renewal if engagement doesn't improve",
            "Suggested connecting Helen with Chris Doyle (Lakeside) re sensor monitoring approach",
        ],
        "action": "Steph to follow up with Helen in January to discuss engagement",
    },
    {
        "org": "Beacon Dwellings",
        "date": "22 November 2024",
        "attendees": "Rob Bryan, David Osei (Sustainability Manager, Beacon)",
        "bullets": [
            "PV programme at 340 installs as part of wider EPC programme",
            "EPC data quality issues \u2014 25% inaccuracy for solid wall stock",
            "Planning full resurvey in 2025/26 to get accurate baseline",
            "Hoping to reach 500 PV installs by end of financial year",
            "Renewed Carbon Club + Finance Directors Club (\u00a37,800/yr)",
        ],
        "action": "Connect David with Mark (Greendale) re Elmhurst Energy EPC surveys",
    },
    {
        "org": "Lakeside Living",
        "date": "5 December 2024",
        "attendees": "John, Chris Doyle (Head of Repairs & Maintenance, Lakeside)",
        "bullets": [
            "Invested in 500 environmental monitoring sensors for damp and mould prevention",
            "Proactive approach \u2014 sensors alert team before mould develops",
            "Cost about \u00a380 per sensor including installation, supplier is Switchee",
            "Reduced reactive damp & mould callouts by 30% in monitored properties",
            "Renewed Carbon Club + Operations Club (\u00a37,400/yr)",
        ],
        "action": "Chris to present on sensor programme at future session",
    },
    {
        "org": "Riverside Community Housing",
        "date": "10 December 2024",
        "attendees": "Steph Hosny, James Thornton (Head of Property Services, Riverside)",
        "bullets": [
            "Stock condition survey now complete \u2014 20% discrepancy with EPC records",
            "340 properties need urgent window replacement",
            "85 properties identified as potential damp & mould risks",
            "Developed excellent tenant comms templates \u2014 willing to share with group",
            "Renewed all three clubs (\u00a312,200/yr)",
        ],
        "action": "James to send tenant comms templates to Steph for distribution",
    },
    {
        "org": "Meridian Housing Group",
        "date": "15 December 2024",
        "attendees": "Steph Hosny, Sarah Linehan (COO, Meridian)",
        "bullets": [
            "Built own asset database for tracking installed measures against properties",
            "Excel-based tool that calculates expected EPC from installed measures",
            "More accurate than official EPC register for investment planning",
            "Several members interested in the methodology",
            "Renewed Carbon Club + Executive Club (\u00a39,800/yr)",
        ],
        "action": "Sarah to share asset database template with interested members",
    },
    {
        "org": "Summit Housing Partnership",
        "date": "8 January 2025",
        "attendees": "John, Darren Walsh (Director of Property, Summit)",
        "bullets": [
            "Planning to recruit dedicated tenant liaison officer for retrofit programme",
            "Inspired by Riverside and Lakeside examples",
            "Looking for example job descriptions from other members",
            "Retrofit programme covering 600 properties over next 2 years",
            "Renewed Carbon Club + Operations Club (\u00a38,200/yr)",
        ],
        "action": "Share tenant liaison officer JD examples from Riverside and Lakeside",
    },
    {
        "org": "Ironbridge Homes",
        "date": "15 January 2025",
        "attendees": "Steph Hosny, Samira Begum (Head of Compliance, Ironbridge)",
        "bullets": [
            "Concerned about Awaab's Law compliance \u2014 current damp response time is 5 working days",
            "Needs to tighten to at least 72 hours for initial inspection",
            "Interested in Northfield's triage system (emergency/urgent/routine)",
            "Renewed Carbon Club only (\u00a34,100/yr)",
            "Samira quite new in role \u2014 finding Carbon Club very valuable for peer learning",
        ],
        "action": "Share damp & mould portal thread and Northfield's triage approach with Samira",
    },
    {
        "org": "Ashworth Housing Trust",
        "date": "22 January 2025",
        "attendees": "Rob Bryan, Nadeem Hussain (Property Director, Ashworth)",
        "bullets": [
            "Exploring SHDF Wave 2.2 application for insulation programme",
            "Clarified match funding must come from own reserves (not AHP)",
            "Boiler Upgrade Scheme can be used alongside SHDF for heat pumps",
            "Planning 200-property insulation programme if SHDF bid successful",
            "Renewed Carbon Club + Finance Directors Club (\u00a37,200/yr)",
        ],
        "action": "Share SHDF guidance document and connect with Claire Whitfield (Pennine Valleys) for application advice",
    },
    {
        "org": "Pennine Valleys Housing",
        "date": "3 February 2025",
        "attendees": "Steph Hosny, Claire Whitfield (Asset Strategy Manager, Pennine Valleys)",
        "bullets": [
            "SHDF Wave 2.1 delivery progressing well \u2014 180 of 450 measures completed",
            "Remaining 270 measures scheduled for Q1-Q2 2025/26",
            "Tenant engagement for external wall insulation has been the biggest challenge",
            "Face-to-face visits working well but time-consuming",
            "Considering SHDF Wave 2.2 bid for additional properties",
        ],
        "action": "Share SHDF Wave 2.1 delivery lessons with members considering Wave 2.2 applications",
    },
    {
        "org": "Harrowfield Homes",
        "date": "10 February 2025",
        "attendees": "Steph Hosny, Joanna Briggs (Asset Manager, Harrowfield)",
        "bullets": [
            "Rolled out damp and mould assessment training for all housing officers",
            "Half-day sessions covering visual identification, moisture meters, escalation",
            "Developed training with external surveying firm \u2014 willing to share materials",
            "Early identification during routine visits catching issues before tenants report",
            "Renewed Carbon Club only (\u00a34,200/yr) \u2014 good engagement",
        ],
        "action": "Circulate D&M training materials to interested members",
    },
    {
        "org": "Foxley Homes",
        "date": "17 February 2025",
        "attendees": "Steph Hosny, Danielle Webb (Head of Sustainability, Foxley)",
        "bullets": [
            "Planning first SHDF Wave 2.2 application \u2014 targeting 200 properties",
            "Requested SHDF guidance document to be shared with full group",
            "Interest in dedicated SHDF bid preparation session for Carbon Club",
            "PV programme at 180 installations \u2014 steady growth",
            "Renewed Carbon Club only (\u00a34,200/yr)",
        ],
        "action": "Share SHDF guidance with full group and consider dedicated bid preparation session",
    },
    {
        "org": "Crestwood Homes",
        "date": "24 February 2025",
        "attendees": "Rob Bryan, Paul Gorman (Head of Asset Management, Crestwood)",
        "bullets": [
            "Planning major cavity wall insulation programme \u2014 400 properties",
            "Struggling with contractor quality \u2014 asked about procurement frameworks",
            "Attendance has been inconsistent \u2014 missed 3 of last 4 sessions",
            "Paul seemed positive about club value but said workload is affecting attendance",
            "Renewed Carbon Club only (\u00a34,200/yr) \u2014 may need attention at next renewal",
        ],
        "action": "Share contractor framework information and encourage more consistent attendance",
    },
    {
        "org": "Thornbury Housing Association",
        "date": "3 March 2025",
        "attendees": "Steph Hosny, Rachel Iqbal (Sustainability Director, Thornbury)",
        "bullets": [
            "Board set ambitious EPC target \u2014 80% C+ by March 2026 (currently 70%)",
            "Need to improve 900 properties in 15 months \u2014 Rachel concerned about pace",
            "Attendance has dropped significantly \u2014 only 7 of 12 sessions this year",
            "Rachel says workload and board pressure making it hard to attend regularly",
            "Renewed Carbon Club + Operations Club (\u00a38,100/yr) but engagement is a concern",
        ],
        "action": "Monitor engagement closely; offer targeted support on EPC improvement planning",
    },
]


# =========================================================================
# SEED FUNCTIONS
# =========================================================================


def seed_emails():
    print(f"\nSeeding {len(EMAILS)} emails...")
    success = 0
    for email in EMAILS:
        folder = "SentItems" if email["direction"] == "sent" else "Inbox"
        message = {
            "subject": email["subject"],
            "body": {"contentType": "Text", "content": email["body"]},
            "from": {
                "emailAddress": {
                    "name": email["from"]["name"],
                    "address": email["from"]["email"],
                },
            },
            "toRecipients": [
                {
                    "emailAddress": {
                        "name": email["to"]["name"],
                        "address": email["to"]["email"],
                    },
                },
            ],
            "receivedDateTime": email["date"],
            "sentDateTime": email["date"],
            "isRead": True,
        }
        r = requests.post(
            f"{BASE}/me/mailFolders/{folder}/messages",
            headers=H_JSON,
            json=message,
        )
        d = "SENT" if email["direction"] == "sent" else "RECV"
        other = (
            email["to"]["name"]
            if email["direction"] == "sent"
            else email["from"]["name"]
        )
        if r.status_code in (200, 201):
            success += 1
            print(
                f"  [OK] {d} | {email['date'][:10]} | {other} | {email['subject'][:60]}",
            )
        else:
            print(f"  [FAIL] {r.status_code} | {email['subject'][:60]}")
        time.sleep(0.3)
    print(f"  {success}/{len(EMAILS)} emails created.")


def seed_onenote():
    print(f"\nSeeding {len(RENEWAL_NOTES)} OneNote pages...")

    r = requests.get(f"{BASE}/me/onenote/notebooks", headers=H_AUTH)
    nb_id = None
    for nb in r.json().get("value", []):
        if nb["displayName"] == "Member Notes":
            nb_id = nb["id"]
    if not nb_id:
        r = requests.post(
            f"{BASE}/me/onenote/notebooks",
            headers=H_JSON,
            json={"displayName": "Member Notes"},
        )
        nb_id = r.json()["id"]
        print("  Created notebook: Member Notes")
    time.sleep(2)

    r = requests.get(f"{BASE}/me/onenote/notebooks/{nb_id}/sections", headers=H_AUTH)
    sec_id = None
    for sec in r.json().get("value", []):
        if sec["displayName"] == "Renewal Calls":
            sec_id = sec["id"]
    if not sec_id:
        r = requests.post(
            f"{BASE}/me/onenote/notebooks/{nb_id}/sections",
            headers=H_JSON,
            json={"displayName": "Renewal Calls"},
        )
        sec_id = r.json()["id"]
        print("  Created section: Renewal Calls")
    time.sleep(2)

    success = 0
    for note in RENEWAL_NOTES:
        bullets_html = "\n".join(f"<li>{b}</li>" for b in note["bullets"])
        html = f"""<!DOCTYPE html>
<html><head><title>{note['org']} \u2014 {note['date']}</title></head>
<body>
<h1>{note['org']}</h1>
<p><strong>Date:</strong> {note['date']}</p>
<p><strong>Attendees:</strong> {note['attendees']}</p>
<h2>Discussion Notes</h2>
<ul>{bullets_html}</ul>
<h2>Action Item</h2>
<p style="background-color:#FFF3CD;padding:8px;border-left:4px solid #FFC107;">
<strong>ACTION:</strong> {note['action']}</p>
</body></html>"""
        r = requests.post(
            f"{BASE}/me/onenote/sections/{sec_id}/pages",
            headers=H_HTML,
            data=html.encode("utf-8"),
        )
        if r.status_code in (200, 201):
            success += 1
            print(f"  [OK] {note['org']} \u2014 {note['date']}")
        else:
            print(f"  [FAIL] {note['org']}: {r.status_code}")
        time.sleep(1)
    print(f"  {success}/{len(RENEWAL_NOTES)} pages created.")


# =========================================================================
# ONEDRIVE FILE GENERATION
# =========================================================================

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill(start_color="3B2D5F", end_color="3B2D5F", fill_type="solid")


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 40)


def _wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── VfM Financial Data ────────────────────────────────────────────────────


def _gen_vfm():
    wb = Workbook()
    ws = wb.active
    ws.title = "VfM Financial Data 2024-25"
    headers = [
        "Organisation",
        "Total Units",
        "Headline Social Housing CPU (£) 2024/25",
        "Headline Social Housing CPU (£) 2023/24",
        "Reinvestment %",
        "Operating Margin (Social Housing) %",
        "Operating Margin (Overall) %",
        "Gearing %",
        "EBITDA MRI Interest Cover %",
        "ROCE %",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for o in ORGS:
        f = o["fin"]
        ws.append(
            [
                o["name"],
                o["homes"],
                f["cpu25"],
                f["cpu24"],
                f["reinvest"],
                f["margin_sh"],
                f["margin_all"],
                f["gearing"],
                f["ebitda"],
                f["roce"],
            ],
        )
    _auto_width(ws)
    return _wb_bytes(wb)


# ── TSM Satisfaction Data ─────────────────────────────────────────────────


def _gen_tsm():
    wb = Workbook()
    ws = wb.active
    ws.title = "TSM Satisfaction Data 2023-24"
    headers = [
        "Organisation",
        "Overall Satisfaction %",
        "Satisfaction with Repairs %",
        "Satisfaction with Time Taken for Repairs %",
        "Home Safety %",
        "Kept Informed %",
        "Fair & Respectful Treatment %",
        "Complaint Handling Satisfaction %",
        "Anti-Social Behaviour Handling %",
        "Contribution to Neighbourhood %",
        "Well-Maintained Home %",
        "Well-Maintained Communal Areas %",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for o in ORGS:
        t = o["tsm"]
        ws.append(
            [
                o["name"],
                t["overall"],
                t["repairs"],
                t["time"],
                t["safety"],
                t["informed"],
                t["fair"],
                t["complaint"],
                t["asb"],
                t["neighbourhood"],
                t["home"],
                t["communal"],
            ],
        )
    _auto_width(ws)
    return _wb_bytes(wb)


# ── Operational KPI Data ──────────────────────────────────────────────────


def _gen_ops_kpi():
    wb = Workbook()
    ws = wb.active
    ws.title = "Operational KPIs 2024-25"
    headers = [
        "Organisation",
        "Void Turnaround (days)",
        "Void Rent Loss (%)",
        "Stock at EPC C or above (%)",
        "Retrofit Measures Installed",
        "PV Installations",
        "Complaints per 1,000 Homes",
        "EV Charging Points Installed",
        "Damp & Mould Cases (per 1,000 Homes)",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for o in ORGS:
        p = o["ops"]
        ws.append(
            [
                o["name"],
                p["void_days"],
                p["rent_loss"],
                p["epc_c"],
                p["retrofit"],
                p["pv"],
                p["complaints"],
                p["ev"],
                p["damp"],
            ],
        )
    _auto_width(ws)
    return _wb_bytes(wb)


# ── 1-to-1 Call Notes ─────────────────────────────────────────────────────


def _1to1_entries(org):
    """Generate 6-8 1:1 call note rows for a single organisation."""
    o = org["ops"]
    c = f"{org['contact']} ({org['name']})"
    rows = []

    rows.append(
        (
            "15/03/2024",
            f"Steph Hosny; {c}",
            "general; engagement",
            3,
            f"Regular check-in. {org['name']} managing {org['homes']:,} homes in "
            f"{org['city']}. Discussed priorities for 2024/25 including "
            f"decarbonisation targets and tenant engagement.",
        ),
    )

    rows.append(
        (
            "20/06/2024",
            f"Rob Bryan; {c}",
            "membership; benchmarking",
            3,
            f"Mid-year review. {org['contact']} positive about club value and keen "
            f"on more benchmarking data. Attending sessions regularly. Discussed "
            f"upcoming workshop topics and peer connections.",
        ),
    )

    if o["pv"] >= 800:
        rows.append(
            (
                "12/01/2025",
                f"Steph Hosny; {c}",
                "solar PV; SHDF Wave 2",
                4,
                f"{org['name']} completed {o['pv']:,} PV installations. Procurement "
                f"through established framework. Tenant satisfaction high. Happy to "
                f"present at a future session and host peer visits.",
            ),
        )
    elif o["pv"] > 0:
        rows.append(
            (
                "12/01/2025",
                f"Steph Hosny; {c}",
                "solar PV; programme update",
                3,
                f"PV programme at {o['pv']} installations. Steady progress. "
                f"Exploring options for scaling up in 2025/26. Interested in "
                f"procurement frameworks used by other members.",
            ),
        )
    else:
        rows.append(
            (
                "12/01/2025",
                f"Steph Hosny; {c}",
                "decarbonisation; planning",
                2,
                f"No PV programme yet. {org['contact']} interested in learning from "
                f"members who have done PV at scale. Asked about procurement routes "
                f"and typical costs. Considering a programme for 2025/26.",
            ),
        )

    rows.append(
        (
            "18/07/2024",
            f"Steph Hosny; {c}",
            "EPC; data quality",
            3 if o["epc_c"] >= 73 else 2,
            f"EPC C+ stock at {o['epc_c']}%. "
            + (
                "Above club average — good position for regulatory reporting. "
                "Discussed plans to maintain momentum on retrofit measures."
                if o["epc_c"] >= 73
                else "Below club average — needs improvement. Discussed prioritisation "
                "of worst-performing stock and potential SHDF funding."
            ),
        ),
    )

    if o["void_days"] >= 23:
        rows.append(
            (
                "04/09/2024",
                f"Steph Hosny; {c}",
                "void turnaround; improvement",
                2,
                f"Void turnaround at {o['void_days']} days — above club average. "
                f"Board raising questions about performance. Discussed benchmarking "
                f"data and improvement strategies from top-performing members.",
            ),
        )
    else:
        rows.append(
            (
                "04/09/2024",
                f"Steph Hosny; {c}",
                "void turnaround; performance",
                4,
                f"Void turnaround at {o['void_days']} days — performing well "
                f"relative to peers. Shared approach with group. Discussed "
                f"maintaining performance as stock condition challenges increase.",
            ),
        )

    rows.append(
        (
            "15/10/2024",
            f"Rob Bryan; {c}",
            "renewal; membership",
            3,
            f"Renewal discussion for {org['renewal']}. Currently in "
            f"{', '.join(org['clubs'].keys())}. Total fee: "
            f"£{sum(org['clubs'].values()):,}/yr. Member since {org['since']}. "
            f"{'Straightforward renewal expected.' if sum(org['att'][-4:]) >= 3 else 'Engagement has dropped — renewal may need discussion.'}",
        ),
    )

    if o["damp"] >= 20:
        rows.append(
            (
                "22/11/2024",
                f"Steph Hosny; {c}",
                "damp and mould; Awaab's Law",
                2,
                f"Damp and mould rate at {o['damp']} per 1,000 homes — above "
                f"sector average. Team stretched dealing with cases. Reviewing "
                f"response times to ensure Awaab's Law compliance.",
            ),
        )

    rows.append(
        (
            "05/02/2025",
            f"Steph Hosny; {c}",
            "retrofit; progress",
            4 if o["retrofit"] >= 600 else 3,
            f"Retrofit programme at {o['retrofit']} measures completed. "
            + (
                "Strong progress — on track for year-end targets. Discussed "
                "plans for 2025/26 expansion."
                if o["retrofit"] >= 600
                else "Steady progress but room to accelerate. Discussed SHDF Wave 2.2 "
                "as potential funding source for additional measures."
            ),
        ),
    )

    return [
        {
            "date": r[0],
            "attendees": r[1],
            "tags": r[2],
            "score": r[3],
            "notes": r[4],
        }
        for r in rows
    ]


def _gen_1to1():
    wb = Workbook()
    wb.remove(wb.active)
    for org in ORGS:
        title = org["name"][:31]
        ws = wb.create_sheet(title=title)
        headers = ["Date", "Attendees", "Topic Tags", "Self-Score", "Notes"]
        ws.append(headers)
        _style_header(ws, len(headers))
        for e in _1to1_entries(org):
            ws.append(
                [
                    e["date"],
                    e["attendees"],
                    e["tags"],
                    e["score"],
                    e["notes"],
                ],
            )
        _auto_width(ws)
    return _wb_bytes(wb)


# ── Membership Master ─────────────────────────────────────────────────────


def _gen_membership():
    wb = Workbook()

    ws_orgs = wb.active
    ws_orgs.title = "Organisations"
    h1 = [
        "Org ID",
        "Organisation Name",
        "Total Homes",
        "HQ Location",
        "Member Since",
        "Account Manager",
        "Primary Contact",
        "Primary Contact Role",
        "Primary Contact Email",
    ]
    ws_orgs.append(h1)
    _style_header(ws_orgs, len(h1))
    for i, o in enumerate(ORGS, 1):
        mgr = "Steph Hosny" if i % 3 != 0 else ("Rob Bryan" if i % 3 == 0 else "Jane")
        ws_orgs.append(
            [
                f"ORG-{i:03d}",
                o["name"],
                o["homes"],
                o["city"],
                o["since"],
                mgr,
                o["contact"],
                o["role"],
                o["email"],
            ],
        )
    _auto_width(ws_orgs)

    ws_mem = wb.create_sheet("Memberships")
    h2 = [
        "Org ID",
        "Organisation Name",
        "Club",
        "Membership Fee (£/yr)",
        "Renewal Date",
        "Contract Length",
        "Status",
    ]
    ws_mem.append(h2)
    _style_header(ws_mem, len(h2))
    for i, o in enumerate(ORGS, 1):
        for club, fee in o["clubs"].items():
            ws_mem.append(
                [
                    f"ORG-{i:03d}",
                    o["name"],
                    club,
                    fee,
                    o["renewal"],
                    "1 year",
                    "Active",
                ],
            )
    _auto_width(ws_mem)

    ws_con = wb.create_sheet("Contacts")
    h3 = ["Org ID", "Name", "Role", "Email", "Club(s)", "Phone"]
    ws_con.append(h3)
    _style_header(ws_con, len(h3))
    for i, o in enumerate(ORGS, 1):
        clubs_str = "; ".join(o["clubs"].keys())
        ws_con.append(
            [
                f"ORG-{i:03d}",
                o["contact"],
                o["role"],
                o["email"],
                clubs_str,
                f"0{1200 + i * 7} {555:03d} {1000 + i * 13:04d}",
            ],
        )
    _auto_width(ws_con)

    return _wb_bytes(wb)


# ── Attendance Log ────────────────────────────────────────────────────────


def _gen_attendance():
    wb = Workbook()
    ws = wb.active
    ws.title = "Carbon Club Attendance"
    headers = ["Session Date", "Session Title"] + [o["name"] for o in ORGS]
    ws.append(headers)
    _style_header(ws, len(headers))
    for idx, (month, date) in enumerate(SESSION_DATES[:11]):
        row = [date, f"Carbon Club Workshop \u2014 {month}"]
        for org in ORGS:
            if org["att"][idx]:
                row.append(org["contact"])
            else:
                row.append("")
        ws.append(row)
    _auto_width(ws)
    return _wb_bytes(wb)


# ── Email Interaction Log ─────────────────────────────────────────────────

_INTERACTIONS = [
    (
        "15/10/2024",
        "Broadland Housing",
        "Lena Morris",
        "Outbound",
        "Solar PV case studies — Greendale & Northfield",
        "Sent PDF case studies for Greendale (1,200 installs) and Northfield (850 installs with battery storage).",
    ),
    (
        "22/11/2024",
        "Broadland Housing",
        "Lena Morris",
        "Outbound",
        "Void benchmarking data pack",
        "Sent void turnaround benchmarking data compiled from Carbon Club members.",
    ),
    (
        "08/01/2025",
        "Broadland Housing",
        "Lena Morris",
        "Outbound",
        "SHDF Wave 2 funding guidance",
        "Sent DESNZ guidance on SHDF Wave 2 application process and eligibility criteria.",
    ),
    (
        "16/10/2024",
        "Broadland Housing",
        "Lena Morris",
        "Inbound",
        "RE: Solar PV case studies",
        "Lena requested peer visit to Greendale and asked for SHDF guidance.",
    ),
    (
        "09/10/2024",
        "Beacon Dwellings",
        "David Osei",
        "Outbound",
        "Greendale EPC contact introduction",
        "Connected David with Mark Jennings (Greendale) re Elmhurst Energy EPC surveys.",
    ),
    (
        "15/09/2024",
        "Westmoor Housing Group",
        "Karen Blackwell",
        "Inbound",
        "Void standards benchmarking request",
        "Karen requested benchmarking data on void standards and turnaround times after board pressure.",
    ),
    (
        "16/09/2024",
        "Westmoor Housing Group",
        "Karen Blackwell",
        "Outbound",
        "Void benchmarking — initial pointers",
        "Shared quick pointers: Severn Vale 15 days (DLO), Northfield 16 days (pre-void inspections).",
    ),
    (
        "22/10/2024",
        "Severn Vale Homes",
        "Andrew Marsh",
        "Inbound",
        "DLO setup costs breakdown",
        "Andrew shared DLO setup costs: £185k total, 8 operatives, broke even in 14 months.",
    ),
    (
        "20/02/2025",
        "Oaktree Living",
        "Priya Sharma",
        "Inbound",
        "In-house void team — business case",
        "Priya interested in setting up in-house void team after Severn Vale discussion.",
    ),
    (
        "21/02/2025",
        "Oaktree Living",
        "Priya Sharma",
        "Outbound",
        "DLO connection — Andrew Marsh",
        "Connected Priya with Andrew Marsh (Severn Vale) re DLO setup costs and experience.",
    ),
    (
        "05/12/2024",
        "Beacon Dwellings",
        "David Osei",
        "Inbound",
        "EPC resurvey programme",
        "David flagged 25% EPC inaccuracy for solid wall stock. Planning full resurvey 2025/26.",
    ),
    (
        "15/01/2025",
        "Lakeside Living",
        "Chris Doyle",
        "Inbound",
        "Environmental monitoring sensors results",
        "Chris shared sensor programme results: 500 deployed, 30% fewer D&M callouts, £80/unit.",
    ),
    (
        "25/02/2025",
        "Lakeside Living",
        "Chris Doyle",
        "Inbound",
        "Sensor supplier details — Switchee",
        "Chris provided Switchee contact for sensor programme. £80/unit, 5-year data contract.",
    ),
    (
        "03/02/2025",
        "Riverside Community Housing",
        "James Thornton",
        "Inbound",
        "Tenant comms templates for retrofit",
        "James shared template notification letters for PV, insulation, windows, heat pumps.",
    ),
    (
        "05/11/2024",
        "Pennine Valleys Housing",
        "Claire Whitfield",
        "Inbound",
        "SHDF Wave 2.1 success",
        "Claire reported successful £1.2m SHDF bid for 450 measures across 280 properties.",
    ),
    (
        "08/12/2024",
        "Ashworth Housing Trust",
        "Nadeem Hussain",
        "Inbound",
        "SHDF match funding question",
        "Nadeem asked whether AHP funding can be used as SHDF match (answer: no, own reserves only).",
    ),
    (
        "09/12/2024",
        "Ashworth Housing Trust",
        "Nadeem Hussain",
        "Outbound",
        "SHDF match funding clarification",
        "Clarified: match must be own resources. Boiler Upgrade Scheme can supplement for heat pumps.",
    ),
    (
        "20/01/2025",
        "Ironbridge Homes",
        "Samira Begum",
        "Inbound",
        "Awaab's Law compliance query",
        "Samira asked about D&M response time targets. Current 5 days may not be fast enough.",
    ),
    (
        "21/01/2025",
        "Ironbridge Homes",
        "Samira Begum",
        "Outbound",
        "D&M response time benchmarks",
        "Shared club range: Northfield 48hrs (best), Lakeside 72hrs, most 3-5 days. Recommended 72hr target.",
    ),
    (
        "10/02/2025",
        "Riverview Estates",
        "Helen Foster",
        "Outbound",
        "Checking in — attendance",
        "Reached out about missed sessions. Helen stretched with D&M caseload (42/1,000 homes).",
    ),
    (
        "12/02/2025",
        "Riverview Estates",
        "Helen Foster",
        "Inbound",
        "RE: Checking in",
        "Helen confirmed team stretched. Agreed to 1:1 on 25th Feb. Requested session notes.",
    ),
    (
        "18/02/2025",
        "Summit Housing Partnership",
        "Darren Walsh",
        "Inbound",
        "Tenant liaison officer JD request",
        "Darren looking for example JDs after hearing Riverside and Lakeside success with liaison officers.",
    ),
    (
        "06/03/2025",
        "Greendale Homes",
        "Mark Jennings",
        "Inbound",
        "PV procurement framework details",
        "Mark shared Solarfix framework details: £5.4m/3yrs, 12-week procurement, contact James Wright.",
    ),
    (
        "07/03/2025",
        "Northfield Housing Association",
        "Fiona Clarke",
        "Inbound",
        "Battery storage cost-benefit analysis",
        "Fiona shared battery storage analysis: £1,800 extra, £150/yr additional saving, 10-12yr payback.",
    ),
    (
        "24/02/2025",
        "Internal",
        "Rob Bryan",
        "Inbound",
        "New member enquiry — Pinnacle Living",
        "Rob spoke with Sarah Thompson at Pinnacle Living (Doncaster, 6,800 homes) interested in Carbon Club.",
    ),
    (
        "25/02/2025",
        "Internal",
        "Steph Hosny",
        "Outbound",
        "RE: Pinnacle Living follow-up",
        "Confirmed will reach out to Sarah Thompson today with membership info.",
    ),
    (
        "28/02/2025",
        "Internal",
        "Steph Hosny",
        "Outbound",
        "Carbon Club — March session agenda",
        "Sent draft agenda to Rob. Flagged Riverview non-attendance and April renewal risk.",
    ),
    (
        "03/03/2025",
        "Internal",
        "Steph Hosny",
        "Outbound",
        "Broadland renewal brief preparation",
        "Notified Rob about Broadland renewal 1 April. Total £8,500/yr. Strong engagement, PV interest is retention hook.",
    ),
    (
        "08/02/2025",
        "Harrowfield Homes",
        "Joanna Briggs",
        "Inbound",
        "D&M training programme materials",
        "Joanna shared half-day D&M assessment training materials developed with surveying firm.",
    ),
    (
        "04/03/2025",
        "Foxley Homes",
        "Danielle Webb",
        "Inbound",
        "SHDF guidance request",
        "Danielle asked for SHDF Wave 2 guidance to be shared with whole group. Interest in bid preparation sessions.",
    ),
]


def _gen_email_log():
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Interaction Log"
    headers = [
        "Date",
        "Organisation",
        "Contact",
        "Direction",
        "Subject",
        "Summary",
        "Vantage Staff",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for row in _INTERACTIONS:
        staff = "Steph Hosny"
        if "Rob" in row[2] or "Internal" in row[1]:
            staff = "Steph Hosny" if row[3] == "Outbound" else "Rob Bryan"
        ws.append(list(row) + [staff])
    _auto_width(ws)
    return _wb_bytes(wb)


# ── Workshop Transcripts ──────────────────────────────────────────────────


def _gen_transcript_feb():
    return """Carbon Club Workshop — 14 February 2025
Participants: Steph Hosny (Vantage), Rob Bryan (Vantage), Mark Jennings (Greendale Homes),
              Fiona Clarke (Northfield HA), David Osei (Beacon Dwellings),
              Lena Morris (Broadland Housing), Karen Blackwell (Westmoor Housing Group),
              James Thornton (Riverside Community Housing), Priya Sharma (Oaktree Living),
              Andrew Marsh (Severn Vale Homes), Oliver Grant (Thameside Housing Trust),
              Claire Whitfield (Pennine Valleys Housing), Chris Doyle (Lakeside Living),
              Sarah Linehan (Meridian Housing Group), Nadeem Hussain (Ashworth Housing Trust),
              Darren Walsh (Summit Housing Partnership), Samira Begum (Ironbridge Homes),
              Joanna Briggs (Harrowfield Homes), Danielle Webb (Foxley Homes),
              Mei-Lin Chen (Wychwood HA), Wayne Kirkpatrick (Stonebridge Housing Group)

[00:00:15] Steph Hosny: Welcome everyone to the February Carbon Club session. Today we have two spotlights — Mark from Greendale is going to update us on their solar PV programme, and Fiona from Northfield will talk about their experience with battery storage. Then we'll do a roundtable on EPC data quality.

[00:02:30] Steph Hosny: Mark, do you want to kick us off?

[00:02:45] Mark Jennings: Sure, thanks Steph. So as most of you know, we started our PV programme back in late 2023. We've now completed 1,200 installations across our general needs stock — mainly three-bed semis and end-terraces which have the best roof orientation. We've been using 4kW systems through the Solarfix framework, accessed via Fusion21.

[00:05:10] Mark Jennings: The average cost has come in at about £4,200 per property including scaffolding, which is slightly below our original budget of £4,500. Tenant feedback has been really positive — we're seeing about 89% satisfaction in our post-install surveys. The average tenant is saving around £180 per year on electricity.

[00:07:30] Mark Jennings: One thing I'd flag for anyone starting a programme — talk to your DNO early. We had some grid connection delays with Western Power Distribution that held us up for about six weeks on one estate. Batch applications help but you need to submit well in advance.

[00:09:15] Lena Morris: Mark, what was your tenant refusal rate? We're a bit worried about that.

[00:09:30] Mark Jennings: About 8% overall. Mostly elderly residents who were concerned about disruption. We found that face-to-face home visits brought that down significantly compared to just sending letters.

[00:10:45] Steph Hosny: Thanks Mark. That's really useful. We'll circulate the Solarfix framework details after the session. Fiona, over to you.

[00:11:00] Fiona Clarke: Thanks Steph. So Northfield took a slightly different approach — we went with PV plus battery storage from day one. We've completed 850 installations across three estates in south Birmingham — Selly Oak, Kings Heath, and Moseley Gardens.

[00:13:20] Fiona Clarke: The battery adds about £1,800 per property on top of the PV cost, bringing our total to around £6,200 per unit. But we're seeing tenants save an additional £150 per year compared to PV-only installations. We used GivEnergy 5.2kWh batteries with a 12-year warranty. Failure rate has been less than 1% — only 6 out of 850.

[00:16:00] Oliver Grant: Fiona, what's the payback period on the battery component specifically?

[00:16:15] Fiona Clarke: We estimate 10 to 12 years on the battery alone. So it's marginal financially, but the additional tenant benefit and the energy resilience argument helped us get board approval.

[00:18:00] David Osei: For Beacon, we're at 340 installations now as part of our wider EPC programme. We're using the Fusion21 framework as well, coming in at about £4,400 per unit. We're targeting 500 by end of March. Our focus has been getting everything up to EPC C first, so PV is one component alongside cavity wall insulation and new windows.

[00:20:30] Steph Hosny: Good progress David. Now let's move on to EPC data quality — this has been a hot topic. James, you raised this at the last session. Want to kick off the discussion?

[00:20:50] James Thornton: Yes — we completed our stock condition survey and found about a 20% discrepancy between our EPC records and what's actually in the properties. That's significant when you're trying to target retrofit programmes.

[00:22:15] Mark Jennings: We're seeing similar issues. We're working with Elmhurst Energy to resurvey about 2,000 properties. Initial estimate is 15% have inaccurate EPCs, mainly pre-1940 solid wall stock where internal insulation was done but the EPC wasn't updated.

[00:24:00] Sarah Linehan: We took a different approach at Meridian — we built our own asset database that tracks installed measures against each property and calculates what the EPC should be. It's Excel-based but gives us a much more accurate picture than relying on the EPC register.

[00:25:30] David Osei: We found about 25% inaccuracy in our solid wall stock. Real problem when you're targeting properties below EPC C. We've budgeted for a full resurvey in 2025/26.

[00:27:00] Oliver Grant: The regulatory implications worry me. If we're reporting 75% at EPC C but the reality is 65%, that's a governance issue under the new consumer standards.

[00:29:00] Steph Hosny: Good points. I think we need a dedicated session on this — maybe we can get someone from the regulator to come and talk about expectations. I'll look into that for a future session.

[00:30:15] Claire Whitfield: Quick update on SHDF — for those thinking about Wave 2.2, the application window opens 1 April 2025 and closes 30 June. Start pulling your property-level data together now. It took us about six weeks to put our Wave 2.1 application together and we got £1.2 million for 450 measures across 280 properties.

[00:32:00] Karen Blackwell: Can I also raise tenant communications? We're about to start a major insulation programme and I'd appreciate any examples of good tenant engagement approaches.

[00:32:30] James Thornton: Happy to share our template letters Karen — we've got them for PV, insulation, windows, and heat pumps. The key is being specific about savings figures. Happy for Steph to circulate.

[00:34:00] Darren Walsh: We recruited a dedicated tenant liaison officer and it's been transformative. Single point of contact through the whole process. Tenants actually thank us now.

[00:35:30] Steph Hosny: Thanks everyone. Let me summarise the actions. I'll circulate Greendale's PV procurement framework details. Fiona will share the battery storage cost-benefit analysis. I'll connect Lena with Mark for a peer visit. And we'll set up an EPC data quality focused session for later in the spring. See you all next month.
"""


def _gen_transcript_mar():
    return """Carbon Club Workshop — 5 March 2025
Participants: Steph Hosny (Vantage), Mark Jennings (Greendale Homes),
              Fiona Clarke (Northfield HA), David Osei (Beacon Dwellings),
              Lena Morris (Broadland Housing), Karen Blackwell (Westmoor Housing Group),
              James Thornton (Riverside Community Housing), Priya Sharma (Oaktree Living),
              Tom Henderson (Millbrook Homes), Andrew Marsh (Severn Vale Homes),
              Oliver Grant (Thameside Housing Trust), Claire Whitfield (Pennine Valleys Housing),
              Chris Doyle (Lakeside Living), Sarah Linehan (Meridian Housing Group),
              Nadeem Hussain (Ashworth Housing Trust), Darren Walsh (Summit Housing Partnership),
              Samira Begum (Ironbridge Homes), Danielle Webb (Foxley Homes)

[00:00:12] Steph Hosny: Good morning everyone, welcome to the March Carbon Club session. Quick housekeeping — I notice Riverview Estates aren't with us again today — that's the fourth session in a row now. I'll be reaching out to Helen after this. Also, I'd like to welcome Tom Henderson from Millbrook Homes — Tom is usually in the Executive Club but he's joining Carbon Club today.

[00:01:30] Tom Henderson: Thanks Steph. Yes, this is my first Carbon Club session — I'm usually in the Exec Club but given we're ramping up our decarbonisation programme, I thought I should start attending. Looking forward to it.

[00:02:15] Steph Hosny: Great to have you Tom. Right, today's agenda — we'll start with a follow-up from last month on PV procurement, then SHDF Wave 2 update from Claire, and we'll finish with a discussion on tenant communications for retrofit works.

[00:03:00] Mark Jennings: Just a quick follow-up from February — I sent Steph the Solarfix framework details as promised. Key points: framework operator is Solarfix Ltd accessed via Fusion21, contract value was £5.4 million over three years, procurement timeline about 12 weeks. Contact at Solarfix is James Wright. Happy for anyone to reach out.

[00:04:45] Fiona Clarke: And I've shared our battery storage cost-benefit analysis with Steph. Key finding — the additional £1,800 per property for the battery gives tenants an extra £150 per year saving. Payback period is 10 to 12 years. We used GivEnergy 5.2kWh units with a 12-year warranty.

[00:06:30] David Osei: Update from Beacon — we hit 420 PV installs by end of January. On track for 500 by end of March. Costs holding steady at about £4,400 per unit through the Fusion21 framework. We've also started our EPC resurvey — first batch of 300 properties done, confirming about 22% inaccuracy rate.

[00:08:15] Claire Whitfield: SHDF update — for those planning Wave 2.2 applications, the window opens 1 April and closes 30 June 2025. I'd really recommend starting now. You need property-level EPC data, contractor quotes or at least indicative pricing, and a tenant engagement plan. It took us six weeks and we got £1.2 million.

[00:10:00] Nadeem Hussain: Claire, on the match funding — does it have to be from our own reserves? Our finance director is asking about using Affordable Homes Programme money.

[00:10:20] Claire Whitfield: Has to be your own resources or private borrowing. You can't use other government grants as match. But you can use the Boiler Upgrade Scheme alongside SHDF for heat pump installations — they're treated as separate funding pots.

[00:12:00] Andrew Marsh: I wanted to offer to do a spotlight on our whole-house retrofit programme at a future session. We've done 200 properties with insulation, windows, and heat pumps — EPC E to B in most cases. I could also cover the in-house DLO approach since several people have been asking about it.

[00:13:30] Priya Sharma: Yes please Andrew — we're seriously looking at setting up an in-house void team. Your experience with the setup costs and break-even timeline would be really helpful.

[00:15:00] Karen Blackwell: Following up on the tenant comms discussion from last month — we've started our insulation programme and we're using face-to-face visits for elderly tenants. Early results are promising — refusal rate dropped from 15% to under 5% with home visits. More resource-intensive but definitely worth it.

[00:16:45] James Thornton: Our stock condition survey flagged 85 properties as potential damp and mould risks. We're prioritising those for the retrofit programme. Also identified 340 properties needing urgent window replacement.

[00:18:30] Chris Doyle: Quick plug for environmental monitoring sensors — we've now got 500 deployed and they're flagging high humidity before mould develops. Reduced reactive callouts by 30%. If anyone wants the supplier details — Switchee, about £80 per unit including installation, free pilot of 50 units available.

[00:20:00] Samira Begum: With Awaab's Law now in force, what response times are people targeting for damp and mould? We're at five working days for initial inspection and I'm not sure that's fast enough.

[00:20:30] Fiona Clarke: We've moved to 48 hours. Triaged into emergency, urgent, and routine. It's been effective — open cases down from 180 to 45 in six months.

[00:22:00] Steph Hosny: Right, let me wrap up with actions. Mark's PV framework details are going out to everyone today. Fiona's battery analysis likewise. I'll introduce Lena to Mark for a peer visit. Everyone thinking about SHDF — start your applications now. Andrew, I'll schedule your spotlight for April or May. And I'll submit the EPC data quality survey to everyone by end of March. Thanks all — see you next month.
"""


def _gen_participants():
    return """Meeting: Carbon Club Workshop — 5 March 2025
Duration: 10:00 – 11:15 GMT

Participants:
- Steph Hosny (Vantage) — joined 09:58, left 11:15
- Mark Jennings (Greendale Homes) — joined 10:01, left 11:14
- Fiona Clarke (Northfield HA) — joined 10:00, left 11:15
- David Osei (Beacon Dwellings) — joined 10:02, left 11:12
- Lena Morris (Broadland Housing) — joined 10:05, left 10:58
- Karen Blackwell (Westmoor Housing Group) — joined 10:00, left 11:15
- James Thornton (Riverside Community Housing) — joined 10:08, left 11:15
- Priya Sharma (Oaktree Living) — joined 10:00, left 11:10
- Tom Henderson (Millbrook Homes) — joined 10:03, left 11:15
- Andrew Marsh (Severn Vale Homes) — joined 10:01, left 11:14
- Oliver Grant (Thameside Housing Trust) — joined 10:04, left 11:13
- Claire Whitfield (Pennine Valleys Housing) — joined 10:00, left 11:15
- Chris Doyle (Lakeside Living) — joined 10:02, left 11:12
- Sarah Linehan (Meridian Housing Group) — joined 10:06, left 11:00
- Nadeem Hussain (Ashworth Housing Trust) — joined 10:01, left 11:15
- Darren Walsh (Summit Housing Partnership) — joined 10:00, left 11:14
- Samira Begum (Ironbridge Homes) — joined 10:03, left 11:15
- Danielle Webb (Foxley Homes) — joined 10:07, left 11:10
"""


# ── Case Study Text Files ─────────────────────────────────────────────────


def _gen_case_study_greendale():
    return """GREENDALE HOMES — Solar PV Programme Case Study
===============================================================

Organisation: Greendale Homes
Location: Nottingham
Stock Size: 12,500 homes
Date Published: November 2024

BACKGROUND
----------
Greendale Homes began exploring rooftop solar PV in 2023 as part of their
wider decarbonisation strategy. With 68% of stock at EPC C or above and a
target of 80% by 2027, PV was identified as a cost-effective measure for
improving energy performance while directly reducing tenant energy bills.

APPROACH
--------
- Procurement: Solarfix Ltd framework via Fusion21
- System: 4kW rooftop PV on suitable properties (south/west-facing roofs)
- Target stock: 3-bed semi-detached and end-terrace properties
- Installation timeline: November 2023 – ongoing (1,200 completed by Jan 2025)
- Tenant engagement: Face-to-face home visits for elderly tenants, letters
  with specific savings figures for all others

KEY METRICS
-----------
Installations completed:      1,200
Cost per property:            £4,200 (including scaffolding)
Average generation:           3,800 kWh per property per year
Average tenant saving:        £180 per year on electricity
Tenant satisfaction:          89%
Tenant refusal rate:          8%
System reliability:           96% (48 faults across 1,200 installations)
Procurement timeline:         12 weeks from market to appointment

RESULTS
-------
- EPC improvement: Average uplift of 5-8 SAP points per property
- 94% of installations completed within the planned 2-day window
- Grid connection: Some delays with Western Power Distribution (batch
  applications recommended — submit 6-8 weeks before installation start)

LESSONS LEARNED
---------------
1. Engage with DNO early — grid connection capacity can be a bottleneck
2. Face-to-face visits reduce refusal rates significantly (8% vs 15% with
   letters only)
3. Quality assurance: 10% sample inspection — 96% pass rate
4. Scaffold sharing between adjacent properties reduces costs by ~£200/unit
5. Winter installations take 20% longer due to weather delays

CONTACT
-------
Mark Jennings, Head of Sustainability
m.jennings@greendalehousing.org.uk
"""


def _gen_case_study_northfield():
    return """NORTHFIELD HOUSING ASSOCIATION — Solar PV with Battery Storage Case Study
===============================================================

Organisation: Northfield Housing Association
Location: Birmingham
Stock Size: 8,200 homes
Date Published: January 2025

BACKGROUND
----------
Northfield HA decided to pair rooftop solar PV with battery storage from
the outset of their programme, believing the additional investment would
deliver significantly better outcomes for tenants. The programme focused on
three estates in south Birmingham with high proportions of fuel-poor tenants.

APPROACH
--------
- System: 4kW solar PV + GivEnergy 5.2kWh battery storage
- Target estates: Selly Oak, Kings Heath, Moseley Gardens
- Installation timeline: March 2024 – ongoing (850 completed by Jan 2025)
- Funding: Combination of SHDF Wave 2 grant and own capital investment
- Tenant engagement: Pre-installation education sessions (1-hour home visit),
  post-installation follow-up at 2 weeks

KEY METRICS
-----------
Installations completed:      850
Total cost per property:      £6,200
  — PV component:             £4,400
  — Battery component:        £1,800
Annual tenant saving (PV only): £180
Annual tenant saving (PV+battery): £330 (additional £150 from battery)
Battery payback period:       10-12 years
Battery warranty:             12 years (GivEnergy)
Battery failure rate:         <1% (6 faults out of 850)
Tenant satisfaction:          92%

RESULTS
-------
- Tenants with batteries save an additional £150/year compared to PV-only
- Battery allows tenants to use stored solar energy in evenings when
  electricity tariffs are highest
- 92% tenant satisfaction — higher than PV-only programmes (typically 85-89%)
- Properties improved by 8-12 SAP points on average

LESSONS LEARNED
---------------
1. Battery storage improves the tenant proposition significantly, especially
   for households with evening-heavy usage patterns
2. GivEnergy 5.2kWh unit is well-suited to social housing — compact, quiet,
   12-year warranty matches expected lifecycle
3. Pre-installation education sessions are essential — tenants need to
   understand how to maximise self-consumption
4. Follow-up visits at 2 weeks catch early issues before complaints develop
5. Combined PV+battery installation takes 2.5 days vs 1.5 for PV only

COST-BENEFIT ANALYSIS
----------------------
The battery adds £1,800 to the installation cost. At £150/year additional
saving, the simple payback is 12 years. However, considering:
- Rising electricity prices (5% annual increase assumption)
- Avoided grid export losses
- Tenant satisfaction uplift
- Future smart grid revenue potential
The adjusted payback reduces to approximately 10 years.

CONTACT
-------
Fiona Clarke, Director of Assets
f.clarke@northfieldhousing.org.uk
"""


# ── Void Standards Portal Export ──────────────────────────────────────────


def _gen_void_standards_export():
    return """Void Standards and Turnaround Thresholds — Portal Discussion Export
Exported from Vantage Connect Portal, November 2024

Thread: "Void Standards and Turnaround Thresholds"
Posted: 22 November 2024

[Original Post — Karen Blackwell, Westmoor Housing Group]
Does anyone have documented void standards they'd be willing to share?
We're reviewing our lettable standard and trying to benchmark our turnaround
times. Our board is pushing hard on this — we're at 28 days average which
I know is above what some of you are achieving. What targets are others
working to? And what lettable standards do you use — Basic, Standard, or
Enhanced?

    [Reply — Fiona Clarke, Northfield HA — 23 Nov 2024]
    Our target is 18 calendar days key-to-key for standard voids. We're
    currently achieving 16 days on average. We use an Enhanced lettable
    standard which includes full redecoration and new flooring where needed.
    The key change for us was introducing pre-void inspections 4 weeks
    before the tenancy ends — lets us plan the works schedule in advance
    and order materials early.

    [Reply — James Thornton, Riverside Community Housing — 24 Nov 2024]
    We're targeting 16 days but our actual is 22 at the moment — so we've
    got work to do. We use a Basic standard (safety checks plus clean).
    We're piloting 'void packs' with new tenant welcome kits which include
    cleaning supplies, a small toolkit, and a guide to the property. Early
    feedback from tenants has been positive.

    [Reply — Mark Jennings, Greendale Homes — 25 Nov 2024]
    We target 20 days and we're at 19 on average. Standard lettable standard
    — safety checks plus essential repairs. We set up a dedicated void
    contractor team on a 5-day SLA for each property, which has helped
    enormously with consistency. Having a single team who only do void
    works means they know exactly what's expected.

    [Reply — David Osei, Beacon Dwellings — 25 Nov 2024]
    Target is 15 days, actual is 21. We use an Enhanced standard. We've
    been using the Plentific platform for void works scheduling which has
    helped with visibility and contractor management, but we're still
    not hitting target consistently. Contractor availability is the
    main bottleneck.

    [Reply — Priya Sharma, Oaktree Living — 26 Nov 2024]
    We're at 25 days average with a 20-day target and a Basic standard.
    Really struggling with contractor availability — it's our biggest
    challenge. We're seriously exploring setting up an in-house team
    after hearing about Severn Vale's experience.

    [Reply — Tom Henderson, Millbrook Homes — 27 Nov 2024]
    18-day target, 17 days actual. Standard lettable standard. The key
    for us was integrating void and lettings into a single team — one
    team handles everything from the pre-void inspection through to
    the new tenant sign-up. Eliminated a 3-4 day handover gap we used
    to have between void works completion and lettings processing.

    [Reply — Claire Whitfield, Pennine Valleys Housing — 28 Nov 2024]
    We recently tightened from a 28-day target to 22 days after board
    scrutiny. Currently at 24 days with a Basic standard. It's been a
    challenge but we're making progress. The board presentation on void
    costs (£8,400 per void period for properties taking >30 days) was
    the catalyst for change.

    [Reply — Andrew Marsh, Severn Vale Homes — 29 Nov 2024]
    We target 16 days and we're consistently at 15. Enhanced lettable
    standard. We attribute our performance almost entirely to having an
    in-house DLO (Direct Labour Organisation). Setup cost was about £185k
    — 8 operatives, vans, tools. Broke even within 14 months through
    reduced contractor costs and dramatically lower rent loss. Our void
    turnaround went from 24 days to 15 days after setting up the DLO.
    Happy to share more detail with anyone considering this route.
"""


# ── Upload helper ─────────────────────────────────────────────────────────

_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_TXT_CT = "text/plain"


def _upload_to_onedrive(path, data, content_type=_TXT_CT):
    url = f"{BASE}/me/drive/root:/{path}:/content"
    headers = {**H_AUTH, "Content-Type": content_type}
    r = requests.put(url, headers=headers, data=data)
    return r.status_code in (200, 201)


# ── Seed OneDrive ─────────────────────────────────────────────────────────


def seed_onedrive():
    print(f"\nSeeding OneDrive files...")
    files = [
        ("Vantage/Benchmarking Data/vfm_financial_data.xlsx", _gen_vfm(), _XLSX_CT),
        ("Vantage/Benchmarking Data/tsm_satisfaction_data.xlsx", _gen_tsm(), _XLSX_CT),
        (
            "Vantage/Benchmarking Data/operational_kpi_data.xlsx",
            _gen_ops_kpi(),
            _XLSX_CT,
        ),
        (
            "Vantage/Carbon Club/1-to-1 Notes/carbon_club_1to1_notes.xlsx",
            _gen_1to1(),
            _XLSX_CT,
        ),
        ("Vantage/Membership/membership_master.xlsx", _gen_membership(), _XLSX_CT),
        (
            "Vantage/Carbon Club/Attendance/attendance_log.xlsx",
            _gen_attendance(),
            _XLSX_CT,
        ),
        (
            "Vantage/Email Interaction Log/email_interaction_log.xlsx",
            _gen_email_log(),
            _XLSX_CT,
        ),
        (
            "Vantage/Carbon Club/Workshop Transcripts/carbon_club_workshop_14feb2025.txt",
            _gen_transcript_feb().encode("utf-8"),
            _TXT_CT,
        ),
        (
            "Vantage/Carbon Club/Workshop Transcripts/carbon_club_workshop_5mar2025.txt",
            _gen_transcript_mar().encode("utf-8"),
            _TXT_CT,
        ),
        (
            "Vantage/Carbon Club/Teams Participant Lists/teams_participants_5mar2025.txt",
            _gen_participants().encode("utf-8"),
            _TXT_CT,
        ),
        (
            "Vantage/Knowledge Hub/greendale_pv_case_study.txt",
            _gen_case_study_greendale().encode("utf-8"),
            _TXT_CT,
        ),
        (
            "Vantage/Knowledge Hub/northfield_pv_battery_case_study.txt",
            _gen_case_study_northfield().encode("utf-8"),
            _TXT_CT,
        ),
        (
            "Vantage/Portal Exports/void_standards_portal_thread.txt",
            _gen_void_standards_export().encode("utf-8"),
            _TXT_CT,
        ),
    ]
    success = 0
    for path, data, ct in files:
        if _upload_to_onedrive(path, data, ct):
            success += 1
            print(f"  [OK] {path}")
        else:
            print(f"  [FAIL] {path}")
        time.sleep(0.5)
    print(f"  {success}/{len(files)} files uploaded.")


# =========================================================================
# MAIN
# =========================================================================


def _resolve_token(args):
    try:
        return resolve_access_token(token_file=args.token_file)
    except FileNotFoundError as exc:
        print(f"[ERROR] {exc}")
        sys.exit(1)
    except Exception as exc:
        print(f"[ERROR] Could not get access token: {exc}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Seed Vantage demo data into M365")
    parser.add_argument(
        "--wipe",
        action="store_true",
        help="Wipe all existing data before seeding",
    )
    parser.add_argument(
        "--token-file",
        type=str,
        help="Path to file containing the Graph API token",
    )
    args = parser.parse_args()

    token = _resolve_token(args)
    _init_headers(token)

    r = requests.get(f"{BASE}/me", headers=H_AUTH)
    if r.status_code != 200:
        print(f"[ERROR] Token validation failed ({r.status_code}). Get a fresh token.")
        sys.exit(1)
    print(
        f"Authenticated as: {r.json()['displayName']} ({r.json()['userPrincipalName']})",
    )

    print("=" * 60)
    print(f"VANTAGE DEMO — {'WIPE & ' if args.wipe else ''}SEED")
    print("=" * 60)

    if args.wipe:
        wipe_emails()
        wipe_onenote()
        wipe_onedrive()

    seed_emails()
    seed_onenote()
    seed_onedrive()

    print("\n" + "=" * 60)
    n_drive = 13
    print(
        f"DONE \u2014 {len(EMAILS)} emails, {len(RENEWAL_NOTES)} OneNote pages, {n_drive} OneDrive files",
    )
    print("=" * 60)


if __name__ == "__main__":
    main()
