"""Seed Midland Heart data into ``MidlandHeart/`` contexts via ``dm.ingest()``.

Reads two Excel workbooks and ingests their sheets as first-class data
contexts, simulating API-sourced data that lives outside the FileManager's
file lifecycle.

Usage::

    from unity.data_manager.data_manager import DataManager
    from unity.customization.clients.midland_heart.seed_data import seed_all

    dm = DataManager()
    results = seed_all(dm)
    for ctx, result in results.items():
        print(f"{ctx}: {result.rows_inserted} rows")

Excel sources (under ``intranet/repairs/``):

- ``MDH Repairs Data July - Nov 25 - DL V1.xlsx``
  → ``MidlandHeart/Repairs2025`` (sheet: ``Raised 01-07-2025 to 30-11-2025``)

- ``MDH Telematics Data July - Nov 25 - DL V1.xlsx``
  → ``MidlandHeart/Telematics2025/July`` … ``/November`` (one sheet per month)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from unity.data_manager.data_manager import DataManager
    from unity.data_manager.types.ingest import IngestResult

logger = logging.getLogger(__name__)

_INTRANET_DIR = Path(__file__).resolve().parents[3] / "intranet" / "repairs"

REPAIRS_FILE = "MDH Repairs Data July - Nov 25 - DL V1.xlsx"
REPAIRS_SHEET = "Raised 01-07-2025 to 30-11-2025"
REPAIRS_CONTEXT = "MidlandHeart/Repairs2025"
REPAIRS_DESCRIPTION = (
    "Reactive repairs raised July–November 2025. One row per job ticket line. "
    "Key columns: JobTicketReference, WorksOrderRef, WorksOrderStatusDescription, "
    "WorksOrderRaisedDate, WorksOrderPriorityDescription, OperativeName, "
    "PropertyReference, FullAddress, NoAccess, FollowOn, FirstTimeFix, "
    "ScheduledAppointmentStart, ArrivedOnSite, CompletedVisit."
)

TELEMATICS_FILE = "MDH Telematics Data July - Nov 25 - DL V1.xlsx"
TELEMATICS_MONTHS = [
    "July 2025",
    "August 2025",
    "September 2025",
    "October 2025",
    "November 2025",
]
TELEMATICS_CONTEXT_PREFIX = "MidlandHeart/Telematics2025"
TELEMATICS_DESCRIPTION = (
    "Monthly vehicle telematics data. One row per trip/sub-trip. "
    "Key columns: Trip, Driver, Vehicle, Departure, Arrival, StartLocation, "
    "EndLocation, Total distance, Avg speed, Max speed, Idling time."
)


# ---------------------------------------------------------------------------
# Excel reading helpers
# ---------------------------------------------------------------------------


def _read_xlsx_sheet(
    path: Path,
    sheet_name: str,
) -> List[Dict[str, Any]]:
    """Read an Excel sheet into a list of row dicts.

    Uses ``openpyxl`` in read-only mode for speed on large workbooks.
    The first row is treated as headers; subsequent rows become dicts.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for seed_data.  Install with: pip install openpyxl",
        ) from e

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
    except KeyError:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {path.name}. Available: {available}",
        )

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or f"col_{i}").strip() for i, h in enumerate(next(rows_iter))]

    data: List[Dict[str, Any]] = []
    for raw_row in rows_iter:
        row_dict = {}
        for col_name, value in zip(headers, raw_row):
            if value is not None:
                row_dict[col_name] = value
        if row_dict:
            data.append(row_dict)

    wb.close()
    return data


# ---------------------------------------------------------------------------
# Ingestion functions
# ---------------------------------------------------------------------------


def seed_repairs(
    dm: "DataManager",
    *,
    excel_dir: Optional[Path] = None,
    embed_columns: Optional[List[str]] = None,
) -> "IngestResult":
    """Ingest the repairs workbook into ``MidlandHeart/Repairs2025``.

    Parameters
    ----------
    dm : DataManager
        Active DataManager instance.
    excel_dir : Path | None
        Directory containing the Excel files.  Defaults to ``intranet/repairs/``.
    embed_columns : list[str] | None
        Columns to embed.  ``None`` skips embedding.

    Returns
    -------
    IngestResult
    """
    base = excel_dir or _INTRANET_DIR
    path = base / REPAIRS_FILE

    logger.info("Reading repairs data from %s (sheet: %s)", path.name, REPAIRS_SHEET)
    rows = _read_xlsx_sheet(path, REPAIRS_SHEET)
    logger.info("Read %d repairs rows", len(rows))

    return dm.ingest(
        REPAIRS_CONTEXT,
        rows,
        description=REPAIRS_DESCRIPTION,
        embed_columns=embed_columns,
        embed_strategy="along" if embed_columns else "off",
        chunk_size=1000,
        unique_keys={"JobTicketReference": "str"},
    )


def seed_telematics(
    dm: "DataManager",
    *,
    excel_dir: Optional[Path] = None,
    embed_columns: Optional[List[str]] = None,
) -> Dict[str, "IngestResult"]:
    """Ingest the telematics workbook (one context per month).

    Parameters
    ----------
    dm : DataManager
        Active DataManager instance.
    excel_dir : Path | None
        Directory containing the Excel files.
    embed_columns : list[str] | None
        Columns to embed per month.

    Returns
    -------
    dict[str, IngestResult]
        Mapping of ``MidlandHeart/Telematics2025/<Month>`` → result.
    """
    base = excel_dir or _INTRANET_DIR
    path = base / TELEMATICS_FILE

    results: Dict[str, "IngestResult"] = {}

    for sheet_name in TELEMATICS_MONTHS:
        month_label = sheet_name.split()[0]  # "July 2025" → "July"
        context = f"{TELEMATICS_CONTEXT_PREFIX}/{month_label}"
        logger.info("Reading telematics %s from %s", sheet_name, path.name)

        try:
            rows = _read_xlsx_sheet(path, sheet_name)
        except (ValueError, KeyError) as e:
            logger.warning("Skipping telematics sheet %s: %s", sheet_name, e)
            continue

        logger.info("Read %d telematics rows for %s", len(rows), sheet_name)

        result = dm.ingest(
            context,
            rows,
            description=f"{TELEMATICS_DESCRIPTION} Month: {sheet_name}.",
            embed_columns=embed_columns,
            embed_strategy="along" if embed_columns else "off",
            chunk_size=500,
        )
        results[context] = result

    return results


def seed_all(
    dm: "DataManager",
    *,
    excel_dir: Optional[Path] = None,
    embed_columns_repairs: Optional[List[str]] = None,
    embed_columns_telematics: Optional[List[str]] = None,
) -> Dict[str, "IngestResult"]:
    """Seed all Midland Heart data contexts.

    Returns
    -------
    dict[str, IngestResult]
        Mapping of context path → ingest result for every seeded context.
    """
    results: Dict[str, "IngestResult"] = {}

    repairs_result = seed_repairs(
        dm,
        excel_dir=excel_dir,
        embed_columns=embed_columns_repairs,
    )
    results[REPAIRS_CONTEXT] = repairs_result

    telem_results = seed_telematics(
        dm,
        excel_dir=excel_dir,
        embed_columns=embed_columns_telematics,
    )
    results.update(telem_results)

    total_rows = sum(r.rows_inserted for r in results.values())
    logger.info(
        "Midland Heart seed complete: %d contexts, %d total rows",
        len(results),
        total_rows,
    )
    return results
