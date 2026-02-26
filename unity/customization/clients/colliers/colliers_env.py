"""Colliers environment: domain-specific tools for financial data extraction
and web deal research.
"""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from unity.customization.clients.colliers.colliers_schemas import (
    CATEGORY_TO_SECTION,
    FIELD_NAME_MAPPING,
    FINANCIAL_DATA_SCHEMA,
    DealRow,
    FiscalYearData,
)

from unity.actor.environments.base import BaseEnvironment, ToolMetadata


class ColliersService:
    """Domain-specific tools for Colliers deliverables."""

    def create_financial_data_excel(
        self,
        json_file_path: str,
        output_path: str,
    ) -> str:
        """Read validated financial JSON and produce a formatted HISTORIC ACCOUNTS Excel.

        Parameters
        ----------
        json_file_path : str
            Path to a JSON file containing ``list[FiscalYearData]``.
        output_path : str
            Where to write the ``.xlsx`` output.

        Returns
        -------
        str
            The path to the created Excel file.
        """
        raw = json.loads(Path(json_file_path).read_text(encoding="utf-8"))
        fiscal_data = [FiscalYearData.model_validate(item) for item in raw]

        properties: dict[str, list[FiscalYearData]] = defaultdict(list)
        for item in fiscal_data:
            properties[item.property_name].append(item)
        for prop_name in properties:
            properties[prop_name].sort(key=lambda x: x.fiscal_year)

        wb = Workbook()
        wb.remove(wb.active)

        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_font_white = Font(bold=True, size=10, color="FFFFFF")
        data_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid",
        )
        category_font = Font(bold=True, size=10)
        italic_font = Font(italic=True, size=10)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for prop_name, fiscal_years in properties.items():
            safe_name = prop_name[:31]
            for ch in "/\\*?[]":
                safe_name = safe_name.replace(ch, "-")
            ws = wb.create_sheet(title=safe_name)
            num_years = len(fiscal_years)

            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=num_years + 1,
            )
            ws.cell(row=1, column=1, value="HISTORIC ACCOUNTS").font = title_font
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            for col_idx, fy in enumerate(fiscal_years, start=2):
                cell = ws.cell(
                    row=3,
                    column=col_idx,
                    value=f"Financial Statements to YE\n{fy.fiscal_year}",
                )
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin_border

            current_row = 5
            for category, fields in FINANCIAL_DATA_SCHEMA.items():
                ws.cell(
                    row=current_row,
                    column=1,
                    value=category.replace("_", " "),
                ).font = category_font
                current_row += 1
                for field_name in fields:
                    is_italic = "%" in field_name or "PSUPW" in field_name
                    cell = ws.cell(row=current_row, column=1, value=field_name)
                    if is_italic:
                        cell.font = italic_font
                    pydantic_field = FIELD_NAME_MAPPING.get(
                        field_name,
                        field_name.lower().replace(" ", "_"),
                    )
                    section_name = CATEGORY_TO_SECTION[category]
                    for col_idx, fy in enumerate(fiscal_years, start=2):
                        section_data = getattr(fy, section_name, None)
                        value = None
                        source = None
                        if section_data is not None:
                            field_obj = getattr(section_data, pydantic_field, None)
                            if field_obj is not None and hasattr(field_obj, "value"):
                                value = field_obj.value
                                source = getattr(field_obj, "source", None)
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = thin_border
                        if value is not None:
                            cell.fill = data_fill
                            if isinstance(value, (int, float)) and "%" in field_name:
                                cell.number_format = (
                                    "0.0%" if abs(value) <= 1 else "0.00"
                                )
                            elif isinstance(value, (int, float)) and abs(value) >= 1000:
                                cell.number_format = "#,##0"
                            if source:
                                cell.comment = Comment(source, "Source")
                    current_row += 1
                current_row += 1

            ws.column_dimensions["A"].width = 35
            for col_idx in range(2, num_years + 2):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

        out = Path(output_path)
        wb.save(out)
        return str(out)

    _PRIMARY_COL_COUNT = 18

    _CURRENCY_FIELDS = frozenset(
        {
            "rent",
            "rent_pb",
            "price",
            "cv_pb",
            "quote_price",
            "quote_cv_pb",
            "achieved_vs_quote_gbp",
            "achieved_vs_quote_cv_pb",
        },
    )

    _PERCENT_FIELDS = frozenset({"achieved_vs_quote_pct"})

    _DECIMAL_NUMBER_FIELDS = frozenset(
        {
            "achieved_vs_quote_niy_basis_point",
            "transaction_time_weeks",
            "transaction_time_months",
        },
    )

    def create_web_search_excel(
        self,
        json_file_path: str,
        output_path: str,
    ) -> str:
        """Read validated web-search deal JSON and produce a Colliers-styled Excel.

        Produces a single wide table matching the Colliers deal tracker format
        with a two-tone header row (dark blue for primary columns, dark grey
        for secondary columns), auto-filters, and formatted currency/date cells.

        Parameters
        ----------
        json_file_path : str
            Path to a JSON file containing ``list[DealRow]``.
        output_path : str
            Where to write the ``.xlsx`` output.

        Returns
        -------
        str
            The path to the created Excel file.
        """
        raw = json.loads(Path(json_file_path).read_text(encoding="utf-8"))
        deals = [DealRow.model_validate(item) for item in raw]

        wb = Workbook()
        ws = wb.active
        ws.title = "Deal Tracker"

        field_names = list(DealRow.model_fields.keys())
        headers = [DealRow.model_fields[name].alias or name for name in field_names]
        num_cols = len(headers)

        blue_fill = PatternFill(
            start_color="1F3864",
            end_color="1F3864",
            fill_type="solid",
        )
        grey_fill = PatternFill(
            start_color="404040",
            end_color="404040",
            fill_type="solid",
        )
        header_font = Font(bold=True, size=10, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        data_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = blue_fill if col_idx <= self._PRIMARY_COL_COUNT else grey_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"
        ws.row_dimensions[1].height = 30

        for row_idx, deal in enumerate(deals, start=2):
            data = deal.model_dump()
            for col_idx, fname in enumerate(field_names, start=1):
                val = data.get(fname)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = data_border

                if val is None:
                    continue

                if isinstance(val, Decimal):
                    val = float(val)

                if isinstance(val, date):
                    cell.value = val
                    cell.number_format = "MMM-YY"
                elif fname in self._CURRENCY_FIELDS and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = "#,##0"
                elif fname in self._PERCENT_FIELDS and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = "0.00%"
                elif fname in self._DECIMAL_NUMBER_FIELDS and isinstance(
                    val,
                    (int, float),
                ):
                    cell.value = val
                    cell.number_format = "0.0"
                else:
                    cell.value = val

        for col_idx, header in enumerate(headers, start=1):
            width = max(12, min(35, len(header) + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "C2"

        out = Path(output_path)
        wb.save(out)
        return str(out)


class ColliersEnvironment(BaseEnvironment):
    """Colliers-specific environment providing domain tools and schema docs.

    Namespace: ``colliers``
    """

    NAMESPACE = "colliers"

    def __init__(self) -> None:
        self._service = ColliersService()
        super().__init__(instance=self._service, namespace=self.NAMESPACE)

    def get_tools(self) -> Dict[str, ToolMetadata]:
        return {
            f"{self.NAMESPACE}.create_financial_data_excel": ToolMetadata(
                name=f"{self.NAMESPACE}.create_financial_data_excel",
                is_impure=True,
            ),
            f"{self.NAMESPACE}.create_web_search_excel": ToolMetadata(
                name=f"{self.NAMESPACE}.create_web_search_excel",
                is_impure=True,
            ),
        }

    def get_prompt_context(self) -> str:
        financial_schema = json.dumps(
            FiscalYearData.model_json_schema(),
            indent=2,
        )
        deal_schema = json.dumps(
            DealRow.model_json_schema(),
            indent=2,
        )

        return (
            f"### `{self.NAMESPACE}` — Colliers Domain Tools\n\n"
            "#### Tools\n\n"
            f"**`{self.NAMESPACE}.create_financial_data_excel(json_file_path: str, output_path: str) -> str`**\n"
            "  Read validated financial JSON (list of FiscalYearData) and produce a\n"
            "  formatted HISTORIC ACCOUNTS Excel spreadsheet.\n\n"
            f"**`{self.NAMESPACE}.create_web_search_excel(json_file_path: str, output_path: str) -> str`**\n"
            "  Read validated web-search deal JSON (list of DealRow) and produce a\n"
            "  summary Excel spreadsheet.\n\n"
            "#### Data Schemas\n\n"
            "When extracting data, save your output as a JSON file conforming to\n"
            "the schemas below, then call the appropriate `colliers.*` tool to\n"
            "produce the final Excel deliverable.\n\n"
            "##### FiscalYearData (financial extraction)\n\n"
            f"```json\n{financial_schema}\n```\n\n"
            "Each numeric field is a ``FieldValue`` object: "
            '``{"value": <number|string|null>, "source": "<explanation>"}``.\n'
            "Percentages must be decimal (0.9783 for 97.83%). Use null for missing fields.\n\n"
            "##### DealRow (web deal research)\n\n"
            f"```json\n{deal_schema}\n```\n\n"
            "Currency values are numeric (no currency symbols). Dates are ISO format.\n"
            "Yield fields may be strings (e.g., '6.5-7%').\n"
        )

    async def capture_state(self) -> Dict[str, Any]:
        return {"type": "colliers"}


colliers_env = ColliersEnvironment()
