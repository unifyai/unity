"""Produce a formatted HISTORIC ACCOUNTS Excel workbook from extracted
financial JSON data.  Pre-seeded for Colliers Healthcare valuations.
"""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from unity.function_manager.custom import custom_function

# Section → ordered field names (display labels).
FINANCIAL_DATA_SCHEMA = {
    "INCOME": [
        "No. of Days' Management Accounts",
        "Fee Deflation/Inflation Factor",
        "Average fee per week",
        "Registered places",
        "Operational places",
        "Occupancy %",
        "Average fill",
        "Occupancy assuming 100% singles",
        "Fee income per annum",
        "Other income",
        "Income per annum",
    ],
    "EXPENDITURE": [
        "Employee Wages",
        "Agency",
        "National Insurance",
        "Pension Contribution",
        "Payroll",
        "Payroll %",
        "Payroll PSUPW",
        "Provisions",
        "Provisions PSUPW",
        "Heat & Light",
        "Heat & Light %",
        "Accountancy",
        "Bank Charges",
        "Clinical Waste",
        "Council tax",
        "Gardening Costs",
        "Insurance",
        "Laundry & Cleaning",
        "Advertising",
        "Medical Costs",
        "Motor Costs",
        "Print,Post,Stationery, IT",
        "Professional Fees",
        "Registration",
        "Repairs, Renewal, Maintenance",
        "Resident's Activities",
        "Staff Training and Uniforms",
        "Subscriptions",
        "Telephone",
        "Water rates",
        "Sundries",
    ],
    "SUMMARY_TOTALS": [
        "Total Expenses",
        "Total expenses PSUPW",
        "Non Payroll",
        "Non Payroll %",
        "Non Payroll PSUPW",
        "EBITDA / FMOP",
        "EBITDA / FMOP %",
        "EBITDA / FMOP PSUPW",
    ],
    "ADDITIONAL_ITEMS": [
        "Director's Drawings",
        "Director's NI",
        "Rent",
        "Equipment Hire",
        "Depreciation",
        "Interest Payments",
        "Head Office Costs",
        "Management Costs",
        "Motor Vehicles",
        "Fixtures and Fittings",
        "Other 'one off' costs",
        "Amortisation of Goodwill",
        "Other",
        "Total Expenses in accounts",
        "TOTAL NET PROFIT/LOSS",
    ],
}

# Display label → JSON key for looking up values inside each section dict.
FIELD_NAME_TO_KEY = {
    "No. of Days' Management Accounts": "no_of_days_management_accounts",
    "Fee Deflation/Inflation Factor": "fee_deflation_inflation_factor",
    "Average fee per week": "average_fee_per_week",
    "Registered places": "registered_places",
    "Operational places": "operational_places",
    "Occupancy %": "occupancy_percent",
    "Average fill": "average_fill",
    "Occupancy assuming 100% singles": "occupancy_assuming_100_singles",
    "Fee income per annum": "fee_income_per_annum",
    "Other income": "other_income",
    "Income per annum": "income_per_annum",
    "Employee Wages": "employee_wages",
    "Agency": "agency",
    "National Insurance": "national_insurance",
    "Pension Contribution": "pension_contribution",
    "Payroll": "payroll",
    "Payroll %": "payroll_percent",
    "Payroll PSUPW": "payroll_psupw",
    "Provisions": "provisions",
    "Provisions PSUPW": "provisions_psupw",
    "Heat & Light": "heat_and_light",
    "Heat & Light %": "heat_and_light_percent",
    "Accountancy": "accountancy",
    "Bank Charges": "bank_charges",
    "Clinical Waste": "clinical_waste",
    "Council tax": "council_tax",
    "Gardening Costs": "gardening_costs",
    "Insurance": "insurance",
    "Laundry & Cleaning": "laundry_and_cleaning",
    "Advertising": "advertising",
    "Medical Costs": "medical_costs",
    "Motor Costs": "motor_costs",
    "Print,Post,Stationery, IT": "print_post_stationery_it",
    "Professional Fees": "professional_fees",
    "Registration": "registration",
    "Repairs, Renewal, Maintenance": "repairs_renewal_maintenance",
    "Resident's Activities": "residents_activities",
    "Staff Training and Uniforms": "staff_training_and_uniforms",
    "Subscriptions": "subscriptions",
    "Telephone": "telephone",
    "Water rates": "water_rates",
    "Sundries": "sundries",
    "Total Expenses": "total_expenses",
    "Total expenses PSUPW": "total_expenses_psupw",
    "Non Payroll": "non_payroll",
    "Non Payroll %": "non_payroll_percent",
    "Non Payroll PSUPW": "non_payroll_psupw",
    "EBITDA / FMOP": "ebitda_fmop",
    "EBITDA / FMOP %": "ebitda_fmop_percent",
    "EBITDA / FMOP PSUPW": "ebitda_fmop_psupw",
    "Director's Drawings": "directors_drawings",
    "Director's NI": "directors_ni",
    "Rent": "rent",
    "Equipment Hire": "equipment_hire",
    "Depreciation": "depreciation",
    "Interest Payments": "interest_payments",
    "Head Office Costs": "head_office_costs",
    "Management Costs": "management_costs",
    "Motor Vehicles": "motor_vehicles",
    "Fixtures and Fittings": "fixtures_and_fittings",
    "Other 'one off' costs": "other_one_off_costs",
    "Amortisation of Goodwill": "amortisation_of_goodwill",
    "Other": "other",
    "Total Expenses in accounts": "total_expenses_in_accounts",
    "TOTAL NET PROFIT/LOSS": "total_net_profit_loss",
}


@custom_function()
def create_financial_data_excel(json_file_path: str, output_path: str) -> str:
    """Read extracted financial JSON and produce a formatted HISTORIC ACCOUNTS
    Excel workbook for Colliers healthcare valuations.

    The input JSON should be a list of fiscal-year objects, each with:
      - property_name (str)
      - fiscal_year (str)
      - INCOME, EXPENDITURE, SUMMARY_TOTALS, ADDITIONAL_ITEMS sections
        where each field is a FieldValue dict: {"value": ..., "source": ...}

    Parameters
    ----------
    json_file_path : str
        Path to a JSON file containing a list of fiscal year records.
    output_path : str
        Where to write the .xlsx output.

    Returns
    -------
    str
        The path to the created Excel file.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    raw = json.loads(Path(json_file_path).read_text(encoding="utf-8"))

    properties: dict[str, list[dict]] = defaultdict(list)
    for item in raw:
        properties[item["property_name"]].append(item)
    for prop_name in properties:
        properties[prop_name].sort(key=lambda x: x.get("fiscal_year", ""))

    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid",
    )
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    data_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid",
    )
    category_font = Font(bold=True, size=10)
    italic_font = Font(italic=True, size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for prop_name, fiscal_years in properties.items():
        safe_name = prop_name[:31]
        for ch in "/\\*?[]":
            safe_name = safe_name.replace(ch, "-")
        ws = wb.create_sheet(title=safe_name)
        num_years = len(fiscal_years)

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=num_years + 1,
        )
        ws.cell(row=1, column=1, value="HISTORIC ACCOUNTS").font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        for col_idx, fy in enumerate(fiscal_years, start=2):
            cell = ws.cell(
                row=3,
                column=col_idx,
                value=f"Financial Statements to YE\n{fy['fiscal_year']}",
            )
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        current_row = 5
        for category, fields in FINANCIAL_DATA_SCHEMA.items():
            ws.cell(
                row=current_row,
                column=1,
                value=category.replace("_", " "),
            ).font = category_font
            current_row += 1
            for field_name in fields:
                is_italic = "%" in field_name or "PSUPW" in field_name
                cell = ws.cell(row=current_row, column=1, value=field_name)
                if is_italic:
                    cell.font = italic_font
                json_key = FIELD_NAME_TO_KEY.get(
                    field_name,
                    field_name.lower().replace(" ", "_"),
                )
                for col_idx, fy in enumerate(fiscal_years, start=2):
                    section_data = fy.get(category)
                    value = None
                    source = None
                    if isinstance(section_data, dict):
                        field_obj = section_data.get(json_key)
                        if isinstance(field_obj, dict):
                            value = field_obj.get("value")
                            source = field_obj.get("source")
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                    if value is not None:
                        cell.fill = data_fill
                        if isinstance(value, (int, float)) and "%" in field_name:
                            cell.number_format = "0.0%" if abs(value) <= 1 else "0.00"
                        elif isinstance(value, (int, float)) and abs(value) >= 1000:
                            cell.number_format = "#,##0"
                        if source:
                            cell.comment = Comment(source, "Source")
                current_row += 1
            current_row += 1

        ws.column_dimensions["A"].width = 35
        for col_idx in range(2, num_years + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    out = Path(output_path)
    wb.save(out)
    return str(out)
