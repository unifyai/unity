"""Produce a Colliers-styled Deal Tracker Excel workbook from extracted
web-research deal JSON data.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from unity.function_manager.custom import custom_function

# Ordered (json_key, display_header) pairs defining the column layout.
DEAL_COLUMNS = [
    ("deal_date", "Date"),
    ("name", "Name"),
    ("status", "Status"),
    ("address", "Address"),
    ("description", "Description"),
    ("deal_type", "Deal Type"),
    ("tenant", "Tenant"),
    ("lease_terms", "Lease Terms"),
    ("rent", "Rent"),
    ("rent_pb", "Rent PB"),
    ("price", "Price"),
    ("cv_pb", "CV PB"),
    ("yield_pct", "Yield"),
    ("vendor", "Vendor"),
    ("vendor_agent", "Vendor Agent"),
    ("purchaser", "Purchaser"),
    ("comments", "Comments"),
    ("inputter", "Inputter"),
    ("date_launched_added", "Date Launched / Added"),
    ("postcode", "Postcode"),
    ("region", "Region"),
    ("tenure", "Tenure"),
    ("single_asset_or_portfolio", "Single Asset or Portfolio"),
    ("homes", "Homes"),
    ("beds", "Beds"),
    ("build_type", "Build Type"),
    ("age", "Age"),
    ("quote_price", "Quote Price"),
    ("quote_cv_pb", "Quote CV PB"),
    ("quote_yield", "Quote Yield"),
    ("yield_grouping_category", "Yield Grouping / Category"),
    ("purchaser_agent", "Purchaser Agent"),
    ("epc_rating", "EPC Rating"),
    ("vendor_type", "Vendor Type"),
    ("vendor_nationality", "Vendor Nationality"),
    ("vendors_global_territory", "Vendor's Global Territory"),
    ("purchaser_type", "Purchaser Type"),
    ("purchaser_nationality", "Purchaser Nationality"),
    ("purchasers_global_territory", "Purchaser's Global Territory"),
    ("achieved_vs_quote_gbp", "Achieved vs Quote (\u00a3)"),
    ("achieved_vs_quote_pct", "Achieved vs Quote (%)"),
    ("achieved_vs_quote_cv_pb", "Achieved vs Quote (CV PB)"),
    ("achieved_vs_quote_niy_basis_point", "Achieved vs Quote (NIY Basis Point)"),
    ("transaction_quarter", "Transaction Quarter"),
    ("transaction_year", "Transaction Year"),
    ("transaction_time_weeks", "Transaction Time (Weeks)"),
    ("transaction_time_months", "Transaction Time (Months)"),
]

PRIMARY_COL_COUNT = 18

CURRENCY_FIELDS = frozenset(
    {
        "rent",
        "rent_pb",
        "price",
        "cv_pb",
        "quote_price",
        "quote_cv_pb",
        "achieved_vs_quote_gbp",
        "achieved_vs_quote_cv_pb",
    },
)

PERCENT_FIELDS = frozenset({"achieved_vs_quote_pct"})

DECIMAL_NUMBER_FIELDS = frozenset(
    {
        "achieved_vs_quote_niy_basis_point",
        "transaction_time_weeks",
        "transaction_time_months",
    },
)


@custom_function()
def create_web_search_excel(json_file_path: str, output_path: str) -> str:
    """Read extracted deal-research JSON and produce a Colliers-styled Deal
    Tracker Excel workbook.

    The input JSON should be a list of deal objects with keys matching the
    Deal Tracker schema (e.g. deal_date, name, status, address, price, etc.).

    Parameters
    ----------
    json_file_path : str
        Path to a JSON file containing a list of deal records.
    output_path : str
        Where to write the .xlsx output.

    Returns
    -------
    str
        The path to the created Excel file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    raw = json.loads(Path(json_file_path).read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Tracker"

    field_names = [col[0] for col in DEAL_COLUMNS]
    headers = [col[1] for col in DEAL_COLUMNS]
    num_cols = len(headers)

    blue_fill = PatternFill(
        start_color="1F3864",
        end_color="1F3864",
        fill_type="solid",
    )
    grey_fill = PatternFill(
        start_color="404040",
        end_color="404040",
        fill_type="solid",
    )
    header_font = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = blue_fill if col_idx <= PRIMARY_COL_COUNT else grey_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"
    ws.row_dimensions[1].height = 30

    for row_idx, deal in enumerate(raw, start=2):
        for col_idx, fname in enumerate(field_names, start=1):
            val = deal.get(fname)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border

            if val is None:
                continue

            if isinstance(val, str):
                # Attempt date parsing for date fields
                if fname in ("deal_date", "date_launched_added"):
                    try:
                        parsed = date.fromisoformat(val)
                        cell.value = parsed
                        cell.number_format = "MMM-YY"
                        continue
                    except ValueError:
                        pass
                # Attempt numeric coercion for currency/number fields
                if fname in CURRENCY_FIELDS | PERCENT_FIELDS | DECIMAL_NUMBER_FIELDS:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        pass

            if isinstance(val, date):
                cell.value = val
                cell.number_format = "MMM-YY"
            elif fname in CURRENCY_FIELDS and isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = "#,##0"
            elif fname in PERCENT_FIELDS and isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = "0.00%"
            elif fname in DECIMAL_NUMBER_FIELDS and isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = "0.0"
            else:
                cell.value = val

    for col_idx, header in enumerate(headers, start=1):
        width = max(12, min(35, len(header) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "C2"

    out = Path(output_path)
    wb.save(out)
    return str(out)
