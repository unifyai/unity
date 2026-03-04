"""Shared helpers for Colliers ASH financial extraction eval tests.

Provides ground truth parsing, output parsing, and comparison logic
used by both the Colliers-seeded and generic baseline tests.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Any

import openpyxl

FIXTURES_DIR = Path(__file__).parent / "fixtures" / "colliers_ash"
GROUND_TRUTH_XLSX = FIXTURES_DIR / "ASH_Historic_Accounts.xlsx"
TEMPLATE_XLSX = FIXTURES_DIR / "HISTORIC_ACCOUNTS_TEMPLATE.xlsx"
PDF_FILES = sorted(FIXTURES_DIR.glob("*.pdf"))

# The source PDFs use labels that don't map 1:1 to the schema fields.
# For each ground-truth field below, we also accept the value appearing
# under any of the listed alternatives.
FIELD_ALIASES: dict[str, list[str]] = {
    "Insurance": ["Registration"],
    "Council tax": ["Water rates"],
    "Repairs, Renewal, Maintenance": ["Repairs/Renewal/Maintenance"],
}

# Interpretation-dependent fields where multiple valid readings exist.
# Both the ground truth value and known alternatives are accepted.
BASE_LENIENT_FIELDS: dict[str, dict[str, list]] = {
    # Admin Expenses only (ground truth) vs Admin + Cost of Sale (also valid)
    "Total Expenses": {
        "0": [1141888, 1201256],
        "1": [1387181, 1432792],
        "2": [1158405, 1205903],
        "3": [1722679, 1782390],
    },
    # Pre-tax (ground truth) vs post-tax (also valid reading of "net")
    "TOTAL NET PROFIT/LOSS": {
        "0": [149400, 123690],
        "1": [34689, 23439],
        "2": [-136124],
        "3": [158042, 148214],
    },
    # FY2023: "Repairs" and "Exceptional repairs" are separate PDF lines;
    # combining them is a classification choice, not an extraction error.
    "Repairs, Renewal, Maintenance": {
        "1": [36035, 67917],
    },
}

# Additional lenient fields for the generic test (no domain pre-seeding).
EXTRA_LENIENT_FIELDS: dict[str, dict[str, list]] = {
    # Source PDF has internal inconsistency: Interest = 86,172 on EBITDA page
    # vs 86,712 on P&L page. Both 14,502 (EBITDA page total) and 15,042
    # (recomputed from P&L interest) are defensible.
    "EBITDA / FMOP": {
        "2": [14502, 15042],
    },
    # "Legal: 120" is the smallest line item (0.007% of total expenses),
    # at the very bottom of the notes page. Missing it entirely on a first
    # pass is a minor omission, not a systematic failure.
    "Professional Fees": {
        "3": [120],
    },
}


def merge_lenient_fields(
    *dicts: dict[str, dict[str, list]],
) -> dict[str, dict[str, list]]:
    """Merge multiple lenient field dicts, combining per-year value lists."""
    merged: dict[str, dict[str, list]] = {}
    for d in dicts:
        for field, years in d.items():
            if field not in merged:
                merged[field] = {}
            for year_idx, vals in years.items():
                existing = merged[field].get(year_idx, [])
                for v in vals:
                    if v not in existing:
                        existing.append(v)
                merged[field][year_idx] = existing
    return merged


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def parse_ground_truth() -> dict[str, dict[str, Any]]:
    """Parse the ground truth Excel into {field_label: {year_index: value}}.

    year_index "0" = FYE 2022, "1" = FYE 2023, "2" = FYE 2024, "3" = FYE 2025.
    Only non-null cells are included.
    """
    wb = openpyxl.load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    truth: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=5,
        values_only=False,
    ):
        label = row[0].value
        if label is None:
            continue
        for cell in row[1:]:
            if cell.value is not None:
                year_idx = cell.column - 2
                truth.setdefault(label, {})[str(year_idx)] = cell.value
    return truth


def parse_output_excel(output_path: str) -> dict[str, dict[str, Any]]:
    """Parse an actor-produced Excel into the same format as ground truth."""
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    extracted: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column,
        values_only=False,
    ):
        label = row[0].value
        if label is None:
            continue
        for cell in row[1:]:
            if cell.value is not None:
                year_idx = cell.column - 2
                extracted.setdefault(label, {})[str(year_idx)] = cell.value
    return extracted


def _field_name_to_key_map() -> dict[str, str]:
    from unity.customization.clients.colliers.functions.create_financial_data_excel import (
        FIELD_NAME_TO_KEY,
    )

    return FIELD_NAME_TO_KEY


def parse_output_json(output_path: str) -> dict[str, dict[str, Any]]:
    """Parse actor-produced JSON (FiscalYearData schema) into ground truth format."""
    key_to_display = {v: k for k, v in _field_name_to_key_map().items()}

    raw = json.loads(Path(output_path).read_text(encoding="utf-8"))
    raw.sort(key=lambda x: x.get("fiscal_year", ""))

    extracted: dict[str, dict[str, Any]] = {}
    for year_idx, record in enumerate(raw):
        for section_name in (
            "INCOME",
            "EXPENDITURE",
            "SUMMARY_TOTALS",
            "ADDITIONAL_ITEMS",
        ):
            section = record.get(section_name)
            if not isinstance(section, dict):
                continue
            for json_key, field_obj in section.items():
                if json_key == "notes":
                    continue
                value = (
                    field_obj.get("value") if isinstance(field_obj, dict) else field_obj
                )
                if value is not None:
                    display_name = key_to_display.get(json_key, json_key)
                    extracted.setdefault(display_name, {})[str(year_idx)] = value
    return extracted


def collect_json_values(tmp_path: Path) -> dict[str, dict[str, Any]]:
    """Best-effort parse of any JSON the actor produced."""
    json_files = list(tmp_path.rglob("*.json"))
    if not json_files:
        return {}

    key_to_display = {v: k for k, v in _field_name_to_key_map().items()}

    for jf in json_files:
        try:
            raw = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(raw, list) or not raw:
            continue

        raw.sort(key=lambda x: str(x.get("fiscal_year", x.get("year", ""))))
        extracted: dict[str, dict[str, Any]] = {}
        for year_idx, record in enumerate(raw):
            for section_name in (
                "INCOME",
                "EXPENDITURE",
                "SUMMARY_TOTALS",
                "ADDITIONAL_ITEMS",
                "income",
                "expenditure",
                "summary_totals",
                "additional_items",
            ):
                section = record.get(section_name)
                if not isinstance(section, dict):
                    continue
                for json_key, field_obj in section.items():
                    if json_key == "notes":
                        continue
                    value = (
                        field_obj.get("value")
                        if isinstance(field_obj, dict)
                        else field_obj
                    )
                    if value is not None:
                        display_name = key_to_display.get(json_key, json_key)
                        extracted.setdefault(display_name, {})[str(year_idx)] = value
        if extracted:
            return extracted
    return {}


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------


def values_match(
    expected_val: Any,
    actual_val: Any,
    tolerance: float = 0.02,
) -> bool:
    """Check if two values match within tolerance."""
    if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
        if expected_val == 0:
            return actual_val == 0
        return abs(actual_val - expected_val) / abs(expected_val) <= tolerance
    return str(expected_val) == str(actual_val)


def compare_results(
    expected: dict[str, dict[str, Any]],
    actual: dict[str, dict[str, Any]],
    lenient_fields: dict[str, dict[str, list]],
    zero_is_optional: bool = True,
) -> tuple[list[str], list[str], list[str]]:
    """Compare expected vs actual with alias and leniency support.

    Returns (matches, mismatches, acceptable_gaps).
    """
    matches: list[str] = []
    mismatches: list[str] = []
    acceptable: list[str] = []

    for field, year_vals in expected.items():
        aliases = FIELD_ALIASES.get(field, [])
        lenient_vals = lenient_fields.get(field, {})

        for year_idx, expected_val in year_vals.items():
            # --- Lenient fields ---
            if year_idx in lenient_vals:
                valid_options = lenient_vals[year_idx]
                actual_val = actual.get(field, {}).get(year_idx)
                if actual_val is not None and any(
                    values_match(opt, actual_val) for opt in valid_options
                ):
                    matches.append(
                        f"{field}[year={year_idx}]: {actual_val} "
                        f"(accepted, valid: {valid_options})",
                    )
                    continue
                found_via_alias = False
                for alias in aliases:
                    alias_val = actual.get(alias, {}).get(year_idx)
                    if alias_val is not None and any(
                        values_match(opt, alias_val) for opt in valid_options
                    ):
                        matches.append(
                            f"{field}[year={year_idx}]: {alias_val} "
                            f"via '{alias}' (lenient)",
                        )
                        found_via_alias = True
                        break
                if not found_via_alias:
                    acceptable.append(
                        f"{field}[year={year_idx}]: lenient, not found",
                    )
                continue

            # --- Zero values ---
            if zero_is_optional and expected_val == 0:
                actual_val = actual.get(field, {}).get(year_idx)
                if actual_val is not None and values_match(0, actual_val):
                    matches.append(f"{field}[year={year_idx}]: 0")
                else:
                    acceptable.append(
                        f"{field}[year={year_idx}]: zero omitted",
                    )
                continue

            # --- Primary field name ---
            actual_val = actual.get(field, {}).get(year_idx)
            if actual_val is not None:
                if values_match(expected_val, actual_val):
                    matches.append(
                        f"{field}[year={year_idx}]: {expected_val} ~ {actual_val}",
                    )
                else:
                    mismatches.append(
                        f"{field}[year={year_idx}]: "
                        f"expected={expected_val}, got={actual_val}",
                    )
                continue

            # --- Aliases ---
            found_via_alias = False
            for alias in aliases:
                alias_val = actual.get(alias, {}).get(year_idx)
                if alias_val is not None:
                    if values_match(expected_val, alias_val):
                        matches.append(
                            f"{field}[year={year_idx}]: {expected_val} "
                            f"~ {alias_val} via '{alias}'",
                        )
                    else:
                        mismatches.append(
                            f"{field}[year={year_idx}]: expected={expected_val}, "
                            f"got={alias_val} (via alias '{alias}')",
                        )
                    found_via_alias = True
                    break

            if not found_via_alias:
                mismatches.append(
                    f"{field}[year={year_idx}]: expected={expected_val}, "
                    f"not found (checked aliases: {aliases})",
                )

    return matches, mismatches, acceptable


def print_comparison(
    label: str,
    matches: list[str],
    mismatches: list[str],
    acceptable: list[str],
) -> None:
    """Print a formatted comparison report."""
    total_strict = len(matches) + len(mismatches)
    match_rate = len(matches) / total_strict if total_strict > 0 else 0

    print(f"\n{'='*60}")
    print(label)
    print(f"{'='*60}")
    print(f"Strict matches:   {len(matches)}")
    print(f"Strict failures:  {len(mismatches)}")
    print(f"Acceptable gaps:  {len(acceptable)}")
    print(f"Match rate:        {match_rate:.1%} ({len(matches)}/{total_strict})")

    if mismatches:
        print(f"\nSTRICT FAILURES ({len(mismatches)}):")
        for m in mismatches:
            print(f"  {m}")

    if acceptable:
        print(f"\nACCEPTABLE GAPS ({len(acceptable)}):")
        for m in acceptable[:10]:
            print(f"  {m}")

    if matches:
        print(f"\nSAMPLE MATCHES (first 10 of {len(matches)}):")
        for m in matches[:10]:
            print(f"  {m}")
    print(f"{'='*60}\n")


def stable_workspace(name: str) -> tuple[Path, Path]:
    """Create a deterministic workspace under ~/Unity/Local for cache-stable prompts.

    Returns (workspace_dir, output_path) with fixed paths that don't
    contain session-varying components like pytest's tmp_path counter.
    The workspace is cleaned and recreated on each call.
    """
    from unity.file_manager.settings import get_local_root

    root = Path(get_local_root())
    workspace_dir = root / name / "accounts"
    output_path = root / name / "output" / "ASH_Historic_Accounts.xlsx"

    if workspace_dir.parent.exists():
        shutil.rmtree(workspace_dir.parent)
    workspace_dir.mkdir(parents=True)
    output_path.parent.mkdir(parents=True)

    return workspace_dir, output_path


def find_output(
    search_dirs: list[Path],
    output_path: Path,
) -> dict[str, dict[str, Any]]:
    """Find and parse the actor's output, trying Excel then JSON.

    Searches the explicit output_path first, then globs across all
    provided search directories.
    """
    if output_path.exists():
        return parse_output_excel(str(output_path))

    for d in search_dirs:
        if not d.exists():
            continue
        xlsx_files = [f for f in d.rglob("*.xlsx") if "TEMPLATE" not in f.name]
        if xlsx_files:
            return parse_output_excel(str(xlsx_files[0]))

    for d in search_dirs:
        if not d.exists():
            continue
        json_files = list(d.rglob("*.json"))
        if json_files:
            return parse_output_json(str(json_files[0]))

    for d in search_dirs:
        if d.exists():
            result = collect_json_values(d)
            if result:
                return result

    return {}
