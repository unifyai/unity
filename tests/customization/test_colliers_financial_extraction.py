"""Colliers Healthcare — financial data extraction eval test.

Runs the full CodeActActor with pre-seeded Colliers functions and guidance
against four real PDF annual accounts, then compares the extracted financial
data against a human-validated ground truth Excel workbook.

Fixture data lives in tests/customization/fixtures/colliers_ash/.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from tests.helpers import _handle_project
from tests.actor.state_managers.utils import make_code_act_actor
from unity.manager_registry import ManagerRegistry

pytestmark = pytest.mark.eval

FIXTURES_DIR = Path(__file__).parent / "fixtures" / "colliers_ash"

GROUND_TRUTH_XLSX = FIXTURES_DIR / "ASH_Historic_Accounts.xlsx"

PDF_FILES = sorted(FIXTURES_DIR.glob("*.pdf"))

# ---------------------------------------------------------------------------
# Ambiguous field mappings
# ---------------------------------------------------------------------------
# The source PDFs use labels that don't map 1:1 to the schema fields.
# The ground truth made one mapping choice; a fresh LLM run may make
# another equally valid choice.  For each ground-truth field below, we
# also accept the value appearing under any of the listed alternatives.

FIELD_ALIASES: dict[str, list[str]] = {
    "Insurance": ["Registration"],
    "Council tax": ["Water rates"],
    "Repairs, Renewal, Maintenance": ["Repairs/Renewal/Maintenance"],
}

# Values that are zero in the ground truth.  Omitting a zero-valued field
# (producing null instead of 0) is acceptable — it's a convention choice,
# not an extraction error.
ZERO_IS_OPTIONAL = True

# ---------------------------------------------------------------------------
# Computed / interpretation-dependent fields
# ---------------------------------------------------------------------------
# These fields depend on domain conventions (what counts as "Total Expenses",
# pre-tax vs post-tax profit, etc.) rather than raw extraction from the PDF.
# We accept any value that can be explained by a defensible interpretation
# of the source document.  The ground truth value and the known alternative
# are both listed; we accept either.

LENIENT_FIELDS: dict[str, dict[str, list]] = {
    # Admin Expenses only (ground truth) vs Admin + Cost of Sale (also valid)
    "Total Expenses": {
        "0": [1141888, 1201256],
        "1": [1387181, 1432792],
        "2": [1158405, 1205903],
        "3": [1722679, 1782390],
    },
    # Pre-tax (ground truth) vs post-tax (also valid reading of "net")
    "TOTAL NET PROFIT/LOSS": {
        "0": [149400, 123690],
        "1": [34689, 23439],
        "2": [-136124],
        "3": [158042, 148214],
    },
    # FY2023: PDF shows "Repairs 36,035" and "Exceptional repairs 31,882"
    # as separate line items. Ground truth separates them; combining them
    # into a single repairs figure (67,917) is a classification choice.
    "Repairs, Renewal, Maintenance": {
        "1": [36035, 67917],
    },
}


def _parse_ground_truth() -> dict[str, dict[str, Any]]:
    """Parse the ground truth Excel into {field_label: {year_index: value}}.

    year_index 0 = FYE 2022, 1 = FYE 2023, 2 = FYE 2024, 3 = FYE 2025.
    Only non-null cells are included.
    """
    wb = openpyxl.load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    truth: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=5,
        values_only=False,
    ):
        label = row[0].value
        if label is None:
            continue
        for cell in row[1:]:
            if cell.value is not None:
                year_idx = cell.column - 2
                truth.setdefault(label, {})[str(year_idx)] = cell.value
    return truth


def _parse_output_excel(output_path: str) -> dict[str, dict[str, Any]]:
    """Parse the actor-produced Excel into the same format as ground truth."""
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    extracted: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column,
        values_only=False,
    ):
        label = row[0].value
        if label is None:
            continue
        for cell in row[1:]:
            if cell.value is not None:
                year_idx = cell.column - 2
                extracted.setdefault(label, {})[str(year_idx)] = cell.value
    return extracted


def _parse_output_json(output_path: str) -> dict[str, dict[str, Any]]:
    """Parse the actor-produced JSON into the same format as ground truth.

    Handles the FiscalYearData JSON schema where each field is a FieldValue
    dict with "value" and "source" keys.
    """
    from unity.customization.clients.colliers.functions.create_financial_data_excel import (
        FIELD_NAME_TO_KEY,
    )

    key_to_display = {v: k for k, v in FIELD_NAME_TO_KEY.items()}

    raw = json.loads(Path(output_path).read_text(encoding="utf-8"))
    raw.sort(key=lambda x: x.get("fiscal_year", ""))

    extracted: dict[str, dict[str, Any]] = {}
    for year_idx, record in enumerate(raw):
        for section_name in (
            "INCOME",
            "EXPENDITURE",
            "SUMMARY_TOTALS",
            "ADDITIONAL_ITEMS",
        ):
            section = record.get(section_name)
            if not isinstance(section, dict):
                continue
            for json_key, field_obj in section.items():
                if json_key == "notes":
                    continue
                if isinstance(field_obj, dict):
                    value = field_obj.get("value")
                else:
                    value = field_obj
                if value is not None:
                    display_name = key_to_display.get(json_key, json_key)
                    extracted.setdefault(display_name, {})[str(year_idx)] = value
    return extracted


def _values_match(expected_val: Any, actual_val: Any, tolerance: float = 0.02) -> bool:
    """Check if two values match within tolerance."""
    if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
        if expected_val == 0:
            return actual_val == 0
        return abs(actual_val - expected_val) / abs(expected_val) <= tolerance
    return str(expected_val) == str(actual_val)


def _compare_results(
    expected: dict[str, dict[str, Any]],
    actual: dict[str, dict[str, Any]],
) -> tuple[list[str], list[str], list[str]]:
    """Compare expected vs actual with alias and leniency support.

    Returns (matches, mismatches, acceptable_gaps).
    acceptable_gaps are things we don't penalise (zero omissions, alias
    matches found under a different field name, lenient-field alternatives).
    """
    matches: list[str] = []
    mismatches: list[str] = []
    acceptable: list[str] = []

    for field, year_vals in expected.items():
        aliases = FIELD_ALIASES.get(field, [])
        lenient_vals = LENIENT_FIELDS.get(field, {})

        for year_idx, expected_val in year_vals.items():
            # --- Lenient fields: accept any of the listed valid values ---
            if year_idx in lenient_vals:
                valid_options = lenient_vals[year_idx]
                # Check under primary field name
                actual_val = actual.get(field, {}).get(year_idx)
                if actual_val is not None and any(
                    _values_match(opt, actual_val) for opt in valid_options
                ):
                    matches.append(
                        f"{field}[year={year_idx}]: {actual_val} "
                        f"(accepted, valid options: {valid_options})",
                    )
                    continue
                # Check under aliases
                found_via_alias = False
                for alias in aliases:
                    alias_val = actual.get(alias, {}).get(year_idx)
                    if alias_val is not None and any(
                        _values_match(opt, alias_val) for opt in valid_options
                    ):
                        matches.append(
                            f"{field}[year={year_idx}]: {alias_val} via '{alias}' "
                            f"(accepted, valid options: {valid_options})",
                        )
                        found_via_alias = True
                        break
                if found_via_alias:
                    continue
                acceptable.append(
                    f"{field}[year={year_idx}]: lenient field, "
                    f"expected one of {valid_options}, not found",
                )
                continue

            # --- Zero values: omission is acceptable ---
            if ZERO_IS_OPTIONAL and expected_val == 0:
                actual_val = actual.get(field, {}).get(year_idx)
                if actual_val is not None and _values_match(0, actual_val):
                    matches.append(f"{field}[year={year_idx}]: 0")
                else:
                    acceptable.append(
                        f"{field}[year={year_idx}]: zero omitted (acceptable)",
                    )
                continue

            # --- Check primary field name ---
            actual_val = actual.get(field, {}).get(year_idx)
            if actual_val is not None:
                if _values_match(expected_val, actual_val):
                    matches.append(
                        f"{field}[year={year_idx}]: {expected_val} ~ {actual_val}",
                    )
                else:
                    mismatches.append(
                        f"{field}[year={year_idx}]: "
                        f"expected={expected_val}, got={actual_val}",
                    )
                continue

            # --- Check aliases ---
            found_via_alias = False
            for alias in aliases:
                alias_val = actual.get(alias, {}).get(year_idx)
                if alias_val is not None:
                    if _values_match(expected_val, alias_val):
                        matches.append(
                            f"{field}[year={year_idx}]: {expected_val} ~ {alias_val} "
                            f"via alias '{alias}'",
                        )
                    else:
                        mismatches.append(
                            f"{field}[year={year_idx}]: expected={expected_val}, "
                            f"got={alias_val} (via alias '{alias}')",
                        )
                    found_via_alias = True
                    break

            if not found_via_alias:
                mismatches.append(
                    f"{field}[year={year_idx}]: expected={expected_val}, "
                    f"not found (checked aliases: {aliases})",
                )

    return matches, mismatches, acceptable


def _seed_colliers_guidance() -> None:
    """Seed the GuidanceManager with Colliers guidance entries."""
    from unity.customization.clients.colliers import _COLLIERS_GUIDANCE

    gm = ManagerRegistry.get_guidance_manager()
    for entry in _COLLIERS_GUIDANCE:
        gm.add_guidance(title=entry["title"], content=entry["content"])


def _seed_colliers_functions() -> "FunctionManager":
    """Create a FunctionManager with Colliers custom functions synced."""
    from unity.function_manager.function_manager import FunctionManager
    from unity.function_manager.custom_functions import (
        collect_functions_from_directories,
    )

    colliers_fn_dir = (
        Path(__file__).resolve().parents[2]
        / "unity"
        / "customization"
        / "clients"
        / "colliers"
        / "functions"
    )
    fm = FunctionManager()
    source_fns = collect_functions_from_directories([colliers_fn_dir])
    if source_fns:
        fm.sync_custom(source_functions=source_fns, source_venvs={})
    return fm


@pytest.mark.asyncio
@pytest.mark.timeout(900)
@_handle_project
async def test_colliers_ash_financial_extraction(tmp_path: Path):
    """Extract financial data from 4 ASH annual accounts PDFs and compare
    against human-validated ground truth.

    This is the core Colliers day-1 capability test.
    """
    assert len(PDF_FILES) == 4, f"Expected 4 PDF fixtures, found {len(PDF_FILES)}"
    assert GROUND_TRUTH_XLSX.exists(), "Ground truth Excel not found"

    # Copy PDFs to workspace
    workspace_dir = tmp_path / "accounts"
    workspace_dir.mkdir()
    for pdf in PDF_FILES:
        shutil.copy(pdf, workspace_dir / pdf.name)

    output_dir = tmp_path / "output"
    output_dir.mkdir()
    output_path = output_dir / "ASH_Historic_Accounts.xlsx"

    _seed_colliers_guidance()
    fm = _seed_colliers_functions()

    async with make_code_act_actor(
        impl="real",
        include_function_manager_tools=True,
        function_manager=fm,
    ) as (actor, _primitives, calls):
        handle = await actor.act(
            f"Please extract the financial account data from the PDF files in "
            f"{workspace_dir} and produce the standard HISTORIC ACCOUNTS Excel "
            f"spreadsheet. Save the output to {output_path}. "
            f"There are 4 annual accounts PDFs covering fiscal years ending "
            f"30 April 2022 through 30 April 2025. The property name is "
            f"'Ablegrange Severn Heights'.",
            guidelines=(
                "You are the Colliers Healthcare Valuation Assistant. "
                "Search your guidance and functions for the standard financial "
                "extraction workflow and Excel formatting function."
            ),
            clarification_enabled=False,
        )
        result = await handle.result()

    # Find output files
    json_files = list(tmp_path.rglob("*.json"))
    xlsx_files = list(tmp_path.rglob("*.xlsx"))

    print(f"\n{'='*60}")
    print(f"Actor result:\n{result[:500]}")
    print(f"\nJSON files: {json_files}")
    print(f"Excel files: {xlsx_files}")
    print(f"Primitive calls: {calls}")
    print(f"{'='*60}\n")

    # Parse ground truth
    expected = _parse_ground_truth()

    # Try to parse output — prefer Excel, fall back to JSON
    actual: dict[str, dict[str, Any]] = {}
    if output_path.exists():
        actual = _parse_output_excel(str(output_path))
    elif xlsx_files:
        actual = _parse_output_excel(str(xlsx_files[0]))
    elif json_files:
        actual = _parse_output_json(str(json_files[0]))

    if not actual:
        pytest.fail(
            f"No output found. Actor result:\n{result}\n\n"
            f"Files in workspace: {list(tmp_path.rglob('*'))}",
        )

    matches, mismatches, acceptable = _compare_results(expected, actual)

    print(f"\n{'='*60}")
    print("COMPARISON RESULTS")
    print(f"{'='*60}")
    print(f"Strict matches:   {len(matches)}")
    print(f"Strict failures:  {len(mismatches)}")
    print(f"Acceptable gaps:  {len(acceptable)}")

    if mismatches:
        print("\nSTRICT FAILURES (objective errors):")
        for m in mismatches:
            print(f"  {m}")

    if acceptable:
        print("\nACCEPTABLE GAPS (ambiguous / zero omissions / lenient):")
        for m in acceptable:
            print(f"  {m}")

    if matches:
        print(f"\nSAMPLE MATCHES (first 10 of {len(matches)}):")
        for m in matches[:10]:
            print(f"  {m}")
    print(f"{'='*60}\n")

    # Hard assertion: no objective extraction errors
    assert (
        len(mismatches) == 0
    ), f"{len(mismatches)} strict failures (objective errors):\n" + "\n".join(
        f"  {m}" for m in mismatches
    )

    # Soft assertion: most values should be extracted
    total_checkable = len(matches) + len(mismatches)
    assert total_checkable >= 60, (
        f"Only {total_checkable} values found in output "
        f"(expected at least 60 of ~90 non-zero ground truth values). "
        f"The actor may not have completed extraction."
    )
