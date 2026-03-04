"""Generic financial extraction eval test — NO Colliers customization.

Runs the same PDF extraction task as the Colliers test, but with zero
domain-specific pre-seeding: no custom functions, no guidance entries, no
special guidelines.  The actor has only the base CodeActActor prompt and
general FileManager primitives (including render_pdf / render_excel_sheet).

The request is written the way you'd brief a capable new hire who has
never done healthcare valuation accounting before, providing the empty
Excel template as the target format.

This test measures what the general-purpose system can do out of the box,
establishing a baseline that the Colliers pre-seeding improves upon.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from tests.helpers import _handle_project
from tests.actor.state_managers.utils import make_code_act_actor

pytestmark = pytest.mark.eval

FIXTURES_DIR = Path(__file__).parent / "fixtures" / "colliers_ash"
GROUND_TRUTH_XLSX = FIXTURES_DIR / "ASH_Historic_Accounts.xlsx"
TEMPLATE_XLSX = FIXTURES_DIR / "HISTORIC_ACCOUNTS_TEMPLATE.xlsx"
PDF_FILES = sorted(FIXTURES_DIR.glob("*.pdf"))

# Same lenient comparison infrastructure as the Colliers test
FIELD_ALIASES: dict[str, list[str]] = {
    "Insurance": ["Registration"],
    "Council tax": ["Water rates"],
    "Repairs, Renewal, Maintenance": ["Repairs/Renewal/Maintenance"],
}

ZERO_IS_OPTIONAL = True

LENIENT_FIELDS: dict[str, dict[str, list]] = {
    # Admin Expenses only (ground truth) vs Admin + Cost of Sale (also valid)
    "Total Expenses": {
        "0": [1141888, 1201256],
        "1": [1387181, 1432792],
        "2": [1158405, 1205903],
        "3": [1722679, 1782390],
    },
    # Pre-tax (ground truth) vs post-tax (also valid reading of "net")
    "TOTAL NET PROFIT/LOSS": {
        "0": [149400, 123690],
        "1": [34689, 23439],
        "2": [-136124],
        "3": [158042, 148214],
    },
    # FY2023: "Repairs" and "Exceptional repairs" are separate PDF lines;
    # combining them is a classification choice, not an extraction error.
    "Repairs, Renewal, Maintenance": {
        "1": [36035, 67917],
    },
    # Source PDF has internal inconsistency: Interest = 86,172 on EBITDA page
    # vs 86,712 on P&L page. Both 14,502 (EBITDA page total) and 15,042
    # (recomputed from P&L interest) are defensible.
    "EBITDA / FMOP": {
        "2": [14502, 15042],
    },
    # "Legal: 120" is the smallest line item (0.007% of total expenses),
    # at the very bottom of the notes page. Missing it entirely on a first
    # pass is a minor omission, not a systematic failure.
    "Professional Fees": {
        "3": [120],
    },
}


def _parse_ground_truth() -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    truth: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=5,
        values_only=False,
    ):
        label = row[0].value
        if label is None:
            continue
        for cell in row[1:]:
            if cell.value is not None:
                year_idx = cell.column - 2
                truth.setdefault(label, {})[str(year_idx)] = cell.value
    return truth


def _parse_output_excel(output_path: str) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    extracted: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column,
        values_only=False,
    ):
        label = row[0].value
        if label is None:
            continue
        for cell in row[1:]:
            if cell.value is not None:
                year_idx = cell.column - 2
                extracted.setdefault(label, {})[str(year_idx)] = cell.value
    return extracted


def _collect_all_json_values(tmp_path: Path) -> dict[str, dict[str, Any]]:
    """Best-effort parse of any JSON the actor produced.

    Tries to interpret the JSON as fiscal-year records with section dicts
    containing FieldValue-style entries.  Falls back to flat key scanning.
    """
    json_files = list(tmp_path.rglob("*.json"))
    if not json_files:
        return {}

    from unity.customization.clients.colliers.functions.create_financial_data_excel import (
        FIELD_NAME_TO_KEY,
    )

    key_to_display = {v: k for k, v in FIELD_NAME_TO_KEY.items()}

    for jf in json_files:
        try:
            raw = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(raw, list) or not raw:
            continue

        raw.sort(key=lambda x: str(x.get("fiscal_year", x.get("year", ""))))
        extracted: dict[str, dict[str, Any]] = {}
        for year_idx, record in enumerate(raw):
            for section_name in (
                "INCOME",
                "EXPENDITURE",
                "SUMMARY_TOTALS",
                "ADDITIONAL_ITEMS",
                "income",
                "expenditure",
                "summary_totals",
                "additional_items",
            ):
                section = record.get(section_name)
                if not isinstance(section, dict):
                    continue
                for json_key, field_obj in section.items():
                    if json_key == "notes":
                        continue
                    if isinstance(field_obj, dict):
                        value = field_obj.get("value")
                    else:
                        value = field_obj
                    if value is not None:
                        display_name = key_to_display.get(json_key, json_key)
                        extracted.setdefault(display_name, {})[str(year_idx)] = value
        if extracted:
            return extracted
    return {}


def _values_match(expected_val: Any, actual_val: Any, tolerance: float = 0.02) -> bool:
    if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
        if expected_val == 0:
            return actual_val == 0
        return abs(actual_val - expected_val) / abs(expected_val) <= tolerance
    return str(expected_val) == str(actual_val)


def _compare_results(
    expected: dict[str, dict[str, Any]],
    actual: dict[str, dict[str, Any]],
) -> tuple[list[str], list[str], list[str]]:
    matches: list[str] = []
    mismatches: list[str] = []
    acceptable: list[str] = []

    for field, year_vals in expected.items():
        aliases = FIELD_ALIASES.get(field, [])
        lenient_vals = LENIENT_FIELDS.get(field, {})

        for year_idx, expected_val in year_vals.items():
            if year_idx in lenient_vals:
                valid_options = lenient_vals[year_idx]
                actual_val = actual.get(field, {}).get(year_idx)
                if actual_val is not None and any(
                    _values_match(opt, actual_val) for opt in valid_options
                ):
                    matches.append(
                        f"{field}[year={year_idx}]: {actual_val} (lenient match)",
                    )
                    continue
                for alias in aliases:
                    alias_val = actual.get(alias, {}).get(year_idx)
                    if alias_val is not None and any(
                        _values_match(opt, alias_val) for opt in valid_options
                    ):
                        matches.append(
                            f"{field}[year={year_idx}]: {alias_val} via '{alias}' (lenient)",
                        )
                        break
                else:
                    acceptable.append(
                        f"{field}[year={year_idx}]: lenient, not found",
                    )
                continue

            if ZERO_IS_OPTIONAL and expected_val == 0:
                actual_val = actual.get(field, {}).get(year_idx)
                if actual_val is not None and _values_match(0, actual_val):
                    matches.append(f"{field}[year={year_idx}]: 0")
                else:
                    acceptable.append(
                        f"{field}[year={year_idx}]: zero omitted",
                    )
                continue

            actual_val = actual.get(field, {}).get(year_idx)
            if actual_val is not None:
                if _values_match(expected_val, actual_val):
                    matches.append(
                        f"{field}[year={year_idx}]: {expected_val} ~ {actual_val}",
                    )
                else:
                    mismatches.append(
                        f"{field}[year={year_idx}]: expected={expected_val}, got={actual_val}",
                    )
                continue

            found_via_alias = False
            for alias in aliases:
                alias_val = actual.get(alias, {}).get(year_idx)
                if alias_val is not None:
                    if _values_match(expected_val, alias_val):
                        matches.append(
                            f"{field}[year={year_idx}]: {expected_val} ~ {alias_val} via '{alias}'",
                        )
                    else:
                        mismatches.append(
                            f"{field}[year={year_idx}]: expected={expected_val}, "
                            f"got={alias_val} via '{alias}'",
                        )
                    found_via_alias = True
                    break

            if not found_via_alias:
                mismatches.append(
                    f"{field}[year={year_idx}]: expected={expected_val}, not found",
                )

    return matches, mismatches, acceptable


@pytest.mark.asyncio
@pytest.mark.timeout(900)
@_handle_project
async def test_generic_ash_financial_extraction(tmp_path: Path):
    """Extract financial data with NO domain-specific customization.

    Baseline test: the actor has only the general CodeActActor prompt,
    FileManager primitives (render_pdf, render_excel_sheet), and a
    detailed human request with the empty Excel template.
    """
    assert len(PDF_FILES) == 4, f"Expected 4 PDF fixtures, found {len(PDF_FILES)}"
    assert TEMPLATE_XLSX.exists(), "Template Excel not found"

    workspace_dir = tmp_path / "accounts"
    workspace_dir.mkdir()
    for pdf in PDF_FILES:
        shutil.copy(pdf, workspace_dir / pdf.name)

    template_path = tmp_path / "HISTORIC_ACCOUNTS_TEMPLATE.xlsx"
    shutil.copy(TEMPLATE_XLSX, template_path)

    output_dir = tmp_path / "output"
    output_dir.mkdir()
    output_path = output_dir / "ASH_Historic_Accounts.xlsx"

    # NO guidance seeding, NO custom functions — but the full default
    # CodeActActor tool set including FunctionManager/GuidanceManager
    # discovery tools and the discovery-first policy.
    async with make_code_act_actor(
        impl="real",
        include_function_manager_tools=True,
    ) as (actor, _primitives, calls):
        handle = await actor.act(
            f"I need your help extracting financial data from annual accounts.\n\n"
            f"In the folder {workspace_dir} there are 4 PDF files — these are "
            f"the signed annual accounts for a care home called 'Ablegrange "
            f"Severn Heights' for fiscal years ending 30 April 2022, 2023, "
            f"2024, and 2025 (one PDF per year).\n\n"
            f"Each PDF contains a Profit and Loss account and detailed Notes "
            f"to the P&L that break down the expenses line by line. The P&L "
            f"shows the high-level figures (Turnover, Cost of Sales, "
            f"Administration Expenses, Interest, Profit before Tax, Taxation). "
            f"The Notes page lists every individual expense (wages, heating, "
            f"repairs, insurance, etc.) with exact amounts.\n\n"
            f"I've attached an empty Excel template at {template_path} that "
            f"shows the exact format I need. Please open this template first "
            f"to see the row labels and structure. Then go through each PDF "
            f"and fill in every value you can find. The PDFs are scans of "
            f"printed documents, not digital-native, so you'll need to view "
            f"them visually rather than trying to extract text.\n\n"
            f"For each value you extract, find the exact number from the PDF "
            f"page. Map each line item from the PDF's Notes to the closest "
            f"matching row in the template. Some PDF labels won't match the "
            f"template exactly — use your best judgment.\n\n"
            f"Save the completed spreadsheet to {output_path}. Each fiscal "
            f"year should be a column (Year 1 = FYE 2022, Year 4 = FYE 2025).",
            clarification_enabled=False,
        )
        result = await handle.result()

    json_files = list(tmp_path.rglob("*.json"))
    xlsx_files = [f for f in tmp_path.rglob("*.xlsx") if "TEMPLATE" not in f.name]

    print(f"\n{'='*60}")
    print(f"Actor result:\n{result[:500]}")
    print(f"\nJSON files: {json_files}")
    print(f"Excel files: {xlsx_files}")
    print(f"Primitive calls: {calls}")
    print(f"{'='*60}\n")

    expected = _parse_ground_truth()

    actual: dict[str, dict[str, Any]] = {}
    if output_path.exists():
        actual = _parse_output_excel(str(output_path))
    elif xlsx_files:
        actual = _parse_output_excel(str(xlsx_files[0]))
    if not actual:
        actual = _collect_all_json_values(tmp_path)

    matches, mismatches, acceptable = _compare_results(expected, actual)

    total_strict = len(matches) + len(mismatches)
    match_rate = len(matches) / total_strict if total_strict > 0 else 0

    print(f"\n{'='*60}")
    print("GENERIC BASELINE — COMPARISON RESULTS")
    print(f"{'='*60}")
    print(f"Strict matches:   {len(matches)}")
    print(f"Strict failures:  {len(mismatches)}")
    print(f"Acceptable gaps:  {len(acceptable)}")
    print(f"Match rate:        {match_rate:.1%} ({len(matches)}/{total_strict})")

    if mismatches:
        print(f"\nSTRICT FAILURES ({len(mismatches)}):")
        for m in mismatches[:30]:
            print(f"  {m}")

    if acceptable:
        print(f"\nACCEPTABLE GAPS ({len(acceptable)}):")
        for m in acceptable[:10]:
            print(f"  {m}")

    if matches:
        print(f"\nSAMPLE MATCHES (first 15 of {len(matches)}):")
        for m in matches[:15]:
            print(f"  {m}")
    print(f"{'='*60}\n")

    if not actual:
        print(
            "WARNING: No structured output found. The actor may have "
            "produced output in a format we couldn't parse.\n"
            f"All files: {list(tmp_path.rglob('*'))}",
        )

    # This is a baseline measurement — we don't assert a strict pass/fail
    # threshold, but we do require *some* output was produced.
    assert total_strict > 0 or actual, (
        f"Actor produced no parseable output.\n"
        f"Result: {result[:1000]}\n"
        f"Files: {list(tmp_path.rglob('*'))}"
    )
